import hashlib
import secrets
import hmac

import streamlit as st
import streamlit.components.v1 as components
import sqlite3, json, os, re, tempfile, shutil, threading
from datetime import date, datetime, timedelta
from pathlib import Path
from io import BytesIO
from difflib import SequenceMatcher
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import html
from urllib.parse import urljoin
import textwrap
import requests

try:
    from pypdf import PdfReader
except Exception:
    PdfReader=None


try:
    from supabase import create_client
except Exception:
    create_client=None

st.set_page_config(page_title="ChapLab Teacher Hub", page_icon="📘", layout="wide", initial_sidebar_state="expanded")


st.markdown("""
<style>
section[data-testid="stSidebar"]{
    display:block !important;
    visibility:visible !important;
    opacity:1 !important;
    transform:none !important;
}
section[data-testid="stSidebar"] > div{
    display:block !important;
    visibility:visible !important;
}
</style>
""", unsafe_allow_html=True)


LEGACY_DB = Path("teacher_tracker.db")
WEB_DB = Path(tempfile.gettempdir()) / "chaplab_teacher_tracker.db"
UPLOAD_DIR = Path(tempfile.gettempdir()) / "chaplab_student_work"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
_CLOUD_LOCK = threading.Lock()

def _secret_section(name):
    try:
        return st.secrets[name]
    except Exception:
        return None

def cloud_config():
    cfg=_secret_section("supabase")
    if not cfg:
        return None
    try:
        url=str(cfg.get("url","")).strip().rstrip("/")
        # Accept either the project URL or an accidentally pasted REST endpoint.
        url=re.sub(r"/rest/v1/?$","",url,flags=re.I)
        key=str(cfg.get("service_role_key","")).strip()
        bucket=str(cfg.get("bucket","chaplab-private")).strip() or "chaplab-private"
        db_object=str(cfg.get("database_object","teacher_tracker.db")).strip() or "teacher_tracker.db"
    except Exception:
        return None
    if not url or not key:
        return None
    return {"url":url,"key":key,"bucket":bucket,"db_object":db_object}

def _stored_login_username(default_username):
    """Read the chosen ChapLab username before normal DB helpers initialize."""
    for db_path in (WEB_DB,LEGACY_DB):
        try:
            if not Path(db_path).exists():
                continue
            c=sqlite3.connect(str(db_path))
            table=c.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='settings'"
            ).fetchone()
            if not table:
                c.close()
                continue
            r=c.execute(
                "SELECT value FROM settings WHERE key='auth_effective_username'"
            ).fetchone()
            c.close()
            if r and r[0]:
                try:
                    value=json.loads(r[0])
                except Exception:
                    value=r[0]
                cleaned=str(value or "").strip()
                if cleaned:
                    return cleaned
        except Exception:
            continue
    return str(default_username or "").strip()

def auth_config():
    cfg=_secret_section("chaplab_auth")
    if not cfg:
        return None
    recovery_username=str(cfg.get("username","")).strip()
    password=str(cfg.get("password",""))
    if not recovery_username or not password:
        return None

    # The Streamlit Secret remains the recovery credential.
    # After the one-time first-login choice, the user's selected ChapLab
    # username is stored in the database and becomes the everyday login name.
    username=_stored_login_username(recovery_username)
    return {
        "username":username,
        "recovery_username":recovery_username,
        "password":password
    }

def cloud_configured():
    return cloud_config() is not None

def require_login():
    auth=auth_config()
    if cloud_configured() and not auth:
        st.error("ChapLab Web is connected to cloud storage, but Creator login credentials are missing.")
        st.info("Add a [chaplab_auth] section to Streamlit Secrets before using student data online.")
        st.stop()

    if not auth:
        return

    # Enforce a required staff password setup before normal app use.
    if st.session_state.get("chaplab_staff_password_setup"):
        email=str(st.session_state.get("chaplab_staff_email") or "").strip().lower()
        allowed,msg=staff_account_can_login(email)
        if not allowed:
            for k in ("chaplab_authenticated","chaplab_username","chaplab_staff_email","chaplab_staff_password_setup"):
                st.session_state.pop(k,None)
            st.error(msg)
            st.stop()

        st.markdown("## 🔐 Create Your ChapLab Password")
        st.write(f"Account: **{email}**")
        st.info(
            "Your initial access code worked. Before continuing, create your own ChapLab password. "
            "After you save it, the shared access code will no longer work for your account."
        )
        with st.form("staff_create_password_form"):
            p1=st.text_input("Create password",type="password")
            p2=st.text_input("Confirm password",type="password")
            save_password=st.form_submit_button("Save My Password",use_container_width=True,type="primary")

        if save_password:
            if len(str(p1 or ""))<8:
                st.error("Choose a password with at least 8 characters.")
            elif p1!=p2:
                st.error("The passwords do not match.")
            elif str(p1 or "").strip().lower()==INITIAL_STAFF_ACCESS_CODE:
                st.error("Choose a personal password instead of the initial access code.")
            else:
                set_staff_password(email,p1)
                st.session_state.pop("chaplab_staff_password_setup",None)
                st.session_state["chaplab_authenticated"]=True
                record_staff_login(email)
                st.success("Password created. Opening ChapLab…")
                st.rerun()
        st.stop()

    # Re-check active staff accounts on every rerun.
    if st.session_state.get("chaplab_authenticated"):
        staff_email=str(st.session_state.get("chaplab_staff_email") or "").strip().lower()
        if staff_email:
            allowed,msg=staff_account_can_login(staff_email)
            if not allowed:
                for k in ("chaplab_authenticated","chaplab_username","chaplab_staff_email","chaplab_staff_display_name"):
                    st.session_state.pop(k,None)
                st.error(msg)
                st.stop()
            # If Creator reset the account while user was signed in, require setup next.
            if staff_needs_password_setup(staff_email):
                st.session_state["chaplab_authenticated"]=False
                st.session_state["chaplab_staff_password_setup"]=True
                st.rerun()
            touch_staff_last_seen(staff_email)
        elif not st.session_state.get("chaplab_username"):
            st.session_state["chaplab_username"]=auth.get("username","teacher")
        return

    st.markdown("## 📘 ChapLab Teacher Hub")
    st.caption("Private school staff sign-in")
    st.caption(
        "Staff: use your **79.…@nhaschools.com** email. "
        "First-time approved users sign in with access code **bscs**."
    )

    with st.form("chaplab_login"):
        username=st.text_input("School email / Creator username")
        password=st.text_input("Password / Initial access code",type="password")
        submitted=st.form_submit_button("Sign In",use_container_width=True)

    if submitted:
        entered=str(username or "").strip()
        entered_lower=entered.lower()
        recovery_user=str(auth.get("recovery_username","")).strip()
        everyday_creator=str(auth.get("username","")).strip()

        # Creator login stays separate and uses Streamlit Secrets.
        creator_login=entered in {recovery_user,everyday_creator}
        if creator_login and password==auth["password"]:
            st.session_state["chaplab_authenticated"]=True
            st.session_state["chaplab_username"]=everyday_creator or entered
            st.session_state.pop("chaplab_staff_email",None)
            st.session_state["_just_logged_in"]=True
            st.rerun()

        if not valid_school_staff_email(entered_lower):
            st.error(
                "School staff must sign in with an NHA email that starts with **79.** "
                "and ends with **@nhaschools.com**."
            )
            st.stop()

        allowed,msg=staff_account_can_login(entered_lower)
        if not allowed:
            st.error(msg)
            st.stop()

        account=staff_account_by_email(entered_lower)

        # First login / Creator reset: shared access code is accepted case-insensitively.
        if staff_needs_password_setup(entered_lower):
            if initial_access_code_matches(password):
                st.session_state["chaplab_username"]=entered_lower
                st.session_state["chaplab_staff_email"]=entered_lower
                if account and account["display_name"]:
                    st.session_state["chaplab_staff_display_name"]=account["display_name"]
                st.session_state["chaplab_staff_password_setup"]=True
                st.rerun()
            else:
                st.error(
                    "This account still needs password setup. "
                    "Use the initial access code provided by the ChapLab App Creator & Administrator."
                )
            st.stop()

        # Normal returning staff login: individual password only.
        if verify_staff_password(account,password):
            st.session_state["chaplab_authenticated"]=True
            st.session_state["chaplab_username"]=entered_lower
            st.session_state["chaplab_staff_email"]=entered_lower
            if account and account["display_name"]:
                st.session_state["chaplab_staff_display_name"]=account["display_name"]
            record_staff_login(entered_lower)
            st.session_state["_just_logged_in"]=True
            st.rerun()
        else:
            st.error("Email or password is incorrect.")
    st.stop()

require_login()

# Lightweight diagnostic marker: if this appears in Streamlit logs,
# ChapLab passed imports/login setup and entered normal app execution.
print("ChapLab startup: core app loaded; barcode engine deferred until scan.", flush=True)

def _cloud_headers(content_type=None):
    cfg=cloud_config()
    if not cfg:
        return {}
    headers={
        "Authorization":f"Bearer {cfg['key']}",
        "apikey":cfg["key"],
    }
    if content_type:
        headers["Content-Type"]=content_type
    return headers

def _storage_url(path=""):
    cfg=cloud_config()
    if not cfg:
        return ""
    return f"{cfg['url']}/storage/v1/{path.lstrip('/')}"

@st.cache_resource
def ensure_cloud_bucket():
    """Ensure the private bucket exists, but never block app startup for long."""
    if not cloud_configured():
        return False
    cfg=cloud_config()
    try:
        # First check the bucket. A 200 means it already exists.
        check=requests.get(
            _storage_url(f"bucket/{cfg['bucket']}"),
            headers=_cloud_headers(),
            timeout=(3,5)
        )
        if check.status_code==200:
            return True

        # Otherwise try to create it. 400/409 may simply mean it exists.
        create=requests.post(
            _storage_url("bucket"),
            headers=_cloud_headers("application/json"),
            json={
                "id":cfg["bucket"],
                "name":cfg["bucket"],
                "public":False,
                "file_size_limit":52428800
            },
            timeout=(3,5)
        )
        if create.status_code in (200,201,400,409):
            return True
        st.session_state["_cloud_sync_error"]=f"Bucket check returned HTTP {create.status_code}."
        return False
    except Exception as e:
        st.session_state["_cloud_sync_error"]=f"Cloud bucket check timed out or failed: {e}"
        return False

def cloud_download_bytes(remote_path):
    """Download with a short timeout; failure falls back to the local DB."""
    if not cloud_configured():
        return None
    cfg=cloud_config()
    try:
        if not ensure_cloud_bucket():
            return None
        r=requests.get(
            _storage_url(f"object/{cfg['bucket']}/{remote_path.lstrip('/')}"),
            headers=_cloud_headers(),
            timeout=(3,8)
        )
        if r.status_code==200:
            st.session_state.pop("_cloud_sync_error",None)
            return r.content
        if r.status_code not in (400,404):
            st.session_state["_cloud_sync_error"]=f"Database download returned HTTP {r.status_code}."
        return None
    except Exception as e:
        st.session_state["_cloud_sync_error"]=f"Database download timed out or failed: {e}"
        return None

def cloud_upload_bytes(data, remote_path, content_type="application/octet-stream"):
    """Upload with a bounded timeout. It is safe for background use."""
    if not cloud_configured():
        return False
    cfg=cloud_config()
    try:
        if not ensure_cloud_bucket():
            return False
        headers=_cloud_headers(content_type)
        headers["x-upsert"]="true"
        r=requests.post(
            _storage_url(f"object/{cfg['bucket']}/{remote_path.lstrip('/')}"),
            headers=headers,
            data=data,
            timeout=(3,10)
        )
        if r.status_code in (200,201):
            return True
        # Some Storage versions prefer PUT for an upsert.
        r=requests.put(
            _storage_url(f"object/{cfg['bucket']}/{remote_path.lstrip('/')}"),
            headers=headers,
            data=data,
            timeout=(3,10)
        )
        return r.status_code in (200,201)
    except Exception:
        return False

def cloud_upload_file(local_path, remote_path=None):
    local_path=Path(local_path)
    if not local_path.exists() or not cloud_configured():
        return False
    cfg=cloud_config()
    remote_path=remote_path or cfg["db_object"]
    return cloud_upload_bytes(local_path.read_bytes(),remote_path,"application/x-sqlite3")

def _background_cloud_upload(local_path):
    """Snapshot the DB now, then sync without blocking a button click/rerun."""
    if not cloud_configured():
        return
    try:
        payload=Path(local_path).read_bytes()
        cfg=cloud_config()
        remote=cfg["db_object"]
    except Exception:
        return

    def worker():
        with _CLOUD_LOCK:
            cloud_upload_bytes(payload,remote,"application/x-sqlite3")
    threading.Thread(target=worker,daemon=True).start()

def cloud_download_database(target):
    if not cloud_configured():
        return False
    cfg=cloud_config()
    data=cloud_download_bytes(cfg["db_object"])
    if not data:
        return False
    target=Path(target)
    target.parent.mkdir(parents=True,exist_ok=True)
    target.write_bytes(data)
    return True

def prepare_database():
    """Open ChapLab immediately. Cloud is never contacted during startup."""
    if WEB_DB.exists():
        return str(WEB_DB)
    if LEGACY_DB.exists():
        try:
            shutil.copy2(LEGACY_DB,WEB_DB)
            return str(WEB_DB)
        except Exception:
            return str(LEGACY_DB)
    return str(WEB_DB)

DB = prepare_database()

def cloud_status_text():
    if not cloud_configured():
        return "Local mode"
    if st.session_state.get("_cloud_sync_error"):
        return "Cloud configured — app running locally; sync needs attention"
    return "Cloud configured"

# ---------- Composition notebook style ----------
st.markdown("""
<style>
:root{
  --navy:#123f8c;--navy-dark:#0d2f6a;--ink:#253044;--muted:#6c7482;
  --blue:#9fd8f3;--green:#b9df6f;--yellow:#f7d85c;--pink:#f29ab3;
  --purple:#bda0e5;--orange:#f4a64b;--line:#e4e8ef;
}
.stApp{
  background:linear-gradient(rgba(255,255,255,.97),rgba(255,255,255,.97)),
             repeating-linear-gradient(to bottom,#fff 0,#fff 31px,#e5edf5 32px,#fff 33px);
  color:var(--ink);
}
header[data-testid="stHeader"]{
  background:rgba(255,255,255,.94);backdrop-filter:blur(8px);border-bottom:1px solid #edf0f4;
}

.block-container{
  max-width:1450px!important;margin:0 auto!important;padding:120px 34px 52px 300px!important;
  background:transparent!important;border:none!important;box-shadow:none!important;
}
.teacher-sidebar{
  position:fixed;left:0;top:0;bottom:0;width:260px;
  background:linear-gradient(180deg,var(--navy),var(--navy-dark));color:white;z-index:10000;
  padding:28px 18px 24px;box-sizing:border-box;box-shadow:5px 0 18px rgba(0,0,0,.12);overflow-y:auto;
}
.teacher-avatar{
  width:76px;height:76px;border-radius:50%;margin:8px auto 10px;display:flex;align-items:center;
  justify-content:center;background:white;color:var(--navy);border:4px solid #d7b5ff;font-size:25px;font-weight:900;
}
.teacher-brand{
  text-align:center;font-family:"Comic Sans MS","Segoe Print",cursive;font-size:24px;font-weight:800;
  line-height:1.1;margin-bottom:24px;
}
.nav-link{
  display:flex;align-items:center;gap:10px;padding:12px 14px;margin:5px 0;border-radius:10px;
  text-decoration:none!important;color:white!important;font-weight:700;font-size:14px;
}
.nav-link:hover{background:rgba(255,255,255,.13)}
.nav-link.active{background:linear-gradient(90deg,#7669eb,#5f65e8)}
.sidebar-note{
  margin-top:24px;background:linear-gradient(135deg,#f4dbff,#d9cbff);color:#2b2d3f;
  border-radius:18px;padding:16px;font-size:13px;line-height:1.4;
}
.class-bar{
  position:fixed;left:260px;right:0;top:48px;height:62px;background:rgba(255,255,255,.97);
  z-index:9000;display:flex;align-items:flex-end;gap:10px;padding:0 34px;
  border-bottom:1px solid #e6e9ef;box-shadow:0 3px 10px rgba(0,0,0,.04);
}
.class-bar-label{align-self:center;font-weight:800;color:#536072;margin-right:4px}
.class-pill{
  min-width:110px;padding:12px 18px;text-align:center;text-decoration:none!important;color:#253044!important;
  font-family:"Comic Sans MS","Segoe Print",cursive;font-weight:800;border:1px solid rgba(0,0,0,.08);
  border-radius:12px 12px 0 0;box-shadow:0 -2px 5px rgba(0,0,0,.04);
}
.class-pill:nth-of-type(2){background:var(--blue)}
.class-pill:nth-of-type(3){background:var(--green)}
.class-pill:nth-of-type(4){background:#f7f4ed}
.class-pill:nth-of-type(5){background:#d5e99a}
.class-pill.active{box-shadow:inset 0 -4px 0 #5f65e8}
.page-title{
  font-family:"Comic Sans MS","Segoe Print",cursive;font-size:32px;font-weight:900;color:#26344b;margin:0 0 6px;
}
.page-subtitle{color:var(--muted);margin-bottom:22px}
.hero-card{
  background:radial-gradient(circle at 8% 15%,rgba(255,255,255,.8),transparent 30%),
             linear-gradient(135deg,#fffaf1,#fffdf9);
  border:1px solid #e7dfd3;border-radius:22px;padding:28px 30px;box-shadow:0 7px 20px rgba(43,51,70,.08);
  margin-bottom:24px;position:relative;overflow:hidden;
}
.hero-card:after{
  content:"✦   ♡   ☆";position:absolute;right:28px;top:18px;font-size:28px;color:#b399d7;
  font-family:"Comic Sans MS",cursive;
}
.hero-card h1{
  margin:0;font-family:"Comic Sans MS","Segoe Print",cursive;font-size:38px;color:#27344a;
}
.hero-card .accent{
  display:inline-block;margin-top:8px;background:#c5a8e6;padding:6px 18px 8px;border-radius:8px;
  font-family:"Comic Sans MS","Segoe Print",cursive;font-weight:800;
}
.sticky-grid{
  display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:20px;margin:18px 0 26px;
}
.sticky-card{
  min-height:145px;padding:18px;border-radius:6px;box-shadow:7px 8px 15px rgba(0,0,0,.11);
  display:flex;flex-direction:column;justify-content:center;text-align:center;
  font-family:"Comic Sans MS","Segoe Print",cursive;color:#252a33;
}
.sticky-card strong{display:block;font-size:28px;margin-bottom:4px}
.sticky-card small{font-family:Arial,sans-serif;font-size:12px;line-height:1.3}
.s-pink{background:var(--pink)}.s-orange{background:var(--orange)}.s-yellow{background:var(--yellow)}
.s-green{background:var(--green)}.s-blue{background:var(--blue)}.s-purple{background:var(--purple)}
div[data-testid="stForm"],div[data-testid="stDataFrame"],div[data-testid="stExpander"],div[data-testid="stMetric"]{
  background:rgba(255,255,255,.97)!important;border-radius:12px!important;
}
div[data-testid="stMetric"]{border:1px solid var(--line)!important;box-shadow:0 2px 8px rgba(0,0,0,.05);padding:12px 14px}
input,textarea,[data-baseweb="select"]>div{background:#f7f9fc!important}
.profile-cover{
  height:180px;max-width:1050px;margin:0 auto;background:linear-gradient(135deg,#789dbb,#d8e8f3);
  border-radius:18px 18px 0 0;
}
.profile-header{
  max-width:1050px;margin:0 auto 18px;background:white;border:1px solid #dce2ea;border-top:0;
  border-radius:0 0 18px 18px;padding:0 24px 20px;
}
.profile-avatar{
  width:110px;height:110px;border-radius:50%;background:#31536a;color:white;display:flex;align-items:center;
  justify-content:center;font-size:36px;font-weight:900;border:5px solid white;box-shadow:0 0 0 3px #1d2830;margin-top:-58px;
}
.profile-section{
  max-width:1050px;margin:14px auto;background:white;border:1px solid #e0e5ec;border-radius:14px;
  padding:18px 20px;box-shadow:0 2px 8px rgba(0,0,0,.04);
}
@media(max-width:1050px){
  .teacher-sidebar{width:220px}.class-bar{left:220px}.block-container{padding-left:250px!important}
  .sticky-grid{grid-template-columns:repeat(2,minmax(0,1fr))}
}
@media(max-width:780px){
  .teacher-sidebar{display:none}.class-bar{left:0}.block-container{padding:120px 16px 40px!important}
  .sticky-grid{grid-template-columns:1fr}
}
</style>
""", unsafe_allow_html=True)

# ---------- Database ----------
class ChapLabConnection(sqlite3.Connection):
    def commit(self):
        return super().commit()

def conn():
    c=sqlite3.connect(DB,factory=ChapLabConnection)
    c.row_factory=sqlite3.Row
    c.execute("PRAGMA foreign_keys=ON")
    return c

def cols(cur, table):
    return {r[1] for r in cur.execute(f"PRAGMA table_info({table})").fetchall()}

def addcol(cur, table, name, decl):
    table_exists=cur.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?",
        (table,)
    ).fetchone()
    if not table_exists:
        return
    if name not in cols(cur, table):
        cur.execute(f"ALTER TABLE {table} ADD COLUMN {name} {decl}")

def init_db():
    c=conn(); cur=c.cursor()
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS classes(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        class_name TEXT NOT NULL UNIQUE,
        subject_note TEXT DEFAULT '',
        active INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS scholars(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        first_name TEXT NOT NULL,
        last_name TEXT NOT NULL,
        class_name TEXT DEFAULT '',
        gender TEXT DEFAULT '',
        pronouns TEXT DEFAULT '',
        active INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS guardians(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scholar_id INTEGER NOT NULL,
        first_name TEXT DEFAULT '',
        last_name TEXT DEFAULT '',
        relationship TEXT DEFAULT '',
        home_phone TEXT DEFAULT '',
        work_phone TEXT DEFAULT '',
        cell_phone TEXT DEFAULT '',
        email TEXT DEFAULT '',
        UNIQUE(scholar_id, first_name, last_name, relationship)
    );
    CREATE TABLE IF NOT EXISTS standards(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        subject TEXT NOT NULL,
        code TEXT NOT NULL,
        skill TEXT NOT NULL,
        description TEXT DEFAULT '',
        UNIQUE(subject, code)
    );
    CREATE TABLE IF NOT EXISTS assignments(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        subject TEXT NOT NULL,
        category TEXT NOT NULL,
        standard_code TEXT DEFAULT '',
        points_possible REAL DEFAULT 100,
        assignment_date TEXT
    );
    CREATE TABLE IF NOT EXISTS grades(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scholar_id INTEGER NOT NULL,
        assignment_id INTEGER NOT NULL,
        points_earned REAL,
        UNIQUE(scholar_id, assignment_id)
    );
    CREATE TABLE IF NOT EXISTS bulletin_board_work(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scholar_id INTEGER,
        created_at TEXT NOT NULL,
        title TEXT DEFAULT '',
        subject TEXT DEFAULT 'ELA',
        task_text TEXT DEFAULT '',
        vocabulary TEXT DEFAULT '',
        standards TEXT DEFAULT '',
        rubric_text TEXT DEFAULT '',
        work_file_name TEXT DEFAULT '',
        work_file_path TEXT DEFAULT '',
        glow_1 TEXT DEFAULT '',
        glow_2 TEXT DEFAULT '',
        grow TEXT DEFAULT '',
        final_feedback TEXT DEFAULT '',
        grade_value REAL,
        points_possible REAL DEFAULT 100,
        include_in_gradebook INTEGER DEFAULT 0,
        assignment_id INTEGER
    );
    CREATE TABLE IF NOT EXISTS settings(key TEXT PRIMARY KEY,value TEXT);
    CREATE TABLE IF NOT EXISTS username_change_requests(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        requester_key TEXT NOT NULL,
        requester_name TEXT DEFAULT '',
        current_username TEXT DEFAULT '',
        requested_username TEXT NOT NULL,
        reason TEXT DEFAULT '',
        created_at TEXT NOT NULL,
        status TEXT DEFAULT 'Pending',
        reviewed_at TEXT DEFAULT '',
        reviewed_by TEXT DEFAULT '',
        review_note TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS newsletter_blurbs(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        author_key TEXT NOT NULL,
        author_name TEXT DEFAULT '',
        created_at TEXT NOT NULL,
        updated_at TEXT NOT NULL,
        academic_year TEXT DEFAULT '',
        newsletter_period TEXT DEFAULT '',
        blurb_type TEXT DEFAULT 'Subject',
        subject TEXT DEFAULT '',
        title TEXT DEFAULT '',
        recently_taught TEXT DEFAULT '',
        family_help TEXT DEFAULT '',
        coming_next TEXT DEFAULT '',
        prepare_next TEXT DEFAULT '',
        event_type TEXT DEFAULT '',
        event_details TEXT DEFAULT '',
        generated_blurb TEXT DEFAULT '',
        status TEXT DEFAULT 'Draft',
        submitted_at TEXT DEFAULT '',
        finalized_at TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS newsletter_blurb_versions(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        blurb_id INTEGER,
        author_key TEXT NOT NULL,
        author_name TEXT DEFAULT '',
        saved_at TEXT NOT NULL,
        action TEXT DEFAULT 'Saved',
        content_snapshot TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS newsletter_roles(
        author_key TEXT PRIMARY KEY,
        is_newsletter_lead INTEGER DEFAULT 0
    );
    CREATE TABLE IF NOT EXISTS newsletter_requests(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        created_by_key TEXT NOT NULL,
        created_by_name TEXT DEFAULT '',
        created_at TEXT NOT NULL,
        newsletter_period TEXT DEFAULT '',
        due_date TEXT DEFAULT '',
        audience_type TEXT DEFAULT 'Everyone',
        audience_value TEXT DEFAULT '',
        required_subjects TEXT DEFAULT '',
        instructions TEXT DEFAULT '',
        active INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS feature_rollout(
        feature_key TEXT PRIMARY KEY,
        enabled INTEGER DEFAULT 0,
        rollout_stage TEXT DEFAULT 'Creator Only',
        updated_at TEXT DEFAULT '',
        updated_by TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS staff_accounts(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        email TEXT NOT NULL UNIQUE,
        display_name TEXT DEFAULT '',
        role_type TEXT DEFAULT 'Teacher',
        grade_band TEXT DEFAULT '',
        subjects TEXT DEFAULT '',
        approval_status TEXT DEFAULT 'Pending',
        active INTEGER DEFAULT 1,
        created_at TEXT DEFAULT '',
        approved_at TEXT DEFAULT '',
        approved_by TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS staff_login_activity(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        staff_email TEXT NOT NULL,
        logged_in_at TEXT NOT NULL,
        event_type TEXT DEFAULT 'Login',
        details TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS staff_role_assignments(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        grade_band TEXT NOT NULL,
        role_name TEXT NOT NULL,
        staff_email TEXT NOT NULL,
        staff_name TEXT DEFAULT '',
        assigned_by TEXT DEFAULT '',
        assigned_at TEXT DEFAULT '',
        active INTEGER DEFAULT 1
    );
    CREATE TABLE IF NOT EXISTS communications(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scholar_id INTEGER, guardian_id INTEGER,
        created_at TEXT, communication_type TEXT,
        subject TEXT, reason TEXT, details TEXT, generated_text TEXT
    );
    CREATE TABLE IF NOT EXISTS work_samples(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scholar_id INTEGER NOT NULL, uploaded_at TEXT NOT NULL,
        subject TEXT DEFAULT '', skill_code TEXT DEFAULT '', title TEXT DEFAULT '',
        file_name TEXT DEFAULT '', file_path TEXT DEFAULT '',
        teacher_observation TEXT DEFAULT '', strengths TEXT DEFAULT '',
        needs TEXT DEFAULT '', next_steps TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS support_notes(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scholar_id INTEGER NOT NULL, created_at TEXT NOT NULL,
        note_type TEXT DEFAULT '', area TEXT DEFAULT '', observation TEXT DEFAULT '',
        frequency TEXT DEFAULT '', intervention TEXT DEFAULT '',
        response_to_intervention TEXT DEFAULT '', impact TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS report_comments(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scholar_id INTEGER NOT NULL, created_at TEXT NOT NULL,
        subject TEXT DEFAULT '', marking_period TEXT DEFAULT '',
        comment_text TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS contact_reminders(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scholar_id INTEGER NOT NULL,
        guardian_id INTEGER,
        due_date TEXT DEFAULT '',
        reason TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        completed INTEGER DEFAULT 0,
        completed_date TEXT DEFAULT ''
    );
    CREATE TABLE IF NOT EXISTS quarter_settings(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        academic_year TEXT DEFAULT '',
        quarter_name TEXT NOT NULL,
        start_date TEXT DEFAULT '',
        end_date TEXT DEFAULT '',
        locked INTEGER DEFAULT 0,
        UNIQUE(academic_year, quarter_name)
    );

    CREATE TABLE IF NOT EXISTS report_card_deadlines(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        academic_year TEXT DEFAULT '',
        quarter_name TEXT NOT NULL,
        due_date TEXT DEFAULT '',
        UNIQUE(academic_year, quarter_name)
    );

    CREATE TABLE IF NOT EXISTS parent_update_preferences(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scholar_id INTEGER NOT NULL UNIQUE,
        requested_updates INTEGER DEFAULT 0,
        update_frequency TEXT DEFAULT '',
        preferred_guardian_id INTEGER,
        notes TEXT DEFAULT ''
    );

    CREATE TABLE IF NOT EXISTS quarter_closeout(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scholar_id INTEGER NOT NULL,
        quarter_name TEXT NOT NULL,
        academic_year TEXT DEFAULT '',
        grades_reviewed INTEGER DEFAULT 0,
        data_reviewed INTEGER DEFAULT 0,
        comment_generated INTEGER DEFAULT 0,
        comment_finalized INTEGER DEFAULT 0,
        UNIQUE(scholar_id, quarter_name, academic_year)
    );
    CREATE TABLE IF NOT EXISTS benchmark_scores(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        scholar_id INTEGER NOT NULL UNIQUE,
        nwea_fall_reading TEXT DEFAULT '',
        nwea_spring_reading TEXT DEFAULT '',
        nwea_fall_math TEXT DEFAULT '',
        nwea_spring_math TEXT DEFAULT '',
        fp_fall_level TEXT DEFAULT '',
        fp_spring_level TEXT DEFAULT '',
        fp_fall_word_list TEXT DEFAULT '',
        fp_spring_word_list TEXT DEFAULT '',
        notes TEXT DEFAULT ''
    );

    CREATE TABLE IF NOT EXISTS book_levels(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT NOT NULL,
        author TEXT DEFAULT '',
        fp_level TEXT DEFAULT '',
        notes TEXT DEFAULT '',
        UNIQUE(title, author)
    );
    CREATE TABLE IF NOT EXISTS nwea_goals(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        class_id INTEGER,
        season TEXT NOT NULL,
        subject TEXT NOT NULL,
        goal_score REAL,
        UNIQUE(class_id, season, subject)
    );
    CREATE TABLE IF NOT EXISTS grouping_settings(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        class_id INTEGER,
        low_max REAL DEFAULT 69,
        mid_max REAL DEFAULT 84,
        UNIQUE(class_id)
    );
    """)
    addcol(cur,"scholars","class_id","INTEGER")
    addcol(cur,"scholars","school_name","TEXT DEFAULT ''")
    addcol(cur,"scholars","academic_year","TEXT DEFAULT ''")
    addcol(cur,"scholars","grade_level","TEXT DEFAULT ''")
    addcol(cur,"scholars","student_id","TEXT DEFAULT ''")
    addcol(cur,"scholars","address","TEXT DEFAULT ''")
    addcol(cur,"scholars","city","TEXT DEFAULT ''")
    addcol(cur,"scholars","state_code","TEXT DEFAULT ''")
    addcol(cur,"scholars","zip_code","TEXT DEFAULT ''")
    addcol(cur,"scholars","residency","TEXT DEFAULT ''")
    addcol(cur,"scholars","gender","TEXT DEFAULT ''")
    addcol(cur,"scholars","pronouns","TEXT DEFAULT ''")
    addcol(cur,"assignments","class_id","INTEGER")
    addcol(cur,"assignments","marking_period","TEXT DEFAULT ''")
    addcol(cur,"benchmark_scores","nwea_winter_reading","TEXT DEFAULT ''")
    addcol(cur,"benchmark_scores","nwea_winter_math","TEXT DEFAULT ''")
    addcol(cur,"benchmark_scores","nwea_reading_goal","TEXT DEFAULT ''")
    addcol(cur,"benchmark_scores","nwea_math_goal","TEXT DEFAULT ''")
    addcol(cur,"benchmark_scores","fp_winter_level","TEXT DEFAULT ''")
    addcol(cur,"benchmark_scores","fp_winter_word_list","TEXT DEFAULT ''")
    cur.execute("""
        CREATE TABLE IF NOT EXISTS iready_scores(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            scholar_id INTEGER NOT NULL UNIQUE,
            fall_reading TEXT DEFAULT '',
            winter_reading TEXT DEFAULT '',
            spring_reading TEXT DEFAULT '',
            reading_goal TEXT DEFAULT '',
            fall_math TEXT DEFAULT '',
            winter_math TEXT DEFAULT '',
            spring_math TEXT DEFAULT '',
            math_goal TEXT DEFAULT '',
            notes TEXT DEFAULT ''
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS interim_assessments(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            class_id INTEGER,
            subject TEXT NOT NULL,
            interim_number INTEGER NOT NULL,
            title TEXT NOT NULL,
            assessment_date TEXT DEFAULT '',
            source_file TEXT DEFAULT '',
            assignment_id INTEGER,
            imported_at TEXT NOT NULL,
            UNIQUE(class_id,subject,interim_number)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS interim_results(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            assessment_id INTEGER NOT NULL,
            scholar_id INTEGER NOT NULL,
            school_student_id TEXT DEFAULT '',
            overall_score REAL,
            mc_earned REAL, mc_possible REAL,
            cr_earned REAL, cr_possible REAL,
            raw_text TEXT DEFAULT '',
            UNIQUE(assessment_id,scholar_id)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS interim_standard_scores(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            assessment_id INTEGER NOT NULL,
            scholar_id INTEGER NOT NULL,
            standard_code TEXT NOT NULL,
            standard_text TEXT DEFAULT '',
            question_numbers TEXT DEFAULT '',
            standard_score REAL,
            UNIQUE(assessment_id,scholar_id,standard_code)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS interim_question_results(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            assessment_id INTEGER NOT NULL,
            scholar_id INTEGER NOT NULL,
            question_number INTEGER NOT NULL,
            response TEXT DEFAULT '',
            correct_answer TEXT DEFAULT '',
            earned REAL, possible REAL,
            UNIQUE(assessment_id,scholar_id,question_number)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS interim_goals(
            class_id INTEGER NOT NULL,
            subject TEXT NOT NULL,
            interim_number INTEGER NOT NULL,
            proficiency_goal REAL,
            notes TEXT DEFAULT '',
            PRIMARY KEY(class_id,subject,interim_number)
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS book_catalog(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            author TEXT DEFAULT '',
            fp_level TEXT DEFAULT '',
            isbn TEXT DEFAULT '',
            notes TEXT DEFAULT ''
        )
    """)
    addcol(cur,"book_catalog","author","TEXT DEFAULT ''")
    addcol(cur,"book_catalog","fp_level","TEXT DEFAULT ''")
    addcol(cur,"book_catalog","isbn","TEXT DEFAULT ''")
    addcol(cur,"book_catalog","notes","TEXT DEFAULT ''")
    addcol(cur,"classes","is_demo","INTEGER DEFAULT 0")
    addcol(cur,"scholars","is_demo","INTEGER DEFAULT 0")
    addcol(cur,"assignments","is_demo","INTEGER DEFAULT 0")
    addcol(cur,"assignments","include_in_average","INTEGER DEFAULT 1")
    addcol(cur,"communications","guardian_id","INTEGER")
    addcol(cur,"support_notes","concern_category","TEXT DEFAULT ''")
    addcol(cur,"support_notes","frequency_choice","TEXT DEFAULT ''")
    addcol(cur,"support_notes","intervention_choice","TEXT DEFAULT ''")
    addcol(cur,"support_notes","response_choice","TEXT DEFAULT ''")
    addcol(cur,"support_notes","impact_choice","TEXT DEFAULT ''")

    # migrate old class_name values
    for r in cur.execute("""SELECT DISTINCT TRIM(class_name) n FROM scholars
                            WHERE class_name IS NOT NULL AND TRIM(class_name)<>''""").fetchall():
        cur.execute("INSERT OR IGNORE INTO classes(class_name) VALUES (?)",(r["n"],))
    cur.execute("""UPDATE scholars SET class_id=(SELECT id FROM classes WHERE class_name=scholars.class_name)
                   WHERE class_id IS NULL AND TRIM(COALESCE(class_name,''))<>''""")

    defaults={
        "weights":{"Assessment":40,"Classwork":30,"Project/Lab":20,"Homework":10},
        "scale":[["A+",97,100],["A",93,96.99],["A-",90,92.99],
                 ["B+",87,89.99],["B",83,86.99],["B-",80,82.99],
                 ["C+",77,79.99],["C",73,76.99],["C-",70,72.99],
                 ["D",65,69.99],["F",0,64.99]]
    }
    for k,v in defaults.items():
        cur.execute("INSERT OR IGNORE INTO settings(key,value) VALUES (?,?)",(k,json.dumps(v)))
    c.commit(); c.close()

def get_setting(k):
    c=conn(); r=c.execute("SELECT value FROM settings WHERE key=?",(k,)).fetchone(); c.close()
    return json.loads(r["value"]) if r else None

def save_setting(k,v):
    c=conn(); c.execute("INSERT OR REPLACE INTO settings(key,value) VALUES (?,?)",(k,json.dumps(v))); c.commit(); c.close()

def classes_df(active=True):
    c=conn(); q="SELECT * FROM classes"+(" WHERE active=1" if active else "")+" ORDER BY class_name"
    df=pd.read_sql_query(q,c); c.close(); return df

def scholars_df(class_id=None, active=True):
    c=conn()
    q="""SELECT s.*, COALESCE(c.class_name,s.class_name,'') display_class
         FROM scholars s LEFT JOIN classes c ON c.id=s.class_id WHERE 1=1"""
    p=[]
    if active: q+=" AND s.active=1"
    if class_id: q+=" AND s.class_id=?"; p.append(int(class_id))
    q+=" ORDER BY s.last_name,s.first_name"
    df=pd.read_sql_query(q,c,params=p); c.close(); return df

def standards_df(subject=None):
    c=conn(); q="SELECT * FROM standards"; p=[]
    if subject: q+=" WHERE subject=?"; p=[subject]
    q+=" ORDER BY subject,code"
    df=pd.read_sql_query(q,c,params=p); c.close(); return df

def seed():
    data=[
      ("Science","3-PS2-1","Balanced and unbalanced forces","Investigate effects of balanced and unbalanced forces on motion."),
      ("Science","3-PS2-2","Patterns of motion","Use patterns of motion to make predictions."),
      ("Science","3-PS2-3","Electric and magnetic interactions","Determine cause-and-effect relationships of electric or magnetic interactions."),
      ("Science","3-PS2-4","Magnetic design problem","Apply scientific ideas about magnets to solve a problem."),
      ("Science","3-LS1-1","Life cycles","Develop models to describe unique and diverse life cycles."),
      ("Science","3-LS2-1","Social and group behavior","Construct an argument that some animals form groups that help members survive."),
      ("Science","3-LS3-1","Inheritance and variation of traits","Analyze data about inherited traits and variation."),
      ("Science","3-LS3-2","Environmental effects on traits","Use evidence that the environment can influence traits."),
      ("Science","3-LS4-1","Plant and animal extinction","Analyze fossil evidence of organisms and environments from long ago."),
      ("Science","3-LS4-2","Variation and survival","Explain how variation may provide survival advantages."),
      ("Science","3-LS4-3","Habitat and survival","Use evidence about how organisms survive in habitats."),
      ("Science","3-LS4-4","Environmental change solutions","Make a claim about the merit of a solution to environmental change."),
      ("Social Studies","SS3-A1","Develop questions","Develop questions about communities."),
      ("Social Studies","SS3-A2","Use evidence","Recognize and use primary and secondary source evidence."),
      ("Social Studies","SS3-A5","Make inferences","Identify inferences from evidence."),
      ("Social Studies","SS3-B1","Chronological reasoning","Explain relationships among events."),
      ("Social Studies","SS3-B3","Cause and effect","Identify causes and effects."),
      ("Social Studies","SS3-C1","Comparison and context","Compare communities."),
      ("Social Studies","SS3-D1","Geographic reasoning","Ask geographic questions."),
      ("Social Studies","SS3-D2","Maps and geographic tools","Use maps and geographic representations."),
      ("Social Studies","SS3-E1","Economic systems","Identify examples of scarcity, resources, production, and consumption."),
      ("Social Studies","SS3-F1","Civic participation","Participate respectfully in civic/classroom life."),
      ("Social Studies","SS3-COMM","Communities","Describe characteristics and organization of communities."),
      ("Social Studies","SS3-REG","Regions and environment","Explain how geography, climate, resources, and environment influence communities."),
      ("ELA","3R1","Questions and evidence","Develop and answer questions using relevant text evidence."),
      ("ELA","3R2","Theme or central idea","Determine theme/central idea and supporting details."),
      ("ELA","3R3","Characters, events, ideas","Describe and explain development of characters, events, or ideas."),
      ("ELA","3R4","Word meaning","Determine meaning of words and phrases."),
      ("ELA","3R5","Text structure","Explain how parts of a text build on one another."),
      ("ELA","3R6","Point of view / purpose","Discuss point of view and author's purpose."),
      ("ELA","3R7","Illustrations and text features","Explain how illustrations or text features contribute to meaning."),
      ("ELA","3W2","Informative writing","Write informative/explanatory texts."),
      ("ELA","3W3","Narrative writing","Write narratives with details and clear sequence."),
      ("ELA","3W5","Plan, revise, edit","Strengthen writing through planning, revision, and editing."),
      ("ELA","3SL1","Collaborative discussions","Participate effectively in collaborative discussions."),
      ("ELA","3L1","Grammar and usage","Use standard English grammar and usage."),
      ("ELA","3L2","Capitalization, punctuation, spelling","Use capitalization, punctuation, and spelling conventions."),
      ("Math","NY-3.OA.1","Interpret multiplication","Interpret products of whole numbers."),
      ("Math","NY-3.OA.2","Interpret division","Interpret whole-number quotients."),
      ("Math","NY-3.OA.3","Multiplication/division word problems","Solve multiplication and division word problems."),
      ("Math","NY-3.OA.4","Unknown-factor problems","Determine unknowns in multiplication/division equations."),
      ("Math","NY-3.OA.7","Multiply and divide within 100","Fluently multiply and divide within 100."),
      ("Math","NY-3.NBT.1","Rounding","Round whole numbers to nearest 10 or 100."),
      ("Math","NY-3.NBT.2","Add and subtract within 1,000","Fluently add and subtract within 1,000."),
      ("Math","NY-3.NF.1","Understand fractions","Understand fractions as quantities formed by unit fractions."),
      ("Math","NY-3.NF.2","Fractions on a number line","Represent fractions on a number line."),
      ("Math","NY-3.NF.3","Equivalent/comparable fractions","Explain equivalence and compare fractions."),
      ("Math","NY-3.MD.1","Time intervals","Tell/write time and solve interval problems."),
      ("Math","NY-3.MD.3","Graphs","Draw and interpret scaled graphs."),
      ("Math","NY-3.MD.5","Area","Understand area measurement."),
      ("Math","NY-3.MD.8","Perimeter","Solve perimeter problems."),
      ("Math","NY-3.G.1","Classify shapes","Classify shapes by attributes."),
      ("Math","NY-3.G.2","Partition shapes","Partition shapes into equal areas.")
    ]
    c=conn()
    for r in data:
        c.execute("INSERT OR IGNORE INTO standards(subject,code,skill,description) VALUES (?,?,?,?)",r)
    c.commit(); c.close()


init_db(); seed()

# ---------- Staff Account Security / Activity ----------
def valid_school_staff_email(email):
    """Only school staff emails beginning with 79. and ending @nhaschools.com are eligible."""
    email=str(email or "").strip().lower()
    return bool(re.fullmatch(r"79\.[a-z0-9._%+\-]+@nhaschools\.com",email))

def ensure_staff_account_columns():
    c=conn()
    cols={r["name"] for r in c.execute("PRAGMA table_info(staff_accounts)").fetchall()}
    migrations={
        "last_login":"TEXT DEFAULT ''",
        "last_seen":"TEXT DEFAULT ''",
        "login_count":"INTEGER DEFAULT 0",
        "deactivated_at":"TEXT DEFAULT ''",
        "deactivated_by":"TEXT DEFAULT ''",
        "deactivation_reason":"TEXT DEFAULT ''",
        "password_hash":"TEXT DEFAULT ''",
        "password_salt":"TEXT DEFAULT ''",
        "password_set_at":"TEXT DEFAULT ''",
        "password_setup_required":"INTEGER DEFAULT 1",
        "password_reset_at":"TEXT DEFAULT ''",
        "password_reset_by":"TEXT DEFAULT ''",
    }
    for name,definition in migrations.items():
        if name not in cols:
            c.execute(f"ALTER TABLE staff_accounts ADD COLUMN {name} {definition}")
    c.commit(); c.close()

def staff_account_by_email(email):
    email=str(email or "").strip().lower()
    c=conn()
    r=c.execute(
        "SELECT * FROM staff_accounts WHERE lower(email)=lower(?) LIMIT 1",
        (email,)
    ).fetchone()
    c.close()
    return r

def staff_account_can_login(email):
    email=str(email or "").strip().lower()
    if not valid_school_staff_email(email):
        return False,"School staff email must start with 79. and end with @nhaschools.com."
    r=staff_account_by_email(email)
    if not r:
        return False,"This staff email is not registered in ChapLab yet."
    if str(r["approval_status"] or "").strip()!="Approved":
        return False,"This account is waiting for approval."
    if not bool(r["active"]):
        return False,"This ChapLab account has been deactivated."
    return True,""

def record_staff_login(email):
    email=str(email or "").strip().lower()
    if not email:
        return
    now=datetime.now().isoformat(timespec="seconds")
    c=conn()
    c.execute(
        """UPDATE staff_accounts
           SET last_login=?,last_seen=?,login_count=COALESCE(login_count,0)+1
           WHERE lower(email)=lower(?)""",
        (now,now,email)
    )
    c.execute(
        """INSERT INTO staff_login_activity(staff_email,logged_in_at,event_type,details)
           VALUES (?,?,?,?)""",
        (email,now,"Login","Successful ChapLab sign-in")
    )
    c.commit(); c.close()

def touch_staff_last_seen(email):
    email=str(email or "").strip().lower()
    if not email:
        return
    c=conn()
    c.execute(
        "UPDATE staff_accounts SET last_seen=? WHERE lower(email)=lower(?)",
        (datetime.now().isoformat(timespec="seconds"),email)
    )
    c.commit(); c.close()

def all_staff_accounts_df():
    c=conn()
    df=pd.read_sql_query(
        """SELECT id,email,display_name,role_type,grade_band,subjects,
                  approval_status,active,created_at,approved_at,approved_by,
                  COALESCE(last_login,'') last_login,
                  COALESCE(last_seen,'') last_seen,
                  COALESCE(login_count,0) login_count,
                  COALESCE(deactivated_at,'') deactivated_at,
                  COALESCE(deactivated_by,'') deactivated_by,
                  COALESCE(deactivation_reason,'') deactivation_reason,
                  COALESCE(password_set_at,'') password_set_at,
                  COALESCE(password_setup_required,1) password_setup_required,
                  COALESCE(password_reset_at,'') password_reset_at
           FROM staff_accounts
           ORDER BY active DESC,approval_status,display_name,email""",
        c
    )
    c.close()
    return df

def set_staff_account_active(email,active,reason=""):
    email=str(email or "").strip().lower()
    c=conn()
    if active:
        c.execute(
            """UPDATE staff_accounts
               SET active=1,deactivated_at='',deactivated_by='',deactivation_reason=''
               WHERE lower(email)=lower(?)""",
            (email,)
        )
    else:
        now=datetime.now().isoformat(timespec="seconds")
        c.execute(
            """UPDATE staff_accounts
               SET active=0,deactivated_at=?,deactivated_by=?,deactivation_reason=?
               WHERE lower(email)=lower(?)""",
            (now,"ChapLab App Creator & Administrator",str(reason or "").strip(),email)
        )
        c.execute(
            """INSERT INTO staff_login_activity(staff_email,logged_in_at,event_type,details)
               VALUES (?,?,?,?)""",
            (email,now,"Deactivated",str(reason or "").strip())
        )
    c.commit(); c.close()

def set_staff_approval(email,status):
    email=str(email or "").strip().lower()
    if status=="Approved" and not valid_school_staff_email(email):
        return False
    now=datetime.now().isoformat(timespec="seconds")
    c=conn()
    c.execute(
        """UPDATE staff_accounts SET approval_status=?,
           approved_at=CASE WHEN ?='Approved' THEN ? ELSE approved_at END,
           approved_by=CASE WHEN ?='Approved' THEN 'ChapLab App Creator & Administrator' ELSE approved_by END
           WHERE lower(email)=lower(?)""",
        (status,status,now,status,email)
    )
    c.commit(); c.close()
    return True

ensure_staff_account_columns()

INITIAL_STAFF_ACCESS_CODE="bscs"

def _password_hash(password,salt_hex=None):
    """
    PBKDF2-HMAC-SHA256 password hashing using stdlib only.
    Returns (salt_hex, digest_hex).
    """
    password=str(password or "")
    if salt_hex:
        salt=bytes.fromhex(salt_hex)
    else:
        salt=secrets.token_bytes(16)
        salt_hex=salt.hex()
    digest=hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt,
        240000
    )
    return salt_hex,digest.hex()

def staff_password_is_set(account):
    if not account:
        return False
    try:
        return bool(str(account["password_hash"] or "").strip()) and not bool(account["password_setup_required"])
    except Exception:
        return False

def verify_staff_password(account,password):
    if not account or not staff_password_is_set(account):
        return False
    try:
        salt=str(account["password_salt"] or "")
        expected=str(account["password_hash"] or "")
        _,actual=_password_hash(str(password or ""),salt)
        return hmac.compare_digest(actual,expected)
    except Exception:
        return False

def set_staff_password(email,new_password):
    email=str(email or "").strip().lower()
    salt,digest=_password_hash(new_password)
    now=datetime.now().isoformat(timespec="seconds")
    c=conn()
    c.execute(
        """UPDATE staff_accounts
           SET password_hash=?,password_salt=?,password_set_at=?,
               password_setup_required=0,password_reset_at='',password_reset_by=''
           WHERE lower(email)=lower(?)""",
        (digest,salt,now,email)
    )
    c.commit(); c.close()

def reset_staff_password_setup(email):
    """
    Creator reset: removes the staff password and returns the account to the
    one-time access-code setup flow.
    """
    email=str(email or "").strip().lower()
    now=datetime.now().isoformat(timespec="seconds")
    c=conn()
    c.execute(
        """UPDATE staff_accounts
           SET password_hash='',password_salt='',password_set_at='',
               password_setup_required=1,password_reset_at=?,
               password_reset_by='ChapLab App Creator & Administrator'
           WHERE lower(email)=lower(?)""",
        (now,email)
    )
    c.execute(
        """INSERT INTO staff_login_activity(staff_email,logged_in_at,event_type,details)
           VALUES (?,?,?,?)""",
        (email,now,"Password Reset","Returned to initial access-code setup")
    )
    c.commit(); c.close()

def staff_needs_password_setup(email):
    r=staff_account_by_email(email)
    if not r:
        return False
    try:
        return bool(r["password_setup_required"]) or not bool(str(r["password_hash"] or "").strip())
    except Exception:
        return True

def initial_access_code_matches(value):
    return str(value or "").strip().lower()==INITIAL_STAFF_ACCESS_CODE

# ---------- Initial Grade 3 Pilot Team ----------
GRADE3_PILOT_TEACHERS=[
    {
        "email":"79.jschroeder@nhaschools.com",
        "display_name":"Jonathan Schroeder",
        "role_type":"Teacher",
        "grade_band":"Grade 3",
        "subjects":[]
    },
    {
        "email":"79.ncampbell@nhaschools.com",
        "display_name":"Nicole Campbell",
        "role_type":"Teacher",
        "grade_band":"Grade 3",
        "subjects":[]
    },
    {
        "email":"79.adavidson@nhaschools.com",
        "display_name":"Ania Davidson",
        "role_type":"Teacher",
        "grade_band":"Grade 3",
        "subjects":[]
    },
]

def ensure_grade3_pilot_team():
    """
    Pre-approve the initial Grade 3 pilot teachers.
    This does not assign Grade Team Leader or Newsletter Lead.
    Subjects remain blank until each teacher completes/edits their profile.
    """
    c=conn()
    now=datetime.now().isoformat(timespec="minutes")
    for teacher in GRADE3_PILOT_TEACHERS:
        email=teacher["email"].strip().lower()
        if not valid_school_staff_email(email):
            continue
        existing=c.execute(
            "SELECT id FROM staff_accounts WHERE lower(email)=lower(?) LIMIT 1",
            (email,)
        ).fetchone()

        values=(
            teacher["display_name"],
            teacher["role_type"],
            teacher["grade_band"],
            json.dumps(teacher["subjects"]),
            "Approved",
            1,
            now,
            "ChapLab App Creator & Administrator",
        )

        if existing:
            c.execute(
                """UPDATE staff_accounts
                   SET display_name=?,role_type=?,grade_band=?,subjects=?,
                       approval_status=?,active=?,approved_at=?,
                       approved_by=?
                   WHERE id=?""",
                values+(int(existing["id"]),)
            )
        else:
            c.execute(
                """INSERT INTO staff_accounts(
                   email,display_name,role_type,grade_band,subjects,
                   approval_status,active,created_at,approved_at,approved_by)
                   VALUES (?,?,?,?,?,'Approved',1,?,?,?)""",
                (
                    email,
                    teacher["display_name"],
                    teacher["role_type"],
                    teacher["grade_band"],
                    json.dumps(teacher["subjects"]),
                    now,
                    now,
                    "ChapLab App Creator & Administrator",
                )
            )

    # Move staff-account testing to the Grade Team stage without releasing
    # unfinished Dean or wider school features.
    c.execute(
        """INSERT OR REPLACE INTO feature_rollout(
           feature_key,enabled,rollout_stage,updated_at,updated_by)
           VALUES ('self_signup',1,'Creator + Grade Team',?,?)""",
        (now,"ChapLab App Creator & Administrator")
    )
    c.commit()
    c.close()

ensure_grade3_pilot_team()


# ---------- Username Setup / Change Requests ----------
def _username_clean(value):
    return str(value or "").strip()

def _username_valid(value):
    value=_username_clean(value)
    # Simple login-friendly usernames: no spaces, 3–30 characters.
    return bool(re.fullmatch(r"[A-Za-z0-9._-]{3,30}",value))

def auth_username_state():
    c=conn()
    rows=c.execute(
        """SELECT key,value FROM settings
           WHERE key IN ('auth_effective_username','auth_initial_username_choice_used')"""
    ).fetchall()
    c.close()
    out={
        "username":str((auth_config() or {}).get("username","")).strip(),
        "initial_choice_used":False
    }
    for r in rows:
        try:
            v=json.loads(r["value"])
        except Exception:
            v=r["value"]
        if r["key"]=="auth_effective_username":
            out["username"]=_username_clean(v) or out["username"]
        elif r["key"]=="auth_initial_username_choice_used":
            out["initial_choice_used"]=bool(v)
    return out

def save_initial_username_choice(username):
    username=_username_clean(username)
    c=conn()
    c.execute(
        "INSERT OR REPLACE INTO settings(key,value) VALUES ('auth_effective_username',?)",
        (json.dumps(username),)
    )
    c.execute(
        "INSERT OR REPLACE INTO settings(key,value) VALUES ('auth_initial_username_choice_used',?)",
        (json.dumps(True),)
    )
    c.commit(); c.close()
    st.session_state["chaplab_username"]=username

def submit_username_change_request(requested_username,reason=""):
    current=auth_username_state()["username"]
    requested=_username_clean(requested_username)
    c=conn()
    existing=c.execute(
        """SELECT id FROM username_change_requests
           WHERE requester_key=? AND status='Pending'""",
        (str(st.session_state.get("chaplab_username") or current).lower(),)
    ).fetchone()
    if existing:
        c.execute(
            """UPDATE username_change_requests
               SET requested_username=?,reason=?,created_at=? WHERE id=?""",
            (requested,_username_clean(reason),datetime.now().isoformat(timespec="minutes"),int(existing["id"]))
        )
    else:
        c.execute(
            """INSERT INTO username_change_requests(
               requester_key,requester_name,current_username,requested_username,
               reason,created_at,status)
               VALUES (?,?,?,?,?,?,'Pending')""",
            (str(st.session_state.get("chaplab_username") or current).lower(),
             current_author_name() if "current_author_name" in globals() else "Teacher",
             current,requested,_username_clean(reason),datetime.now().isoformat(timespec="minutes"))
        )
    c.commit(); c.close()

def pending_username_requests():
    c=conn()
    df=pd.read_sql_query(
        """SELECT * FROM username_change_requests
           WHERE status='Pending' ORDER BY id DESC""",c
    )
    c.close()
    return df

def review_username_request(request_id,approve,review_note=""):
    c=conn()
    r=c.execute("SELECT * FROM username_change_requests WHERE id=?",(int(request_id),)).fetchone()
    if not r:
        c.close()
        return False
    status="Approved" if approve else "Denied"
    now=datetime.now().isoformat(timespec="minutes")
    c.execute(
        """UPDATE username_change_requests
           SET status=?,reviewed_at=?,reviewed_by=?,review_note=? WHERE id=?""",
        (status,now,current_author_name() if "current_author_name" in globals() else "ChapLab App Creator & Administrator",
         _username_clean(review_note),int(request_id))
    )
    if approve:
        # Current single-account build has one effective login username.
        # Future multi-account rollout will map this same approval workflow
        # to the individual staff account record instead.
        c.execute(
            "INSERT OR REPLACE INTO settings(key,value) VALUES ('auth_effective_username',?)",
            (json.dumps(_username_clean(r["requested_username"])),)
        )
    c.commit(); c.close()
    return True

def enforce_initial_username_choice():
    auth=auth_config()
    if not auth or not st.session_state.get("chaplab_authenticated"):
        return

    state=auth_username_state()
    if state["initial_choice_used"]:
        return

    st.markdown("## 👋 One-Time Username Choice")
    st.info(
        "You may change your username one time from Profile Settings. "
        "After this screen is completed, future username changes require Creator/Admin approval."
    )
    current=state["username"] or str(auth.get("username","")).strip()
    st.write(f"Current username: **{current}**")

    choice=st.radio(
        "What would you like to do?",
        ["Keep my current username","Change my username now"],
        key="initial_username_choice"
    )

    if choice=="Change my username now":
        proposed=st.text_input(
            "Choose your new username",
            value="",
            placeholder="3–30 characters: letters, numbers, . _ -",
            key="initial_new_username"
        )
        cleaned=_username_clean(proposed)
        if proposed and not _username_valid(cleaned):
            st.warning("Use 3–30 characters with no spaces. Letters, numbers, periods, underscores, and hyphens are allowed.")
        if st.button("Save New Username",type="primary",use_container_width=True,key="save_initial_username"):
            if not _username_valid(cleaned):
                st.error("Enter a valid username first.")
            else:
                save_initial_username_choice(cleaned)
                st.success(f"Username changed to {cleaned}.")
                st.rerun()
    else:
        if st.button("Keep Current Username",type="primary",use_container_width=True,key="keep_initial_username"):
            save_initial_username_choice(current)
            st.success("Current username kept.")
            st.rerun()

    st.stop()


# ---------- Helpers ----------
def nm(row): return f"{row['first_name']} {row['last_name']}"

def pronoun_set_from_text(value):
    """Return subject/object/possessive/reflexive pronouns from a saved selection."""
    raw=str(value or "").strip().lower()
    if raw in {"she/her","she / her","female","girl"}:
        return {"subject":"she","object":"her","possessive":"her","possessive_pronoun":"hers","reflexive":"herself"}
    if raw in {"he/him","he / him","male","boy"}:
        return {"subject":"he","object":"him","possessive":"his","possessive_pronoun":"his","reflexive":"himself"}
    if raw in {"they/them","they / them","nonbinary","non-binary"}:
        return {"subject":"they","object":"them","possessive":"their","possessive_pronoun":"theirs","reflexive":"themself"}

    # Accept custom entries written as subject/object/possessive, e.g. ze/hir/hir.
    parts=[p.strip() for p in re.split(r"[/,]",raw) if p.strip()]
    if len(parts)>=2:
        subj=parts[0]
        obj=parts[1]
        poss=parts[2] if len(parts)>=3 else obj
        return {"subject":subj,"object":obj,"possessive":poss,"possessive_pronoun":poss,"reflexive":f"{obj}self"}

    # Neutral default when nothing is entered.
    return {"subject":"they","object":"them","possessive":"their","possessive_pronoun":"theirs","reflexive":"themself"}

def scholar_pronouns(sid):
    c=conn()
    r=c.execute("SELECT gender,pronouns FROM scholars WHERE id=?",(int(sid),)).fetchone()
    c.close()
    if not r:
        return pronoun_set_from_text("")
    return pronoun_set_from_text(r["pronouns"] or r["gender"] or "")

def cap_pronoun(p):
    return p[:1].upper()+p[1:] if p else p
def clean(v):
    if pd.isna(v): return ""
    s=str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit(): s=s[:-2]
    return s

def norm_header(h):
    return re.sub(r'[^a-z0-9]','',str(h).lower())

ALIASES={
 "school_name":["schoolname","school"],
 "academic_year":["academicyear","schoolyear","academicy"],
 "grade_level":["gradelevel","academicgradelevel","grade"],
 "class_name":["course","coursesection","section","class","classname","studentla"],
 "student_id":["studentid","studentnumber","localid","sisid"],
 "student_first":["studentfirstname","studentfirst","studentfir","scholarfirstname","scholarfirst"],
 "student_last":["studentlastname","studentlast","studentla","scholarlastname","scholarlast"],
 "address":["address","streetaddress"],
 "city":["city"],
 "state_code":["statecode","state"],
 "zip_code":["zipcode","zip","postalcode"],
 "residency":["residency"],
 "guardian_first":["firstname","parentfirstname","guardianfirstname"],
 "guardian_last":["lastname","parentlastname","guardianlastname"],
 "relationship":["relationshiptypename","relationship","relation"],
 "home_phone":["homephone","homepho"],
 "work_phone":["workphone","workpho","workphor"],
 "cell_phone":["cellphone","mobilephone","cell"],
 "email":["email","emailaddress"]
}

def auto_mapping(columns):
    normalized={norm_header(c):c for c in columns}
    mapping={}
    for field,als in ALIASES.items():
        for a in als:
            if norm_header(a) in normalized:
                mapping[field]=normalized[norm_header(a)]
                break
    # Screenshot-style special handling:
    # Student first/last headers often start "StudentFir..." and "StudentLa..."
    for c in columns:
        n=norm_header(c)
        if "studentfir" in n and "student_first" not in mapping: mapping["student_first"]=c
        if ("studentlast" in n or n=="studentla") and "student_last" not in mapping: mapping["student_last"]=c
        if "relationship" in n and "relationship" not in mapping: mapping["relationship"]=c
    return mapping

def letter(avg):
    if avg is None: return ""
    for l,lo,hi in get_setting("scale"):
        if float(lo)<=avg<=float(hi): return l
    return ""

def summary(sid,subject):
    c=conn()
    df=pd.read_sql_query("""SELECT a.category,a.standard_code,a.points_possible,g.points_earned
                            FROM grades g JOIN assignments a ON a.id=g.assignment_id
                            WHERE g.scholar_id=? AND a.subject=? AND g.points_earned IS NOT NULL""",
                         c,params=[sid,subject]); c.close()
    if df.empty: return None,pd.DataFrame()
    df["pct"]=df.points_earned/df.points_possible*100
    weights=get_setting("weights")
    cats=df.groupby("category").pct.mean().to_dict()
    used={k:v for k,v in weights.items() if k in cats}
    tw=sum(used.values())
    avg=sum(cats[k]*(used[k]/tw) for k in used) if tw else df.pct.mean()
    skills=df[df.standard_code!=""].groupby("standard_code").pct.mean().reset_index()
    return avg,skills


def guardian_options_for_scholar(sid):
    c=conn()
    g=pd.read_sql_query("""SELECT * FROM guardians WHERE scholar_id=? ORDER BY relationship,last_name,first_name""",c,params=[sid])
    c.close()
    options={0:"— No parent/guardian selected —"}
    for _,r in g.iterrows():
        display=(" ".join(x for x in [r.first_name,r.last_name] if x)).strip() or "Unnamed Guardian"
        if r.relationship:
            display += f" ({r.relationship})"
        options[int(r.id)] = display
    return options

def delete_record(table, record_id):
    allowed={"classes","standards","assignments","work_samples","support_notes","report_comments","communications","guardians"}
    if table not in allowed:
        raise ValueError("Table not allowed.")
    c=conn()
    c.execute(f"DELETE FROM {table} WHERE id=?",(int(record_id),))
    c.commit(); c.close()

def record_select(df, label_col, key):
    if df.empty:
        return None
    ids=list(df["id"].astype(int))
    labels={int(r.id):str(r[label_col]) for _,r in df.iterrows()}
    return st.selectbox("Select record",ids,format_func=lambda x:labels.get(x,str(x)),key=key)


def comment_text(name,subject,avg,skills,next_skill):
    lookup={r.code:r.skill for _,r in standards_df(subject).iterrows()}
    high=[]; low=[]; mid=[]
    for _,r in skills.iterrows():
        item=(lookup.get(r.standard_code,r.standard_code),float(r.pct))
        if r.pct>=85:
            high.append(item)
        elif r.pct<75:
            low.append(item)
        else:
            mid.append(item)
    high=sorted(high,key=lambda x:x[1],reverse=True)[:2]
    low=sorted(low,key=lambda x:x[1])[:2]

    if high:
        strength_text=" and ".join(x[0].lower() for x in high)
        opening=f"{name} has shown particular strength in {strength_text}."
    elif avg is not None and avg>=80:
        opening=f"{name} is showing a solid overall understanding of the {subject.lower()} skills assessed this marking period."
    else:
        opening=f"{name} is continuing to build understanding of grade-level {subject.lower()} skills."

    if low:
        need_text=" and ".join(x[0].lower() for x in low)
        growth=f" An area that would benefit from additional practice is {need_text}."
        home=f" At home, you can support {name} by reviewing corrected work, asking {name} to explain answers in their own words, and completing short practice activities focused on these skills."
        teacher=f" In class, I will continue providing targeted practice, feedback, and opportunities for {name} to revisit these skills."
    else:
        growth=" Continued attention to accuracy, complete explanations, and consistent application of learned skills will help maintain this progress."
        home=f" At home, you can support {name} by discussing what was learned in class, reviewing notes or corrected work, and asking {name} to explain the thinking behind an answer."
        teacher=f" In class, I will continue challenging {name} to apply these skills independently and explain reasoning clearly."

    nxt=f" Our next instructional focus will include {next_skill.lower()}, and I encourage you to ask {name} to share what they are learning about these skills." if next_skill else ""
    return (opening+growth+home+teacher+nxt).strip()

def academic_summary_for_scholar(sid):
    strengths=[]
    needs=[]
    teacher_actions=[]
    for subj in ["ELA","Math","Science","Social Studies"]:
        avg,skills=summary(sid,subj)
        if skills.empty:
            continue
        lookup={r.code:r.skill for _,r in standards_df(subj).iterrows()}
        for _,r in skills.iterrows():
            skill=lookup.get(r.standard_code,r.standard_code)
            pct=float(r.pct)
            if pct>=85:
                strengths.append((subj,skill,pct))
            elif pct<75:
                needs.append((subj,skill,pct))

    strengths=sorted(strengths,key=lambda x:x[2],reverse=True)[:5]
    needs=sorted(needs,key=lambda x:x[2])[:5]

    for subj,skill,pct in needs[:4]:
        teacher_actions.append(f"Provide targeted {subj} practice in {skill.lower()}, review errors with the scholar, and check for understanding after reteaching.")

    if not teacher_actions:
        teacher_actions.append("Continue providing grade-level practice, feedback, and opportunities to explain thinking independently.")

    return strengths,needs,teacher_actions

def guardian_display_map(sid):
    c=conn()
    g=pd.read_sql_query("SELECT * FROM guardians WHERE scholar_id=? ORDER BY relationship,last_name,first_name",c,params=[sid])
    c.close()
    out={0:"— No parent/guardian selected —"}
    for _,r in g.iterrows():
        name=(" ".join(x for x in [r.first_name,r.last_name] if x)).strip() or "Unnamed Guardian"
        if r.relationship:
            name+=f" ({r.relationship})"
        out[int(r.id)]=name
    return out

def log_communication(sid,gid,comm_type,subject,reason,details,generated_text=""):
    c=conn()
    c.execute("""INSERT INTO communications(
        scholar_id,guardian_id,created_at,communication_type,subject,reason,details,generated_text)
        VALUES (?,?,?,?,?,?,?,?)""",
        (int(sid),int(gid) if gid else None,date.today().strftime("%m/%d/%Y"),
         comm_type,subject,reason,details,generated_text))
    c.commit(); c.close()

def reminder_rows(sid=None):
    c=conn()
    q="""SELECT contact_reminders.*, scholars.first_name||' '||scholars.last_name scholar,
         COALESCE(TRIM(guardians.first_name||' '||guardians.last_name),'') guardian,
         guardians.relationship
         FROM contact_reminders
         LEFT JOIN scholars ON scholars.id=contact_reminders.scholar_id
         LEFT JOIN guardians ON guardians.id=contact_reminders.guardian_id
         WHERE 1=1"""
    p=[]
    if sid:
        q+=" AND contact_reminders.scholar_id=?"
        p=[int(sid)]
    q+=" ORDER BY completed ASC,due_date ASC,id DESC"
    df=pd.read_sql_query(q,c,params=p)
    c.close()
    return df


def benchmark_for_scholar(sid):
    c=conn()
    r=c.execute("SELECT * FROM benchmark_scores WHERE scholar_id=?",(int(sid),)).fetchone()
    c.close()
    return r

def fp_to_num(level):
    if level is None:
        return None
    s=str(level).strip().upper()
    if not s:
        return None
    # Supports common F&P-style letter levels A-Z
    if len(s)==1 and "A" <= s <= "Z":
        return ord(s)-ord("A")+1
    return None

def compare_book_to_scholar(book_level, scholar_level):
    b=fp_to_num(book_level)
    s=fp_to_num(scholar_level)
    if b is None or s is None:
        return "I need both a verified book level and the scholar's current F&P level to compare them."
    diff=b-s
    if diff <= -2:
        return "This book is below the scholar's current F&P level and may work well for independent fluency practice."
    if diff == -1:
        return "This book is slightly below the scholar's current F&P level and is likely appropriate for independent reading."
    if diff == 0:
        return "This book matches the scholar's current F&P level and is a strong choice for independent reading."
    if diff == 1:
        return "This book is slightly above the scholar's current F&P level and may be appropriate with light teacher or family support."
    return "This book is above the scholar's current F&P level and is better suited for supported or instructional reading."

def current_fp_level(sid):
    r=benchmark_for_scholar(sid)
    if not r:
        return ""
    return (r["fp_spring_level"] or r["fp_fall_level"] or "").strip()


def parse_score(value):
    try:
        if value is None:
            return None
        s=str(value).strip()
        if not s:
            return None
        return float(s)
    except:
        return None

def get_nwea_goal(class_id, season, subject):
    c=conn()
    if class_id:
        r=c.execute("""SELECT goal_score FROM nwea_goals
                       WHERE class_id=? AND season=? AND subject=?""",
                    (int(class_id),season,subject)).fetchone()
    else:
        r=c.execute("""SELECT goal_score FROM nwea_goals
                       WHERE class_id IS NULL AND season=? AND subject=?""",
                    (season,subject)).fetchone()
    c.close()
    return float(r["goal_score"]) if r and r["goal_score"] is not None else None

def save_nwea_goal(class_id, season, subject, goal):
    c=conn()
    # SQLite NULL values do not participate in UNIQUE exactly the way a normal key does,
    # so handle All Classes/global goals manually.
    if class_id:
        existing=c.execute("""SELECT id FROM nwea_goals
                              WHERE class_id=? AND season=? AND subject=?""",
                           (int(class_id),season,subject)).fetchone()
    else:
        existing=c.execute("""SELECT id FROM nwea_goals
                              WHERE class_id IS NULL AND season=? AND subject=?""",
                           (season,subject)).fetchone()
    if existing:
        c.execute("UPDATE nwea_goals SET goal_score=? WHERE id=?",(float(goal),int(existing["id"])))
    else:
        c.execute("INSERT INTO nwea_goals(class_id,season,subject,goal_score) VALUES (?,?,?,?)",
                  (int(class_id) if class_id else None,season,subject,float(goal)))
    c.commit(); c.close()

def class_nwea_dataframe(class_id, season):
    roster=scholars_df(class_id or None)
    rows=[]
    for _,srow in roster.iterrows():
        br=benchmark_for_scholar(int(srow.id))
        reading=""
        math=""
        if br:
            if season=="Fall":
                reading=br["nwea_fall_reading"] or ""
                math=br["nwea_fall_math"] or ""
            else:
                reading=br["nwea_spring_reading"] or ""
                math=br["nwea_spring_math"] or ""
        rows.append({
            "Scholar":nm(srow),
            "Reading / ELA":parse_score(reading),
            "Math":parse_score(math)
        })
    return pd.DataFrame(rows)

def nwea_rank_summary(df, column):
    valid=df[df[column].notna()].copy()
    if valid.empty:
        return None,None
    max_score=valid[column].max()
    min_score=valid[column].min()
    highest=valid[valid[column]==max_score]["Scholar"].tolist()
    lowest=valid[valid[column]==min_score]["Scholar"].tolist()
    return (highest,max_score),(lowest,min_score)


def get_grouping_cutoffs(class_id):
    c=conn()
    if class_id:
        r=c.execute("SELECT low_max,mid_max FROM grouping_settings WHERE class_id=?",(int(class_id),)).fetchone()
    else:
        r=c.execute("SELECT low_max,mid_max FROM grouping_settings WHERE class_id IS NULL").fetchone()
    c.close()
    if r:
        return float(r["low_max"]),float(r["mid_max"])
    return 69.0,84.0

def save_grouping_cutoffs(class_id, low_max, mid_max):
    c=conn()
    if class_id:
        r=c.execute("SELECT id FROM grouping_settings WHERE class_id=?",(int(class_id),)).fetchone()
    else:
        r=c.execute("SELECT id FROM grouping_settings WHERE class_id IS NULL").fetchone()
    if r:
        c.execute("UPDATE grouping_settings SET low_max=?,mid_max=? WHERE id=?",(float(low_max),float(mid_max),int(r["id"])))
    else:
        c.execute("INSERT INTO grouping_settings(class_id,low_max,mid_max) VALUES (?,?,?)",
                  (int(class_id) if class_id else None,float(low_max),float(mid_max)))
    c.commit(); c.close()

def group_label(score, low_max, mid_max):
    if score is None or pd.isna(score):
        return "No Data"
    score=float(score)
    if score<=low_max:
        return "Low"
    if score<=mid_max:
        return "Mid"
    return "High"

def subject_grouping_df(class_id, subject, low_max, mid_max):
    roster=scholars_df(class_id or None)
    rows=[]
    for _,sr in roster.iterrows():
        avg,_=summary(int(sr.id),subject)
        rows.append({
            "Scholar":nm(sr),
            "Average":None if avg is None else round(float(avg),1),
            "Group":group_label(avg,low_max,mid_max)
        })
    if not rows:
        return pd.DataFrame(columns=["Scholar","Average","Group"])
    df=pd.DataFrame(rows)
    if "Group" not in df.columns:
        df["Group"]="No Data"
    return df

def skill_grouping_df(class_id, subject, skill_code, low_max, mid_max):
    roster=scholars_df(class_id or None)
    rows=[]
    for _,sr in roster.iterrows():
        _,skills=summary(int(sr.id),subject)
        score=None
        if not skills.empty and "standard_code" in skills.columns and skill_code in set(skills["standard_code"]):
            score=float(skills.loc[skills["standard_code"]==skill_code,"pct"].iloc[0])
        rows.append({
            "Scholar":nm(sr),
            "Skill Average":None if score is None else round(score,1),
            "Group":group_label(score,low_max,mid_max)
        })
    if not rows:
        return pd.DataFrame(columns=["Scholar","Skill Average","Group"])
    df=pd.DataFrame(rows)
    if "Group" not in df.columns:
        df["Group"]="No Data"
    return df


def get_teacher_name():
    c=conn()
    r=c.execute("SELECT value FROM settings WHERE key='teacher_name'").fetchone()
    c.close()
    if not r:
        return ""
    try:
        return json.loads(r["value"])
    except:
        return str(r["value"] or "")

def save_teacher_name(name):
    c=conn()
    c.execute("INSERT OR REPLACE INTO settings(key,value) VALUES ('teacher_name',?)",(json.dumps(name.strip()),))
    c.commit(); c.close()

def teacher_dashboard_info():
    c=conn()
    out={
        "display_name":get_teacher_name() or "Ms. Chapman",
        "grade_title":"", "homeroom":"", "school":"", "subjects":[], "tagline":"",
        "show_grade_title":True, "show_homeroom":True, "show_school":True,
        "show_subjects":True, "show_tagline":True
    }
    key_map={
        "teacher_display_name":"display_name", "teacher_grade_title":"grade_title",
        "teacher_homeroom":"homeroom", "teacher_school":"school",
        "teacher_subjects":"subjects", "teacher_tagline":"tagline",
        "teacher_show_grade_title":"show_grade_title", "teacher_show_homeroom":"show_homeroom",
        "teacher_show_school":"show_school", "teacher_show_subjects":"show_subjects",
        "teacher_show_tagline":"show_tagline"
    }
    for db_key,out_key in key_map.items():
        r=c.execute("SELECT value FROM settings WHERE key=?",(db_key,)).fetchone()
        if not r:
            continue
        try:
            val=json.loads(r["value"])
        except Exception:
            val=r["value"]
        if out_key=="subjects":
            out[out_key]=val if isinstance(val,list) else []
        elif out_key.startswith("show_"):
            out[out_key]=bool(val)
        else:
            out[out_key]=str(val or "")
    c.close()
    return out

def save_teacher_dashboard_info(display_name, grade_title, homeroom, school, subjects, tagline,
                                show_grade_title=True, show_homeroom=True, show_school=True,
                                show_subjects=True, show_tagline=True):
    display_name=str(display_name or "").strip() or "Teacher"
    c=conn()
    vals={
        "teacher_display_name":display_name, "teacher_grade_title":str(grade_title or "").strip(),
        "teacher_homeroom":str(homeroom or "").strip(), "teacher_school":str(school or "").strip(),
        "teacher_subjects":list(subjects or []), "teacher_tagline":str(tagline or "").strip(),
        "teacher_show_grade_title":bool(show_grade_title), "teacher_show_homeroom":bool(show_homeroom),
        "teacher_show_school":bool(show_school), "teacher_show_subjects":bool(show_subjects),
        "teacher_show_tagline":bool(show_tagline)
    }
    for k,v in vals.items():
        c.execute("INSERT OR REPLACE INTO settings(key,value) VALUES (?,?)",(k,json.dumps(v)))
    c.execute("INSERT OR REPLACE INTO settings(key,value) VALUES ('teacher_name',?)",(json.dumps(display_name),))
    c.commit(); c.close()

def class_assignments_df(class_id, subject_filter="All Subjects", sort_order="Oldest → Newest"):
    c=conn()
    q="SELECT * FROM assignments WHERE 1=1"
    p=[]
    if class_id:
        q+=" AND (class_id=? OR class_id IS NULL)"
        p.append(int(class_id))
    if subject_filter!="All Subjects":
        q+=" AND subject=?"
        p.append(subject_filter)
    if sort_order=="Newest → Oldest":
        q+=" ORDER BY assignment_date DESC,id DESC"
    elif sort_order=="Subject":
        q+=" ORDER BY subject,assignment_date,id"
    elif sort_order=="Standard Code":
        q+=" ORDER BY standard_code,assignment_date,id"
    else:
        q+=" ORDER BY assignment_date ASC,id ASC"
    df=pd.read_sql_query(q,c,params=p)
    c.close()
    return df

def gradebook_matrix(class_id, subject_filter="All Subjects", sort_order="Oldest → Newest"):
    roster=scholars_df(class_id or None)
    assignments=class_assignments_df(class_id,subject_filter,sort_order)
    if roster.empty:
        return pd.DataFrame(),assignments
    c=conn()
    rows=[]
    for _,sr in roster.iterrows():
        row={"Scholar":nm(sr)}
        for _,a in assignments.iterrows():
            label=f"{a.title}\\n{a.standard_code or 'No Standard'}"
            gr=c.execute("SELECT points_earned FROM grades WHERE scholar_id=? AND assignment_id=?",
                         (int(sr.id),int(a.id))).fetchone()
            if gr and gr["points_earned"] is not None and float(a.points_possible):
                row[label]=round(float(gr["points_earned"])/float(a.points_possible)*100,1)
            else:
                row[label]=None
        rows.append(row)
    c.close()
    return pd.DataFrame(rows),assignments

def make_grade_sheet_xlsx(class_id, class_name, subject_filter="All Subjects", sort_order="Oldest → Newest"):
    roster=scholars_df(class_id or None)
    assignments=class_assignments_df(class_id,subject_filter,sort_order)
    wb=Workbook()
    ws=wb.active
    ws.title="Grade Entry"
    ws["A1"]="ChapLab Grade Entry Sheet"
    ws["A1"].font=Font(size=16,bold=True)
    ws["A2"]="Class"; ws["B2"]=class_name
    ws["D2"]="Subject Filter"; ws["E2"]=subject_filter
    ws["A3"]="Instructions"
    ws["B3"]="Enter points earned only. Do not change scholar names or the hidden assignment ID row. Save and upload this file back into ChapLab."
    ws["A5"]="Scholar ID"; ws["B5"]="Scholar Name"
    ws["A6"]="Scholar ID"; ws["B6"]="Scholar Name"
    ws.row_dimensions[5].hidden=True
    for j,(_,a) in enumerate(assignments.iterrows(),start=3):
        col=get_column_letter(j)
        ws.cell(row=5,column=j,value=int(a.id))
        ws.cell(row=6,column=j,value=f"{a.title}\\n{a.standard_code or 'No Standard'}\\n{a.assignment_date or ''}\\n{a.points_possible:g} pts")
        ws.cell(row=6,column=j).alignment=Alignment(wrap_text=True,horizontal="center",vertical="center")
        ws.column_dimensions[col].width=18
    c=conn()
    for i,(_,sr) in enumerate(roster.iterrows(),start=7):
        ws.cell(row=i,column=1,value=int(sr.id))
        ws.cell(row=i,column=2,value=nm(sr))
        for j,(_,a) in enumerate(assignments.iterrows(),start=3):
            gr=c.execute("SELECT points_earned FROM grades WHERE scholar_id=? AND assignment_id=?",
                         (int(sr.id),int(a.id))).fetchone()
            if gr and gr["points_earned"] is not None:
                ws.cell(row=i,column=j,value=float(gr["points_earned"]))
    c.close()
    thin=Side(style="thin",color="444444")
    for cell in ws[6]:
        if cell.column<=2 or cell.value is not None:
            cell.fill=PatternFill("solid",fgColor="F4D35E")
            cell.font=Font(bold=True,color="111111")
            cell.border=Border(left=thin,right=thin,top=thin,bottom=thin)
    ws.row_dimensions[6].height=62
    max_col=max(2,2+len(assignments))
    for row in ws.iter_rows(min_row=7,max_row=6+len(roster),min_col=1,max_col=max_col):
        for cell in row:
            cell.border=Border(left=thin,right=thin,top=thin,bottom=thin)
            cell.alignment=Alignment(horizontal="left" if cell.column==2 else "center")
    ws.column_dimensions["A"].width=12
    ws.column_dimensions["B"].width=24
    ws.freeze_panes="C7"
    ws.sheet_view.showGridLines=False
    ws.page_setup.orientation="landscape"
    ws.page_setup.fitToWidth=1
    ws.page_setup.fitToHeight=0
    ws.sheet_properties.pageSetUpPr.fitToPage=True
    ws.print_title_rows="1:6"
    bio=BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.getvalue()

def import_grade_sheet_xlsx(uploaded_file):
    wb=load_workbook(uploaded_file,data_only=True)
    ws=wb["Grade Entry"] if "Grade Entry" in wb.sheetnames else wb.active
    assignment_ids=[]
    for col in range(3,ws.max_column+1):
        aid=ws.cell(row=5,column=col).value
        if aid is not None:
            try:
                assignment_ids.append((col,int(aid)))
            except:
                pass
    saved=0; skipped=0
    c=conn()
    for row in range(7,ws.max_row+1):
        sid=ws.cell(row=row,column=1).value
        if sid is None:
            continue
        try:
            sid=int(sid)
        except:
            continue
        for col,aid in assignment_ids:
            value=ws.cell(row=row,column=col).value
            if value in (None,""):
                continue
            try:
                points=float(value)
            except:
                skipped+=1; continue
            ar=c.execute("SELECT points_possible FROM assignments WHERE id=?",(aid,)).fetchone()
            if not ar or points<0 or points>float(ar["points_possible"]):
                skipped+=1; continue
            c.execute("INSERT OR REPLACE INTO grades(scholar_id,assignment_id,points_earned) VALUES (?,?,?)",
                      (sid,aid,points))
            saved+=1
    c.commit(); c.close()
    return saved,skipped

def class_dashboard_metrics(class_id):
    roster=scholars_df(class_id or None)
    total=len(roster)
    contacts=0
    averages=[]
    c=conn()
    if not roster.empty:
        ids=[int(x) for x in roster.id]
        placeholders=",".join(["?"]*len(ids))
        contacts=c.execute(f"SELECT COUNT(*) n FROM communications WHERE scholar_id IN ({placeholders})",ids).fetchone()["n"]
    c.close()
    for _,sr in roster.iterrows():
        vals=[]
        for subj in ["ELA","Math","Science","Social Studies"]:
            av,_=summary(int(sr.id),subj)
            if av is not None:
                vals.append(av)
        if vals:
            averages.append(sum(vals)/len(vals))
    overall=sum(averages)/len(averages) if averages else None
    return total,contacts,overall


def class_name_from_id(class_id):
    if not class_id:
        return ""
    c=conn()
    r=c.execute("SELECT class_name FROM classes WHERE id=?",(int(class_id),)).fetchone()
    c.close()
    return r["class_name"] if r else ""

def scholar_full_profile(sid):
    c=conn()
    r=c.execute("""SELECT s.*, COALESCE(c.class_name,s.class_name,'') display_class
                   FROM scholars s LEFT JOIN classes c ON c.id=s.class_id
                   WHERE s.id=?""",(int(sid),)).fetchone()
    c.close()
    return r

def iep_prefill_summary(sid):
    s=scholar_full_profile(sid)
    if not s:
        return ""
    name=f"{s['first_name']} {s['last_name']}"
    parts=[
        f"Scholar: {name}",
        f"School ID: {s['student_id'] or '—'}",
        f"Grade/Class: {s['grade_level'] or '—'} / {s['display_class'] or '—'}",
        f"School Year: {s['academic_year'] or '—'}",
    ]
    # Grades
    grade_bits=[]
    for subj in ["ELA","Math","Science","Social Studies"]:
        av,_=summary(int(sid),subj)
        if av is not None:
            grade_bits.append(f"{subj} {av:.1f}% ({letter(av)})")
    if grade_bits:
        parts.append("Current grades: " + "; ".join(grade_bits))

    # Benchmarks
    br=benchmark_for_scholar(sid)
    if br:
        bench=[]
        if br["nwea_fall_reading"] or br["nwea_spring_reading"]:
            bench.append(f"NWEA Reading Fall {br['nwea_fall_reading'] or '—'} / Spring {br['nwea_spring_reading'] or '—'}")
        if br["nwea_fall_math"] or br["nwea_spring_math"]:
            bench.append(f"NWEA Math Fall {br['nwea_fall_math'] or '—'} / Spring {br['nwea_spring_math'] or '—'}")
        if br["fp_fall_level"] or br["fp_spring_level"]:
            bench.append(f"F&P Fall {br['fp_fall_level'] or '—'} / Spring {br['fp_spring_level'] or '—'}")
        if br["fp_fall_word_list"] or br["fp_spring_word_list"]:
            bench.append(f"F&P Word List Fall {br['fp_fall_word_list'] or '—'} / Spring {br['fp_spring_word_list'] or '—'}")
        if bench:
            parts.append("Benchmark data: " + "; ".join(bench))

    strengths,needs,teacher_actions=academic_summary_for_scholar(sid)
    if strengths:
        parts.append("Academic strengths: " + "; ".join(f"{subj} - {skill} ({pct:.0f}%)" for subj,skill,pct in strengths[:4]))
    if needs:
        parts.append("Areas needing support: " + "; ".join(f"{subj} - {skill} ({pct:.0f}%)" for subj,skill,pct in needs[:4]))

    c=conn()
    notes=pd.read_sql_query("SELECT * FROM support_notes WHERE scholar_id=? ORDER BY id DESC LIMIT 5",c,params=[sid])
    ws=pd.read_sql_query("SELECT * FROM work_samples WHERE scholar_id=? ORDER BY id DESC LIMIT 5",c,params=[sid])
    c.close()
    if not notes.empty:
        obs=[]
        for _,r in notes.iterrows():
            obs.append(f"{r['area']}: {r['observation']}")
        parts.append("Recent support/IEP observations: " + " | ".join(obs))
    if not ws.empty:
        evidence=[]
        for _,r in ws.iterrows():
            title=r["title"] or r["file_name"] or "Work sample"
            evidence.append(f"{title} — strengths: {r['strengths'] or '—'}; needs: {r['needs'] or '—'}")
        parts.append("Recent work-sample evidence: " + " | ".join(evidence))
    return "\n".join(parts)


def quarter_grade_data(sid, subject, marking_period):
    c=conn()
    df=pd.read_sql_query("""SELECT a.id,a.title,a.category,a.standard_code,a.points_possible,
                                   a.assignment_date,a.marking_period,g.points_earned
                            FROM grades g
                            JOIN assignments a ON a.id=g.assignment_id
                            WHERE g.scholar_id=? AND a.subject=? AND g.points_earned IS NOT NULL
                              AND COALESCE(a.marking_period,'')=?
                            ORDER BY a.assignment_date ASC,a.id ASC""",
                         c,params=[int(sid),subject,marking_period])
    c.close()
    if df.empty:
        return df
    df["pct"]=df["points_earned"]/df["points_possible"]*100
    return df

def weighted_average_from_grade_rows(df):
    if df.empty:
        return None
    weights=get_setting("weights")
    cat_avgs=df.groupby("category")["pct"].mean().to_dict()
    used={k:v for k,v in weights.items() if k in cat_avgs}
    totalw=sum(used.values())
    if totalw:
        return sum(cat_avgs[k]*(used[k]/totalw) for k in used)
    return float(df["pct"].mean())

def quarter_skill_summary(df, subject):
    if df.empty:
        return [],[]
    lookup={r.code:r.skill for _,r in standards_df(subject).iterrows()}
    skill_df=df[df["standard_code"].astype(str)!=""].groupby("standard_code")["pct"].mean().reset_index()
    strengths=[]
    needs=[]
    for _,r in skill_df.iterrows():
        skill=lookup.get(r["standard_code"],r["standard_code"])
        pct=float(r["pct"])
        if pct>=85:
            strengths.append((skill,pct))
        elif pct<75:
            needs.append((skill,pct))
    strengths=sorted(strengths,key=lambda x:x[1],reverse=True)[:3]
    needs=sorted(needs,key=lambda x:x[1])[:3]
    return strengths,needs

def quarter_grade_pattern(df):
    if df.empty:
        return {}
    result={}
    result["count"]=len(df)
    result["average"]=weighted_average_from_grade_rows(df)
    result["highest"]=df.loc[df["pct"].idxmax()]
    result["lowest"]=df.loc[df["pct"].idxmin()]
    result["first_half_avg"]=None
    result["second_half_avg"]=None
    result["trend"]=""

    if len(df)>=4:
        midpoint=max(1,len(df)//2)
        first=float(df.iloc[:midpoint]["pct"].mean())
        second=float(df.iloc[midpoint:]["pct"].mean())
        result["first_half_avg"]=first
        result["second_half_avg"]=second
        if second-first>=5:
            result["trend"]="improved over the course of the marking period"
        elif first-second>=5:
            result["trend"]="showed some decline later in the marking period"
        else:
            result["trend"]="performed fairly consistently across the marking period"

    cat_avgs=df.groupby("category")["pct"].mean().sort_values(ascending=False)
    result["category_best"]=(cat_avgs.index[0],float(cat_avgs.iloc[0])) if len(cat_avgs) else None
    result["category_need"]=(cat_avgs.index[-1],float(cat_avgs.iloc[-1])) if len(cat_avgs)>1 else None
    return result

def quarter_report_comment(name, subject, marking_period, df, next_skills, sid=None):
    pro=scholar_pronouns(sid) if sid else pronoun_set_from_text("")
    subj_pr=pro["subject"]
    poss_pr=pro["possessive"]
    if df.empty:
        return f"{name} does not yet have graded {subject.lower()} assignments tagged to {marking_period}. Add or update assignment marking periods before generating this comment."

    pattern=quarter_grade_pattern(df)
    strengths,needs=quarter_skill_summary(df,subject)
    avg=pattern["average"]
    grade_text=letter(avg) if avg is not None else ""

    opening=f"During {marking_period}, {name} earned a {avg:.1f}% ({grade_text}) in {subject} based on {pattern['count']} graded assignment"
    opening += "s." if pattern["count"]!=1 else "."

    evidence=[]
    if strengths:
        evidence.append("demonstrated strong understanding of " + " and ".join(x[0].lower() for x in strengths[:2]))
    if pattern.get("category_best"):
        cat,catavg=pattern["category_best"]
        evidence.append(f"performed especially well on {cat.lower()} work, averaging {catavg:.1f}%")
    if pattern.get("trend"):
        evidence.append(pattern["trend"])
    if evidence:
        evidence_sentence=f" {cap_pronoun(subj_pr)} " + ", and ".join(evidence) + "."
    else:
        evidence_sentence=""

    if needs:
        need_text=" and ".join(x[0].lower() for x in needs[:2])
        growth=f" An area for continued growth is {need_text}."
    elif pattern.get("category_need") and pattern["category_need"][1] < 80:
        cat,catavg=pattern["category_need"]
        growth=f" Continued practice with {cat.lower()} assignments, where the average was {catavg:.1f}%, will help strengthen overall performance."
    else:
        growth=" Continued attention to accuracy, complete explanations, and consistent application of learned skills will support continued progress."

    high=pattern["highest"]
    low=pattern["lowest"]
    assignment_evidence=""
    if len(df)>=2:
        assignment_evidence=(f" For example, {name}'s strongest recorded performance was on "
                             f"“{high['title']}” ({high['pct']:.1f}%), while “{low['title']}” "
                             f"({low['pct']:.1f}%) shows an area where additional review may be helpful.")

    if needs:
        home=f" At home, you can support {name} by reviewing corrected work connected to these skills, asking {subj_pr} to explain answers in {poss_pr} own words, and completing short practice activities that target the areas needing improvement."
    else:
        home=f" At home, you can support {name} by reviewing classwork, discussing what was learned, and asking {subj_pr} to explain the reasoning behind answers."

    teacher=f" In class, I will continue using grade and skill data to provide targeted feedback, reteaching, and opportunities for {name} to practice independently while building {poss_pr} confidence."

    nxt=""
    if next_skills:
        nxt=f" Our next instructional focus will include {next_skills.lower()}, and practicing these concepts at home will help {name} build confidence with upcoming work."

    return (opening+evidence_sentence+growth+assignment_evidence+home+teacher+nxt).strip()


def current_academic_year():
    # Prefer saved scholar academic year if available, else teacher setting
    c=conn()
    r=c.execute("SELECT academic_year FROM scholars WHERE TRIM(COALESCE(academic_year,''))<>'' LIMIT 1").fetchone()
    c.close()
    return r["academic_year"] if r else ""

def get_quarter_settings(academic_year=""):
    c=conn()
    df=pd.read_sql_query("""SELECT * FROM quarter_settings
                            WHERE academic_year=?
                            ORDER BY CASE quarter_name
                              WHEN 'Quarter 1' THEN 1
                              WHEN 'Quarter 2' THEN 2
                              WHEN 'Quarter 3' THEN 3
                              WHEN 'Quarter 4' THEN 4
                              ELSE 5 END""",c,params=[academic_year])
    c.close()
    return df

def save_quarter_setting(academic_year, quarter_name, start_date, end_date, locked=False):
    c=conn()
    c.execute("""INSERT INTO quarter_settings(academic_year,quarter_name,start_date,end_date,locked)
                 VALUES (?,?,?,?,?)
                 ON CONFLICT(academic_year,quarter_name) DO UPDATE SET
                 start_date=excluded.start_date,end_date=excluded.end_date,locked=excluded.locked""",
              (academic_year,quarter_name,start_date,end_date,1 if locked else 0))
    c.commit(); c.close()

def quarter_for_date(date_text, academic_year=""):
    if not date_text:
        return ""
    try:
        d=pd.to_datetime(date_text).date()
    except:
        return ""
    qdf=get_quarter_settings(academic_year)
    for _,r in qdf.iterrows():
        try:
            start=pd.to_datetime(r["start_date"]).date()
            end=pd.to_datetime(r["end_date"]).date()
            if start <= d <= end:
                return r["quarter_name"]
        except:
            pass
    return ""

def report_deadline(academic_year, quarter_name):
    c=conn()
    r=c.execute("""SELECT due_date FROM report_card_deadlines
                   WHERE academic_year=? AND quarter_name=?""",
                (academic_year,quarter_name)).fetchone()
    c.close()
    return r["due_date"] if r else ""

def save_report_deadline(academic_year, quarter_name, due_date):
    c=conn()
    c.execute("""INSERT INTO report_card_deadlines(academic_year,quarter_name,due_date)
                 VALUES (?,?,?)
                 ON CONFLICT(academic_year,quarter_name) DO UPDATE SET due_date=excluded.due_date""",
              (academic_year,quarter_name,due_date))
    c.commit(); c.close()

def update_preference(sid):
    c=conn()
    r=c.execute("SELECT * FROM parent_update_preferences WHERE scholar_id=?",(int(sid),)).fetchone()
    c.close()
    return r

def save_update_preference(sid, requested, frequency, guardian_id, notes):
    c=conn()
    c.execute("""INSERT INTO parent_update_preferences(
                    scholar_id,requested_updates,update_frequency,preferred_guardian_id,notes)
                 VALUES (?,?,?,?,?)
                 ON CONFLICT(scholar_id) DO UPDATE SET
                    requested_updates=excluded.requested_updates,
                    update_frequency=excluded.update_frequency,
                    preferred_guardian_id=excluded.preferred_guardian_id,
                    notes=excluded.notes""",
              (int(sid),1 if requested else 0,frequency,int(guardian_id) if guardian_id else None,notes))
    c.commit(); c.close()

def latest_subject_average(sid, subject):
    av,_=summary(sid,subject)
    return av

def subject_growth_by_quarter(sid, subject, academic_year=""):
    out=[]
    for q in ["Quarter 1","Quarter 2","Quarter 3","Quarter 4"]:
        qdf=quarter_grade_data(sid,subject,q)
        if not qdf.empty:
            out.append((q,weighted_average_from_grade_rows(qdf)))
    return out

def scholar_risk_and_growth(sid):
    strengths=[]
    risks=[]
    growth_notes=[]
    for subj in ["ELA","Math","Science","Social Studies"]:
        av,_=summary(sid,subj)
        if av is not None:
            if av>=85:
                strengths.append(f"{subj} ({av:.1f}%)")
            elif av<70:
                risks.append(f"{subj} ({av:.1f}%)")
        seq=subject_growth_by_quarter(sid,subj)
        if len(seq)>=2:
            prev=seq[-2][1]
            cur=seq[-1][1]
            delta=cur-prev
            if delta>=5:
                growth_notes.append(f"{subj} improved by {delta:.1f} points from {seq[-2][0]} to {seq[-1][0]}")
            elif delta<=-5:
                risks.append(f"{subj} dropped by {abs(delta):.1f} points from {seq[-2][0]} to {seq[-1][0]}")
    return strengths,risks,growth_notes

def generate_parent_progress_update(sid, subject="General"):
    s=scholar_full_profile(sid)
    if not s:
        return ""
    name=f"{s['first_name']} {s['last_name']}"
    strengths,risks,growth_notes=scholar_risk_and_growth(sid)
    skill_strengths,skill_needs,teacher_actions=academic_summary_for_scholar(sid)

    parts=[f"I wanted to share a progress update about {name}."]
    if subject!="General":
        av=latest_subject_average(sid,subject)
        if av is not None:
            parts.append(f"{name}'s current {subject} average is {av:.1f}% ({letter(av)}).")

    if growth_notes:
        parts.append("A positive trend I am seeing is " + "; ".join(growth_notes[:2]) + ".")
    elif strengths:
        pro=scholar_pronouns(sid)
        parts.append(f"{name} is currently showing strength in " + ", ".join(strengths[:2]) + f", and {pro['subject']} should continue building on this progress.")

    if skill_strengths:
        skilltxt=", ".join(f"{subj}: {skill}" for subj,skill,pct in skill_strengths[:2])
        parts.append(f"Specific strengths include {skilltxt}.")

    if risks:
        parts.append("An area I am watching closely is " + ", ".join(risks[:2]) + ".")
    elif skill_needs:
        needtxt=", ".join(f"{subj}: {skill}" for subj,skill,pct in skill_needs[:2])
        parts.append(f"Skills that would benefit from more practice include {needtxt}.")

    if skill_needs:
        need_names=", ".join(skill.lower() for _,skill,_ in skill_needs[:2])
        parts.append(f"At home, you can help by reviewing corrected work, asking {name} to explain answers aloud, and completing short practice focused on {need_names}.")
    else:
        parts.append(f"At home, you can help by asking {name} to explain what was learned, review classwork, and practice applying skills independently.")

    if teacher_actions:
        if teacher_actions:
            action=teacher_actions[0]
            parts.append("At school, I will continue to " + action[0].lower() + action[1:])
    return " ".join(p for p in parts if p).strip()

def quarter_missing_grade_check(sid, subject, quarter, class_id=None):
    c=conn()
    q="""SELECT a.id,a.title,a.points_possible
         FROM assignments a
         WHERE a.subject=? AND COALESCE(a.marking_period,'')=?"""
    p=[subject,quarter]
    if class_id:
        q+=" AND (a.class_id=? OR a.class_id IS NULL)"
        p.append(int(class_id))
    assignments=pd.read_sql_query(q,c,params=p)
    missing=[]
    for _,a in assignments.iterrows():
        g=c.execute("SELECT points_earned FROM grades WHERE scholar_id=? AND assignment_id=?",
                    (int(sid),int(a.id))).fetchone()
        if not g or g["points_earned"] is None:
            missing.append(a["title"])
    c.close()
    return missing, len(assignments)

def closeout_row(sid, quarter, academic_year):
    c=conn()
    r=c.execute("""SELECT * FROM quarter_closeout
                   WHERE scholar_id=? AND quarter_name=? AND academic_year=?""",
                (int(sid),quarter,academic_year)).fetchone()
    c.close()
    return r

def save_closeout(sid, quarter, academic_year, grades_reviewed, data_reviewed, comment_generated, comment_finalized):
    c=conn()
    c.execute("""INSERT INTO quarter_closeout(
                    scholar_id,quarter_name,academic_year,grades_reviewed,data_reviewed,comment_generated,comment_finalized)
                 VALUES (?,?,?,?,?,?,?)
                 ON CONFLICT(scholar_id,quarter_name,academic_year) DO UPDATE SET
                    grades_reviewed=excluded.grades_reviewed,
                    data_reviewed=excluded.data_reviewed,
                    comment_generated=excluded.comment_generated,
                    comment_finalized=excluded.comment_finalized""",
              (int(sid),quarter,academic_year,1 if grades_reviewed else 0,1 if data_reviewed else 0,
               1 if comment_generated else 0,1 if comment_finalized else 0))
    c.commit(); c.close()


def go_to_page(page_name):
    st.session_state["nav_page"]=page_name

def go_to_section(page_name, binder_tool=None, assistant_tool=None, show_class_settings=False, show_quarter_settings=False):
    st.session_state["nav_page"]=page_name
    if binder_tool is not None:
        st.session_state["class_binder_tool"]=binder_tool
    if assistant_tool is not None:
        st.session_state["assistant_tool"]=assistant_tool
    if show_class_settings:
        st.session_state["show_home_class_settings"]=True
    if show_quarter_settings:
        st.session_state["show_quarter_settings"]=True

def select_class(class_id):
    st.session_state["selected_class"]=int(class_id)
    st.session_state["nav_page"]="Class Dashboard"
    st.session_state["class_binder_tool"]="Overview"

def open_main_dashboard():
    st.session_state["nav_page"]="Home Page"

def open_side_section(section_name):
    mapping={
        "main":"Home Page",
        "scholars":"Scholars",
        "grades":"Scholar Binder",
        "books":"Book Leveler",
        "grouping":"Student Grouping",
        "reports":"Report Card Comments",
        "assistant":"Little Assistant",
        "bulletin":"Bulletin Board",
        "communication":"Communication Log",
        "web":"Web & Backup",
    }
    st.session_state["nav_page"]=mapping.get(section_name,"Main Dashboard")
    if section_name=="grades":
        st.session_state["class_binder_tool"]="Overview"

def open_scholar_profile(sid):
    st.session_state["selected_profile_scholar"]=int(sid)
    st.session_state["nav_page"]="Scholar Profile"

def return_to_scholars():
    st.session_state["nav_page"]="Scholars"

def overall_scholar_average(sid):
    vals=[]
    for subj in ["ELA","Math","Science","Social Studies"]:
        av,_=summary(int(sid),subj)
        if av is not None:
            vals.append(float(av))
    return (sum(vals)/len(vals)) if vals else None

def scholar_status_summary(class_id=None):
    roster=scholars_df(class_id or None)
    low_max,mid_max=get_grouping_cutoffs(class_id)
    counts={"On Track":0,"Approaching":0,"At Risk":0,"No Data":0}
    detail=[]
    for _,sr in roster.iterrows():
        avg=overall_scholar_average(int(sr.id))
        if avg is None:
            status="No Data"
        elif avg<=low_max:
            status="At Risk"
        elif avg<=mid_max:
            status="Approaching"
        else:
            status="On Track"
        counts[status]+=1
        detail.append({"Scholar":nm(sr),"Average":None if avg is None else round(avg,1),"Status":status})
    return counts,pd.DataFrame(detail)

def recent_assignments_df(limit=5):
    c=conn()
    df=pd.read_sql_query("""SELECT a.*,c.class_name
                            FROM assignments a
                            LEFT JOIN classes c ON c.id=a.class_id
                            ORDER BY COALESCE(a.assignment_date,'') DESC,a.id DESC
                            LIMIT ?""",c,params=[int(limit)])
    c.close()
    return df

def home_announcements():
    items=[]
    ay=current_academic_year()
    for q in ["Quarter 1","Quarter 2","Quarter 3","Quarter 4"]:
        due=report_deadline(ay,q)
        if due:
            items.append(("Report Cards",f"{q} comments due",due))
    rem=reminder_rows()
    if not rem.empty:
        for _,r in rem[rem.completed==0].head(4).iterrows():
            items.append(("Parent Update",f"{r.scholar}: {r.reason}",r.due_date or ""))
    return items[:5]


def configure_local_tesseract():
    """Find a local Windows Tesseract install. Returns (ok, message)."""
    try:
        import pytesseract
        candidates=[
            r"C:\Program Files\Tesseract-OCR\tesseract.exe",
            r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        ]
        for path in candidates:
            if os.path.exists(path):
                pytesseract.pytesseract.tesseract_cmd=path
                return True,path
        try:
            _=pytesseract.get_tesseract_version()
            return True,"PATH"
        except:
            return False,"Tesseract OCR is not installed or could not be found."
    except Exception as e:
        return False,str(e)

def normalize_ocr_text(value):
    return re.sub(r'[^a-z0-9 ]',' ',str(value).lower()).strip()

def name_similarity(a,b):
    a=normalize_ocr_text(a)
    b=normalize_ocr_text(b)
    if not a or not b:
        return 0
    return SequenceMatcher(None,a,b).ratio()

def ocr_grade_screenshot(image_file, roster, expected_assignments):
    """
    Extract likely scholar rows and numeric grades from a screenshot.
    expected_assignments is an ordered list of assignment rows/records.
    """
    from PIL import Image, ImageEnhance, ImageFilter
    import pytesseract
    from pytesseract import Output

    ok,msg=configure_local_tesseract()
    if not ok:
        raise RuntimeError(
            "Local OCR is not ready. Install Tesseract OCR for Windows, then restart ChapLab. "
            "The screenshot stays on your computer."
        )

    image=Image.open(image_file).convert("L")
    # Enlarge small screenshots and improve contrast.
    if image.width < 1800:
        scale=max(1.0,1800/image.width)
        image=image.resize((int(image.width*scale),int(image.height*scale)))
    image=ImageEnhance.Contrast(image).enhance(1.8)
    image=image.filter(ImageFilter.SHARPEN)

    data=pytesseract.image_to_data(image,output_type=Output.DATAFRAME,config="--psm 6")
    data=data.dropna(subset=["text"])
    data["text"]=data["text"].astype(str).str.strip()
    data=data[data["text"]!=""]
    if data.empty:
        return pd.DataFrame(), "No readable text was detected."

    # Group OCR words into visual lines.
    line_cols=["block_num","par_num","line_num"]
    lines=[]
    for _,group in data.groupby(line_cols,sort=False):
        group=group.sort_values("left")
        words=group["text"].tolist()
        joined=" ".join(words)
        top=float(group["top"].min())
        left=float(group["left"].min())
        right=float((group["left"]+group["width"]).max())
        lines.append({"text":joined,"words":words,"top":top,"left":left,"right":right,"group":group})

    roster_records=[]
    for _,r in roster.iterrows():
        roster_records.append({
            "Scholar ID":int(r.id),
            "Scholar":nm(r),
            "first":str(r.first_name),
            "last":str(r.last_name)
        })

    results=[]
    n_assign=len(expected_assignments)
    for line in lines:
        line_text=line["text"]
        norm=normalize_ocr_text(line_text)

        # Find best roster match using full name and last name.
        best=None
        best_score=0
        for rr in roster_records:
            full_score=name_similarity(line_text,rr["Scholar"])
            last_score=1.0 if normalize_ocr_text(rr["last"]) in norm and len(rr["last"])>=3 else name_similarity(line_text,rr["last"])*0.92
            first_score=1.0 if normalize_ocr_text(rr["first"]) in norm and len(rr["first"])>=3 else name_similarity(line_text,rr["first"])*0.85
            score=max(full_score,(last_score+first_score)/2)
            if score>best_score:
                best_score=score
                best=rr

        if not best or best_score < 0.55:
            continue

        # Extract numeric-looking tokens from the same row.
        nums=[]
        for token in line["words"]:
            clean=re.sub(r'[^0-9./-]','',token)
            # common forms: 18, 18/20, 90, 90.5
            if re.fullmatch(r'\d+(?:\.\d+)?',clean):
                nums.append(float(clean))
            elif re.fullmatch(r'\d+(?:\.\d+)?/\d+(?:\.\d+)?',clean):
                earned,total=clean.split("/")
                try:
                    nums.append(float(earned))
                except:
                    pass

        # Remove values likely coming from scholar IDs when the screenshot includes them.
        sid_text=str(best["Scholar ID"])
        nums_filtered=[]
        for num in nums:
            if str(int(num))==sid_text and num.is_integer():
                continue
            nums_filtered.append(num)

        if not nums_filtered:
            continue

        # Use the rightmost N numeric values, which most gradebooks place after the name.
        vals=nums_filtered[-n_assign:] if n_assign else nums_filtered
        row={"Scholar ID":best["Scholar ID"],"Scholar":best["Scholar"],"OCR Match %":round(best_score*100)}
        for i,a in enumerate(expected_assignments):
            col=f"{a['title']} ({a['points_possible']:g} pts)"
            row[col]=vals[i] if i < len(vals) else None
        results.append(row)

    if not results:
        return pd.DataFrame(), "OCR ran, but no scholar/grade rows could be matched. Try a clearer crop showing names and grades."

    result=pd.DataFrame(results)
    # Keep the strongest match if scholar appears twice.
    result=result.sort_values("OCR Match %",ascending=False).drop_duplicates("Scholar ID").sort_values("Scholar")
    return result, f"Matched {len(result)} scholar rows. Review every score before importing."

def import_ocr_preview(preview_df, assignments_by_col):
    saved=0
    skipped=0
    c=conn()
    for _,row in preview_df.iterrows():
        try:
            sid=int(row["Scholar ID"])
        except:
            skipped+=1
            continue
        for col,aid in assignments_by_col.items():
            value=row.get(col)
            if pd.isna(value) or value=="":
                continue
            try:
                score=float(value)
            except:
                skipped+=1
                continue
            ar=c.execute("SELECT points_possible FROM assignments WHERE id=?",(int(aid),)).fetchone()
            if not ar or score<0 or score>float(ar["points_possible"]):
                skipped+=1
                continue
            c.execute("INSERT OR REPLACE INTO grades(scholar_id,assignment_id,points_earned) VALUES (?,?,?)",
                      (sid,int(aid),score))
            saved+=1
    c.commit()
    c.close()
    return saved,skipped



def decode_isbn_barcode(image_file):
    """Aggressively decode a book ISBN barcode from phone/computer captures."""
    if image_file is None:
        return None, "No barcode image was provided."

    try:
        import zxingcpp
    except Exception as e:
        return None, f"Barcode reader could not start: {e}"

    try:
        from PIL import Image, ImageEnhance, ImageOps, ImageFilter
        import numpy as np

        if hasattr(image_file,"convert") and hasattr(image_file,"size"):
            img=image_file.convert("RGB")
        else:
            raw=image_file.getvalue() if hasattr(image_file,"getvalue") else bytes(image_file.getbuffer())
            img=Image.open(BytesIO(raw)).convert("RGB")

        w,h=img.size

        # The compact scanner guide is centered, so prioritize center-band crops.
        # Also keep broad/full frame passes in case the barcode is slightly off-center.
        regions=[
            img,
            img.crop((int(w*.03),int(h*.15),int(w*.97),int(h*.85))),
            img.crop((int(w*.08),int(h*.25),int(w*.92),int(h*.75))),
            img.crop((int(w*.14),int(h*.31),int(w*.86),int(h*.69))),
            img.crop((int(w*.20),int(h*.36),int(w*.80),int(h*.64))),
        ]

        decoded=[]
        def harvest(candidate):
            arrays=[]
            try:
                arr=np.array(candidate)
                arrays.append(arr)
            except Exception:
                return None

            # Optional OpenCV passes give barcode lines more separation.
            try:
                import cv2
                arr=np.array(candidate)
                if arr.ndim==3:
                    gray=cv2.cvtColor(arr,cv2.COLOR_RGB2GRAY)
                else:
                    gray=arr
                arrays.append(gray)
                blur=cv2.GaussianBlur(gray,(3,3),0)
                arrays.append(blur)
                arrays.append(cv2.equalizeHist(gray))
                arrays.append(cv2.adaptiveThreshold(
                    gray,255,cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                    cv2.THRESH_BINARY,31,7
                ))
                _,otsu=cv2.threshold(gray,0,255,cv2.THRESH_BINARY+cv2.THRESH_OTSU)
                arrays.append(otsu)
            except Exception:
                pass

            for arr in arrays:
                for rotate in (True,False):
                    try:
                        results=zxingcpp.read_barcodes(
                            arr,
                            try_rotate=rotate,
                            try_downscale=True,
                            try_invert=True
                        )
                    except TypeError:
                        try:
                            results=zxingcpp.read_barcodes(arr)
                        except Exception:
                            results=[]
                    except Exception:
                        results=[]

                    for result in results:
                        value=str(getattr(result,"text","") or "").strip()
                        cleaned=re.sub(r"[^0-9Xx]","",value)
                        if cleaned and cleaned not in decoded:
                            decoded.append(cleaned)
                        if len(cleaned)==13 and cleaned.startswith(("978","979")):
                            return cleaned
            return None

        for region in regions:
            base=[region]
            gray=ImageOps.grayscale(region)
            sharp=gray.filter(ImageFilter.UnsharpMask(radius=2,percent=220,threshold=1))
            base.extend([
                gray,
                sharp,
                ImageEnhance.Contrast(sharp).enhance(1.6),
                ImageEnhance.Contrast(sharp).enhance(2.2),
                ImageEnhance.Contrast(sharp).enhance(3.0),
            ])

            # Upscale every region aggressively; barcode bars need pixel separation.
            for target_w in (1800,2400):
                if region.width < target_w:
                    scale=target_w/max(region.width,1)
                    up=region.resize(
                        (int(region.width*scale),int(region.height*scale)),
                        Image.Resampling.LANCZOS
                    )
                    upg=ImageOps.grayscale(up)
                    ups=upg.filter(ImageFilter.UnsharpMask(radius=2,percent=240,threshold=1))
                    base.extend([
                        up,upg,ups,
                        ImageEnhance.Contrast(ups).enhance(1.8),
                        ImageEnhance.Contrast(ups).enhance(2.6),
                    ])

            for candidate in base:
                found=harvest(candidate)
                if found:
                    return found,None

        for code in decoded:
            if len(code) in (10,13):
                return code,None

        if decoded:
            return None, "A barcode was detected, but it was not a valid ISBN. Keep the 978/979 ISBN barcode inside the guide box."
        return None, (
            "ChapLab couldn't read the barcode automatically. Keep the whole ISBN barcode inside the guide box, "
            "move the phone slightly farther away until the black lines separate clearly, avoid glare, and capture again."
        )
    except Exception as e:
        return None, f"Barcode scan failed: {e}"


def lookup_and_store_isbn(isbn):
    """Look up an ISBN and sync the Book Leveler search controls/results."""
    cleaned=re.sub(r"[^0-9Xx]","",str(isbn or ""))
    if len(cleaned) not in (10,13):
        return False, "A valid ISBN must contain 10 or 13 digits."
    found=openlibrary_lookup_isbn(cleaned)
    st.session_state["book_online_search_mode"]="ISBN"
    st.session_state["book_online_query"]=cleaned
    st.session_state["book_camera_isbn"]=cleaned
    st.session_state["book_online_results"]=[found] if found else []
    if found:
        return True, f"ISBN {cleaned} found."
    return False, f"ISBN {cleaned} was read correctly, but no matching Open Library book was found."

@st.cache_resource
def _chaplab_auto_barcode_component():
    component_path=Path(__file__).parent/"chaplab_auto_barcode_component"
    return components.declare_component("chaplab_auto_barcode", path=str(component_path))

def chaplab_auto_barcode_scanner(key):
    component=_chaplab_auto_barcode_component()
    return component(key=key, default="")

def openlibrary_search_books(query, limit=8):
    """Search Open Library by title, author, ISBN, or keywords."""
    query=str(query or "").strip()
    if not query:
        return []
    try:
        params={
            "q":query,
            "limit":max(1,min(int(limit),20)),
            "fields":"key,title,author_name,first_publish_year,isbn,cover_i,edition_count,publisher,language"
        }
        r=requests.get(
            "https://openlibrary.org/search.json",
            params=params,
            timeout=10,
            headers={"User-Agent":"ChapLabTeacherHub/3.2"}
        )
        r.raise_for_status()
        docs=r.json().get("docs",[])
        results=[]
        for d in docs:
            isbns=d.get("isbn") or []
            authors=d.get("author_name") or []
            publishers=d.get("publisher") or []
            results.append({
                "key":d.get("key") or "",
                "title":d.get("title") or "Untitled",
                "author":", ".join(authors[:3]),
                "year":d.get("first_publish_year") or "",
                "isbn":isbns[0] if isbns else "",
                "all_isbns":isbns[:10],
                "cover_i":d.get("cover_i"),
                "publisher":publishers[0] if publishers else "",
                "edition_count":d.get("edition_count") or 0,
            })
        return results
    except Exception as e:
        raise RuntimeError(f"Open Library search failed: {e}")

def openlibrary_lookup_isbn(isbn):
    """Look up a specific ISBN via Open Library Books API."""
    isbn=re.sub(r"[^0-9Xx]","",str(isbn or ""))
    if not isbn:
        return None
    try:
        r=requests.get(
            "https://openlibrary.org/api/books",
            params={
                "bibkeys":f"ISBN:{isbn}",
                "format":"json",
                "jscmd":"data"
            },
            timeout=10,
            headers={"User-Agent":"ChapLabTeacherHub/3.2"}
        )
        r.raise_for_status()
        data=r.json().get(f"ISBN:{isbn}")
        if not data:
            return None
        authors=[a.get("name","") for a in data.get("authors",[]) if a.get("name")]
        publishers=[p.get("name","") for p in data.get("publishers",[]) if p.get("name")]
        cover=data.get("cover") or {}
        return {
            "title":data.get("title") or "Untitled",
            "author":", ".join(authors[:3]),
            "isbn":isbn,
            "publisher":publishers[0] if publishers else "",
            "publish_date":data.get("publish_date") or "",
            "cover_url":cover.get("medium") or cover.get("large") or cover.get("small") or "",
            "url":data.get("url") or "",
        }
    except Exception as e:
        raise RuntimeError(f"Open Library ISBN lookup failed: {e}")



def lexile_hub_lookup_isbn(isbn):
    """Best-effort lookup against MetaMetrics/Lexile Find a Book pages."""
    isbn=re.sub(r"[^0-9Xx]","",str(isbn or ""))
    if len(isbn) not in (10,13):
        return {"found":False,"lexile":"","url":"","status":"Invalid ISBN"}

    urls=[
        f"https://hub.lexile.com/find-a-book/details/{isbn}/",
        f"https://fab-titles.lexile.com/",
    ]
    headers={
        "User-Agent":"Mozilla/5.0 (compatible; ChapLabTeacherHub/4.0)",
        "Accept-Language":"en-US,en;q=0.9"
    }

    # Direct detail page first.
    try:
        r=requests.get(urls[0],headers=headers,timeout=10,allow_redirects=True)
        if r.status_code==200 and r.text:
            plain=_scholastic_plain_text(r.text) if "_scholastic_plain_text" in globals() else re.sub(r"<[^>]+>"," ",r.text)
            # Prefer measures near explicit Lexile language.
            pats=[
                r"Lexile(?:®)?(?: Measure| Level)?\s*:?\s*((?:BR)?[0-9]{1,4}L)\b",
                r"\b((?:BR)?[0-9]{1,4}L)\b",
            ]
            for pat in pats:
                m=re.search(pat,plain,re.I)
                if m:
                    return {
                        "found":True,
                        "lexile":m.group(1).upper(),
                        "url":r.url,
                        "status":"Lexile Hub measure found"
                    }
            # ISBN detail page exists but measure may be rendered client-side.
            if isbn in re.sub(r"[^0-9Xx]","",plain):
                return {
                    "found":True,
                    "lexile":"",
                    "url":r.url,
                    "status":"Lexile Hub matched ISBN; measure not exposed in page HTML"
                }
    except Exception:
        pass

    # Legacy Find-a-Book landing page can sometimes expose search metadata.
    try:
        r=requests.get(
            urls[1],
            params={"q":isbn},
            headers=headers,
            timeout=10,
            allow_redirects=True
        )
        if r.status_code==200 and r.text:
            plain=re.sub(r"<[^>]+>"," ",r.text)
            plain=html.unescape(plain) if "html" in globals() else plain
            m=re.search(r"Lexile(?:®)?(?: Measure| Level)?\s*:?\s*((?:BR)?[0-9]{1,4}L)\b",plain,re.I)
            if m:
                return {
                    "found":True,
                    "lexile":m.group(1).upper(),
                    "url":r.url,
                    "status":"Lexile Find a Book measure found"
                }
    except Exception:
        pass

    return {
        "found":False,
        "lexile":"",
        "url":f"https://hub.lexile.com/find-a-book/details/{isbn}/",
        "status":"No Lexile measure found automatically"
    }

def normalize_lexile_measure(value):
    """Return numeric Lexile when possible; BR values are preserved separately."""
    raw=str(value or "").strip().upper().replace(" ","")
    if not raw:
        return None,raw
    if raw.startswith("BR"):
        # Beginning Reader measures aren't safely converted using a simple numeric crosswalk.
        return None,raw
    m=re.search(r"([0-9]{1,4})L?",raw)
    if not m:
        return None,raw
    return int(m.group(1)),f"{int(m.group(1))}L"

def estimate_fp_from_lexile(value):
    """
    Approximate F&P-style band from Lexile using a general published correlation crosswalk.
    This is intentionally labeled an estimate; Fountas & Pinnell does not publish
    an official Lexile conversion chart.
    """
    n,display=normalize_lexile_measure(value)
    if not display:
        return {"estimate":"","lexile":"","note":""}
    if n is None:
        return {
            "estimate":"A–B (very approximate)",
            "lexile":display,
            "note":"Beginning Reader (BR) Lexiles do not map cleanly to a single F&P level."
        }

    # Approximate bands adapted from commonly used text-level correlation charts.
    bands=[
        (0,50,"A–B"),
        (51,150,"C–E"),
        (151,200,"F–G"),
        (201,249,"H"),
        (250,299,"I"),
        (300,349,"J"),
        (350,399,"K"),
        (400,449,"L"),
        (450,499,"M"),
        (500,549,"N"),
        (550,599,"O"),
        (600,699,"P"),
        (700,749,"Q"),
        (750,799,"R"),
        (800,849,"S–T"),
        (850,899,"U–W"),
        (900,949,"X"),
        (950,999,"Y"),
        (1000,9999,"Z+"),
    ]
    estimate=""
    for lo,hi,label in bands:
        if lo<=n<=hi:
            estimate=label
            break
    return {
        "estimate":estimate,
        "lexile":display,
        "note":"Approximate crosswalk only — use the book's directly published Guided Reading/F&P level when available."
    }

def _fp_rank(level):
    """Approximate ordered rank for F&P letters A-Z; ranges use midpoint."""
    raw=str(level or "").strip().upper()
    letters=re.findall(r"[A-Z]",raw)
    if not letters:
        return None
    vals=[ord(x)-ord("A")+1 for x in letters]
    return sum(vals)/len(vals)

def _lexile_number(value):
    raw=str(value or "").strip().upper()
    if not raw or raw.startswith("BR"):
        return None
    m=re.search(r"([0-9]{1,4})",raw)
    return int(m.group(1)) if m else None

def evaluate_book_fit(book_fp="",book_lexile="",scholar_fp="",scholar_lexile=""):
    """
    Teacher-facing fit indicator. Prefer like-for-like direct level comparisons.
    This is guidance, not a replacement for teacher judgment or comprehension data.
    """
    bfp=_fp_rank(book_fp)
    sfp=_fp_rank(scholar_fp)
    blx=_lexile_number(book_lexile)
    slx=_lexile_number(scholar_lexile)

    # Prefer direct F&P/Guided Reading comparison when both exist.
    if bfp is not None and sfp is not None:
        gap=bfp-sfp
        if gap <= -3:
            return ("Likely Too Easy","Independent practice/review",
                    "This book is several F&P levels below the scholar's saved level.")
        if gap <= 1:
            return ("Good Fit","Independent Reading",
                    "This book is at or very close to the scholar's saved F&P level.")
        if gap <= 3:
            return ("Slightly Challenging","Instructional / With Support",
                    "This book is somewhat above the scholar's saved F&P level.")
        return ("Too Difficult Right Now","Read-aloud / Teacher Support",
                "This book is well above the scholar's saved F&P level for independent reading.")

    # Lexile comparison if both measures exist. Use broad, teacher-friendly bands.
    if blx is not None and slx is not None:
        gap=blx-slx
        if gap < -150:
            return ("Likely Too Easy","Independent practice/review",
                    "The book's Lexile is substantially below the scholar's saved Lexile.")
        if gap <= 75:
            return ("Good Fit","Independent Reading",
                    "The book's Lexile is close to the scholar's saved Lexile.")
        if gap <= 200:
            return ("Slightly Challenging","Instructional / With Support",
                    "The book's Lexile is above the scholar's saved Lexile but may work with support.")
        return ("Too Difficult Right Now","Read-aloud / Teacher Support",
                "The book's Lexile is substantially above the scholar's saved Lexile for independent reading.")

    return ("Need More Data","Teacher Review",
            "ChapLab needs a comparable saved scholar level and book level to make a fit recommendation.")

def _scholar_level_from_record(record):
    """Best-effort extraction from existing ChapLab scholar records without changing their schema."""
    if not isinstance(record,dict):
        return "",""
    fp_keys=["fp_level","f_and_p","f&p","fountas_pinnell","reading_level","guided_reading","gr_level"]
    lex_keys=["lexile","lexile_level","lexile_measure"]
    fp=""
    lx=""
    for k in fp_keys:
        if record.get(k) not in (None,""):
            fp=str(record.get(k)).strip().upper()
            break
    for k in lex_keys:
        if record.get(k) not in (None,""):
            lx=str(record.get(k)).strip().upper()
            break
    return fp,lx


def google_books_lookup_isbn(isbn):
    """Look up an exact ISBN through the public Google Books volumes API."""
    isbn=re.sub(r"[^0-9Xx]","",str(isbn or ""))
    if len(isbn) not in (10,13):
        return None
    try:
        r=requests.get(
            "https://www.googleapis.com/books/v1/volumes",
            params={"q":f"isbn:{isbn}","maxResults":5,"printType":"books"},
            timeout=10,
            headers={"User-Agent":"ChapLabTeacherHub/4.0"}
        )
        r.raise_for_status()
        items=r.json().get("items") or []
        if not items:
            return None

        chosen=None
        for item in items:
            vi=item.get("volumeInfo") or {}
            ids=vi.get("industryIdentifiers") or []
            values={re.sub(r"[^0-9Xx]","",str(x.get("identifier",""))) for x in ids}
            if isbn in values:
                chosen=item
                break
        chosen=chosen or items[0]
        vi=chosen.get("volumeInfo") or {}
        images=vi.get("imageLinks") or {}
        return {
            "google_id":chosen.get("id") or "",
            "title":vi.get("title") or "",
            "subtitle":vi.get("subtitle") or "",
            "author":", ".join(vi.get("authors") or []),
            "publisher":vi.get("publisher") or "",
            "publish_date":vi.get("publishedDate") or "",
            "description":vi.get("description") or "",
            "page_count":vi.get("pageCount") or "",
            "categories":vi.get("categories") or [],
            "cover_url":images.get("thumbnail") or images.get("smallThumbnail") or "",
            "info_link":vi.get("infoLink") or "",
            "isbn":isbn,
        }
    except Exception:
        return None

def _scholastic_plain_text(raw_html):
    raw_html=re.sub(r"(?is)<script.*?</script>"," ",raw_html or "")
    raw_html=re.sub(r"(?is)<style.*?</style>"," ",raw_html)
    raw_html=re.sub(r"(?s)<[^>]+>"," ",raw_html)
    raw_html=html.unescape(raw_html)
    return re.sub(r"\s+"," ",raw_html).strip()

def _parse_scholastic_levels(raw_html):
    """Extract only leveling information explicitly present in Scholastic HTML."""
    plain=_scholastic_plain_text(raw_html)
    levels={}
    patterns=[
        ("guided_reading",[
            r"Guided Reading Level\s*:?\s*([A-Z])\b",
            r"Guided Reading\s*:?\s*([A-Z])\b",
            r"GRL\s*:?\s*([A-Z])\b",
        ]),
        ("fountas_pinnell",[
            r"Fountas\s*(?:&|and)\s*Pinnell(?: Level)?\s*:?\s*([A-Z])\b",
            r"F&P(?: Level)?\s*:?\s*([A-Z])\b",
        ]),
        ("lexile",[
            r"Reading Level\s*:?\s*LEX\s*:?\s*([0-9]{2,4}L)\b",
            r"Lexile(?:®| Measure| Level)?\s*:?\s*([0-9]{2,4}L)\b",
            r"\b([0-9]{2,4}L)\b",
        ]),
        ("dra",[
            r"DRA(?: Level)?\s*:?\s*([0-9]{1,3})\b",
        ]),
        ("grade_level",[
            r"Grade Level(?: Equivalent)?\s*:?\s*([0-9K][0-9K.\-– ]{0,12})",
            r"Grades\s*:?\s*([K0-9]+\s*(?:-|–|to)\s*[K0-9]+)",
        ]),
    ]
    for key,plist in patterns:
        for pat in plist:
            m=re.search(pat,plain,re.I)
            if m:
                levels[key]=m.group(1).strip().upper() if key!="grade_level" else m.group(1).strip()
                break
    return levels

def scholastic_lookup_isbn(isbn,title=""):
    """Best-effort Scholastic lookup for explicitly published reading-level data."""
    isbn=re.sub(r"[^0-9Xx]","",str(isbn or ""))
    if len(isbn) not in (10,13):
        return {"found":False,"levels":{},"url":"","status":"Invalid ISBN"}

    attempts=[
        ("https://clubs.scholastic.com/search",{"q":isbn}),
        ("https://shop.scholastic.com/teachers-ecommerce/books/search-results.html",{"keyword":isbn}),
        ("https://www.scholastic.com/site/search.html",{"query":isbn}),
    ]
    headers={
        "User-Agent":"Mozilla/5.0 (compatible; ChapLabTeacherHub/4.0)",
        "Accept-Language":"en-US,en;q=0.9"
    }

    visited=[]
    candidate_pages=[]
    for url,params in attempts:
        try:
            r=requests.get(url,params=params,headers=headers,timeout=10,allow_redirects=True)
            if r.status_code!=200 or not r.text:
                continue
            visited.append(r.url)
            candidate_pages.append((r.url,r.text))

            for href in re.findall(r"""href=["']([^"']+)["']""",r.text,re.I):
                normalized_href=href.replace("-","")
                if isbn not in normalized_href:
                    continue
                full=urljoin(r.url,html.unescape(href))
                if "scholastic.com" in full and full not in visited:
                    try:
                        pr=requests.get(full,headers=headers,timeout=10,allow_redirects=True)
                        if pr.status_code==200 and pr.text:
                            visited.append(pr.url)
                            candidate_pages.insert(0,(pr.url,pr.text))
                    except Exception:
                        pass
        except Exception:
            continue

    for url,raw in candidate_pages:
        plain=_scholastic_plain_text(raw)
        compact=re.sub(r"[^0-9Xx]","",plain)
        levels=_parse_scholastic_levels(raw)
        title_match=True
        if title:
            title_match=title.lower() in plain.lower()
        if levels and (isbn in compact or title_match):
            return {
                "found":True,
                "levels":levels,
                "url":url,
                "status":"Scholastic level data found"
            }

    for url,raw in candidate_pages:
        plain=_scholastic_plain_text(raw)
        if isbn in re.sub(r"[^0-9Xx]","",plain):
            return {
                "found":True,
                "levels":{},
                "url":url,
                "status":"Scholastic matched the ISBN, but no level was visible"
            }

    return {
        "found":False,
        "levels":{},
        "url":"",
        "status":"No Scholastic level match found"
    }

def internet_book_lookup_isbn(isbn):
    """Combine Google Books, Open Library, and Scholastic into one book record."""
    isbn=re.sub(r"[^0-9Xx]","",str(isbn or ""))
    google=google_books_lookup_isbn(isbn)

    try:
        openlib=openlibrary_lookup_isbn(isbn)
    except Exception:
        openlib=None

    base={}
    for source in (google or {},openlib or {}):
        for k,v in source.items():
            if v not in (None,"",[],{}) and not base.get(k):
                base[k]=v

    if not base:
        base={"title":"Unknown title","author":"","isbn":isbn}
    base["isbn"]=isbn

    scholastic=scholastic_lookup_isbn(
        isbn,
        title=str(base.get("title") or "")
    )
    lexile_hub=lexile_hub_lookup_isbn(isbn)

    # Prefer an explicit Lexile from Lexile Hub. Fall back to Scholastic.
    scholastic_levels=scholastic.get("levels") or {}
    lexile_measure=(lexile_hub.get("lexile") or scholastic_levels.get("lexile") or "").strip().upper()
    fp_estimate=estimate_fp_from_lexile(lexile_measure) if lexile_measure else {
        "estimate":"","lexile":"","note":""
    }

    base["google_books"]=google or {}
    base["open_library"]=openlib or {}
    base["scholastic"]=scholastic
    base["lexile_hub"]=lexile_hub
    base["lexile_measure"]=lexile_measure
    base["estimated_fp_from_lexile"]=fp_estimate
    base["lookup_sources"]={
        "Lexile Find a Book":bool(lexile_hub.get("found")),
        "Scholastic":bool(scholastic.get("found")),
        "Google Books":bool(google),
        "Open Library":bool(openlib),
    }
    return base

def queue_book_isbn_lookup(isbn):
    """Queue widget fill + internet research for the next rerun."""
    cleaned=re.sub(r"[^0-9Xx]","",str(isbn or ""))
    if len(cleaned) not in (10,13):
        return False
    st.session_state["_book_pending_query"]=cleaned
    st.session_state["_book_pending_lookup_isbn"]=cleaned
    st.session_state["book_camera_isbn"]=cleaned
    st.session_state["_last_autolookup_isbn"]=cleaned
    return True

def openlibrary_cover_url(book):
    cover_i=book.get("cover_i") if isinstance(book,dict) else None
    if cover_i:
        return f"https://covers.openlibrary.org/b/id/{cover_i}-M.jpg"
    return ""

def fp_level_relation(student_level, book_level):
    """Compare two F&P letter levels A-Z without pretending non-letter levels."""
    if not student_level or not book_level:
        return None
    s=str(student_level).strip().upper()
    b=str(book_level).strip().upper()
    if not re.fullmatch(r"[A-Z]",s) or not re.fullmatch(r"[A-Z]",b):
        return None
    diff=ord(b)-ord(s)
    if diff <= -2:
        return ("Below current level",diff)
    if diff == -1:
        return ("Slightly below current level",diff)
    if diff == 0:
        return ("On current level",diff)
    if diff == 1:
        return ("Slightly above current level",diff)
    return ("Above current level",diff)



# ---------- Interim Assessment Import / Analysis ----------
def _norm_person_name(s):
    s=re.sub(r"[^A-Za-z0-9 ]+"," ",str(s or "").lower())
    return " ".join(s.split())

def _roster_name_keys(row):
    first=str(row.get("first_name","") or "").strip()
    last=str(row.get("last_name","") or "").strip()
    return {
        _norm_person_name(f"{first} {last}"),
        _norm_person_name(f"{last} {first}"),
        _norm_person_name(f"{last}, {first}"),
    }

def parse_interim_pdf(uploaded_file):
    if PdfReader is None:
        raise RuntimeError("PDF reader is not installed yet. Redeploy after adding pypdf to requirements.txt.")
    data=bytes(uploaded_file.getbuffer())
    reader=PdfReader(BytesIO(data))
    parsed=[]
    for page_num,page in enumerate(reader.pages,1):
        raw=page.extract_text() or ""
        if not raw.strip():
            continue
        flat=re.sub(r"[ \t]+"," ",raw)
        title_match=re.search(r"(?im)^\s*(\d+)\s+(ELA|MATH)\s+NY\s+INTERIM\s+(\d+)",raw)
        score_match=re.search(r"(?im)^\s*Score\s+([0-4](?:\.\d+)?)\s*$",raw)
        mc_match=re.search(r"Multiple Choice\s*:?\s*(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)",raw,re.I)
        cr_match=re.search(r"Constructed Response\s*:?\s*(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)",raw,re.I)
        date_match=re.search(r"Assessment Date\s*:\s*([0-9/\-]+)",raw,re.I)
        person_match=re.search(r"Assessment Date\s*:\s*[0-9/\-]+\s*\n\s*([^\n]+?)\s*\((\d+)\)",raw,re.I)
        if not person_match:
            person_match=re.search(r"Assessment Date\s*:\s*[0-9/\-]+\s+(.+?)\s*\((\d+)\)\s*Q\s*1",flat,re.I|re.S)
        if not score_match or not person_match:
            continue
        name=person_match.group(1).strip()
        student_id=person_match.group(2).strip()
        subject=(title_match.group(2).upper() if title_match else "").replace("MATH","Math").replace("ELA","ELA")
        interim_no=int(title_match.group(3)) if title_match else None

        questions=[]
        for m in re.finditer(r"Q\s*(\d+)\s*:\s*([A-D\-])\s*\(([A-D])\)",raw,re.I):
            q=int(m.group(1)); resp=m.group(2).upper(); correct=m.group(3).upper()
            questions.append({"question_number":q,"response":resp,"correct_answer":correct,"earned":1.0 if resp==correct else 0.0,"possible":1.0})
        for m in re.finditer(r"Q\s*(\d+)\s*:\s*(\d+(?:\.\d+)?)\s*/\s*(\d+(?:\.\d+)?)",raw,re.I):
            questions.append({"question_number":int(m.group(1)),"response":m.group(2),"correct_answer":"","earned":float(m.group(2)),"possible":float(m.group(3))})
        # Deduplicate by question number.
        qmap={q["question_number"]:q for q in questions}
        questions=[qmap[k] for k in sorted(qmap)]

        standards=[]
        if "STANDARD SUMMARY" in raw.upper():
            std_part=re.split(r"STANDARD SUMMARY",raw,flags=re.I,maxsplit=1)[1]
            starts=list(re.finditer(r"(?m)^\s*(\d+[A-Z]\d+)\s+",std_part))
            for i,m in enumerate(starts):
                code=m.group(1).strip()
                seg=std_part[m.end(): starts[i+1].start() if i+1<len(starts) else len(std_part)]
                normalized=" ".join(seg.split())
                tail=re.search(r"((?:\d{1,2}\s*,\s*)*\d{1,2})\s+([0-4](?:\.\d+)?)\s*$",normalized)
                if not tail:
                    continue
                qnums=tail.group(1).strip()
                std_score=float(tail.group(2))
                desc=normalized[:tail.start()].strip()
                standards.append({"standard_code":code,"standard_text":desc,"question_numbers":qnums,"standard_score":std_score})

        parsed.append({
            "page":page_num,"name":name,"student_id":student_id,"subject":subject,"interim_number":interim_no,
            "assessment_date":date_match.group(1) if date_match else "","overall_score":float(score_match.group(1)),
            "mc_earned":float(mc_match.group(1)) if mc_match else None,"mc_possible":float(mc_match.group(2)) if mc_match else None,
            "cr_earned":float(cr_match.group(1)) if cr_match else None,"cr_possible":float(cr_match.group(2)) if cr_match else None,
            "questions":questions,"standards":standards,"raw_text":raw
        })
    return parsed

def match_interim_results_to_roster(parsed, roster):
    by_id={str(r.get("student_id","") or "").strip():int(r.id) for _,r in roster.iterrows() if str(r.get("student_id","") or "").strip()}
    name_map={}
    for _,r in roster.iterrows():
        for key in _roster_name_keys(r):
            name_map.setdefault(key,int(r.id))
    rows=[]
    for item in parsed:
        sid=by_id.get(item["student_id"])
        method="School ID" if sid else ""
        if not sid:
            key=_norm_person_name(item["name"].replace(","," "))
            sid=name_map.get(key)
            if sid: method="Name"
        row=dict(item)
        row["scholar_id"]=sid
        row["match_method"]=method if sid else "Unmatched"
        rows.append(row)
    return rows

def _ensure_interim_assignment(class_id,subject,interim_number,assessment_date):
    title=f"{subject} Interim {interim_number}"
    c=conn()
    r=c.execute("SELECT id FROM assignments WHERE class_id=? AND subject=? AND title=? ORDER BY id DESC LIMIT 1",(int(class_id),subject,title)).fetchone()
    if r:
        aid=int(r["id"])
    else:
        mp=quarter_for_date(assessment_date,current_academic_year()) or "Quarter 1"
        cur=c.execute("""INSERT INTO assignments(title,subject,category,standard_code,points_possible,assignment_date,class_id,marking_period)
                         VALUES (?,?,?,?,?,?,?,?)""",(title,subject,"Assessment","",4.0,assessment_date or str(date.today()),int(class_id),mp))
        aid=int(cur.lastrowid)
    c.commit(); c.close(); return aid

def save_interim_import(class_id,subject,interim_number,source_file,matched_rows):
    valid=[r for r in matched_rows if r.get("scholar_id")]
    if not valid:
        return 0,0,None
    assessment_date=next((r.get("assessment_date") for r in valid if r.get("assessment_date")),str(date.today()))
    aid=_ensure_interim_assignment(class_id,subject,interim_number,assessment_date)
    c=conn()
    c.execute("""INSERT INTO interim_assessments(class_id,subject,interim_number,title,assessment_date,source_file,assignment_id,imported_at)
                 VALUES (?,?,?,?,?,?,?,?)
                 ON CONFLICT(class_id,subject,interim_number) DO UPDATE SET
                 title=excluded.title,assessment_date=excluded.assessment_date,source_file=excluded.source_file,
                 assignment_id=excluded.assignment_id,imported_at=excluded.imported_at""",
              (int(class_id),subject,int(interim_number),f"{subject} Interim {interim_number}",assessment_date,source_file,aid,datetime.now().isoformat(timespec="seconds")))
    assessment_id=int(c.execute("SELECT id FROM interim_assessments WHERE class_id=? AND subject=? AND interim_number=?",(int(class_id),subject,int(interim_number))).fetchone()["id"])
    saved=0; skipped=0
    for r in matched_rows:
        sid=r.get("scholar_id")
        if not sid:
            skipped+=1; continue
        c.execute("""INSERT INTO interim_results(assessment_id,scholar_id,school_student_id,overall_score,mc_earned,mc_possible,cr_earned,cr_possible,raw_text)
                     VALUES (?,?,?,?,?,?,?,?,?)
                     ON CONFLICT(assessment_id,scholar_id) DO UPDATE SET
                     school_student_id=excluded.school_student_id,overall_score=excluded.overall_score,
                     mc_earned=excluded.mc_earned,mc_possible=excluded.mc_possible,cr_earned=excluded.cr_earned,
                     cr_possible=excluded.cr_possible,raw_text=excluded.raw_text""",
                  (assessment_id,int(sid),r.get("student_id",""),r.get("overall_score"),r.get("mc_earned"),r.get("mc_possible"),r.get("cr_earned"),r.get("cr_possible"),r.get("raw_text","")))
        c.execute("DELETE FROM interim_standard_scores WHERE assessment_id=? AND scholar_id=?",(assessment_id,int(sid)))
        for s in r.get("standards",[]):
            c.execute("""INSERT OR REPLACE INTO interim_standard_scores(assessment_id,scholar_id,standard_code,standard_text,question_numbers,standard_score)
                         VALUES (?,?,?,?,?,?)""",(assessment_id,int(sid),s.get("standard_code",""),s.get("standard_text",""),s.get("question_numbers",""),s.get("standard_score")))
        c.execute("DELETE FROM interim_question_results WHERE assessment_id=? AND scholar_id=?",(assessment_id,int(sid)))
        for q in r.get("questions",[]):
            c.execute("""INSERT OR REPLACE INTO interim_question_results(assessment_id,scholar_id,question_number,response,correct_answer,earned,possible)
                         VALUES (?,?,?,?,?,?,?)""",(assessment_id,int(sid),int(q["question_number"]),q.get("response",""),q.get("correct_answer",""),q.get("earned"),q.get("possible")))
        c.execute("""INSERT INTO grades(scholar_id,assignment_id,points_earned) VALUES (?,?,?)
                     ON CONFLICT(scholar_id,assignment_id) DO UPDATE SET points_earned=excluded.points_earned""",(int(sid),aid,float(r.get("overall_score") or 0)))
        saved+=1
    c.commit(); c.close()
    return saved,skipped,assessment_id

def interim_assessment_record(class_id,subject,interim_number):
    if not class_id: return None
    c=conn(); r=c.execute("SELECT * FROM interim_assessments WHERE class_id=? AND subject=? AND interim_number=?",(int(class_id),subject,int(interim_number))).fetchone(); c.close(); return r

def interim_results_df(assessment_id):
    if not assessment_id: return pd.DataFrame()
    c=conn(); df=pd.read_sql_query("""SELECT r.*,s.first_name,s.last_name FROM interim_results r
                                      JOIN scholars s ON s.id=r.scholar_id WHERE r.assessment_id=? ORDER BY s.last_name,s.first_name""",c,params=[int(assessment_id)]); c.close()
    if not df.empty: df["Scholar"]=df["first_name"].fillna("")+" "+df["last_name"].fillna("")
    return df

def interim_standard_summary_df(assessment_id):
    if not assessment_id: return pd.DataFrame()
    c=conn(); df=pd.read_sql_query("SELECT * FROM interim_standard_scores WHERE assessment_id=?",c,params=[int(assessment_id)]); c.close()
    if df.empty: return pd.DataFrame()
    rows=[]
    for code,g in df.groupby("standard_code"):
        vals=pd.to_numeric(g["standard_score"],errors="coerce").dropna()
        if vals.empty: continue
        textv=str(g.iloc[0]["standard_text"] or "")
        qnums=str(g.iloc[0]["question_numbers"] or "")
        prof=float((vals>=3.0).mean()*100)
        below=float((vals<3.0).mean()*100)
        rows.append({"Standard":code,"Standard Text":textv,"Question #s":qnums,"Average":round(vals.mean(),2),"% 3.0+":round(prof,1),"% Below 3.0":round(below,1),"N":len(vals)})
    return pd.DataFrame(rows).sort_values(["% Below 3.0","Average"],ascending=[False,True]) if rows else pd.DataFrame()

def interim_question_summary_df(assessment_id):
    if not assessment_id: return pd.DataFrame()
    c=conn(); df=pd.read_sql_query("SELECT * FROM interim_question_results WHERE assessment_id=?",c,params=[int(assessment_id)]); c.close()
    if df.empty: return pd.DataFrame()
    rows=[]
    for q,g in df.groupby("question_number"):
        earned=pd.to_numeric(g["earned"],errors="coerce"); possible=pd.to_numeric(g["possible"],errors="coerce")
        pct=(earned.sum()/possible.sum()*100) if possible.sum()>0 else None
        rows.append({"Question":int(q),"% Earned":round(pct,1) if pct is not None else None,"N":len(g)})
    return pd.DataFrame(rows).sort_values("Question")

def interim_potential_risers(results):
    if results.empty: return []
    vals=pd.to_numeric(results["overall_score"],errors="coerce")
    return results[(vals-2.5).abs()<0.001]["Scholar"].tolist()

def interim_proficiency(results):
    if results.empty: return None
    vals=pd.to_numeric(results["overall_score"],errors="coerce").dropna()
    return float((vals>=3.0).mean()*100) if not vals.empty else None

def get_interim_goal(class_id,subject,interim_number):
    if not class_id: return None
    c=conn(); r=c.execute("SELECT proficiency_goal FROM interim_goals WHERE class_id=? AND subject=? AND interim_number=?",(int(class_id),subject,int(interim_number))).fetchone(); c.close()
    return float(r["proficiency_goal"]) if r and r["proficiency_goal"] is not None else None

def save_interim_goal(class_id,subject,interim_number,goal):
    c=conn(); c.execute("""INSERT INTO interim_goals(class_id,subject,interim_number,proficiency_goal) VALUES (?,?,?,?)
                         ON CONFLICT(class_id,subject,interim_number) DO UPDATE SET proficiency_goal=excluded.proficiency_goal""",(int(class_id),subject,int(interim_number),float(goal))); c.commit(); c.close()

def interim_copy_blocks(class_id,subject,interim_number,assessment_id):
    cname=class_name_from_id(class_id) or "Class"
    results=interim_results_df(assessment_id)
    std=interim_standard_summary_df(assessment_id)
    qdf=interim_question_summary_df(assessment_id)
    prof=interim_proficiency(results)
    risers=interim_potential_risers(results)
    next_no=min(3,int(interim_number)+1)
    next_goal=get_interim_goal(class_id,subject,next_no)
    goal_text=(f"{next_goal:.0f}%" if next_goal is not None else "[enter next interim goal]")
    goals=(f"{cname} | {subject} Interim {interim_number}\n"
           f"Proficiency % (3.0/B and above): {prof:.0f}%\n" if prof is not None else f"{cname} | {subject} Interim {interim_number}\nProficiency %: No data\n")
    goals+=f"Potential Risers 2.5 (C): {', '.join(risers) if risers else 'None'}\n"
    if interim_number<3: goals+=f"Interim {next_no} {subject} Goal: {goal_text}\n"

    strengths=std[std["% 3.0+"]>=70] if not std.empty else pd.DataFrame()
    reteach=std[std["% Below 3.0"]>=60] if not std.empty else pd.DataFrame()
    sec1_strength="; ".join([f"{r.Standard} ({r['% 3.0+']:.0f}% at 3.0+)" for _,r in strengths.iterrows()]) or "No standard met the 70% strength threshold."
    sec1_growth="; ".join([f"{r.Standard} ({r['% Below 3.0']:.0f}% below 3.0)" for _,r in reteach.iterrows()]) or "No standard met the 60% whole-group reteach threshold."
    lines=["SECTION I: HIGH-LEVEL OBSERVATION",f"{cname} Strengths: {sec1_strength}",f"{cname} Growth Areas: {sec1_growth}","", "SECTION 2: WHOLE GROUP RETEACH"]
    for _,r in reteach.iterrows():
        desc=str(r["Standard Text"] or "").strip()
        why=f"Scholars need additional practice with {desc[:220] if desc else r.Standard}."
        plan=f"Model the skill with a think-aloud, underline/key evidence, complete guided practice, then spiral {r.Standard} in Do Now/weekly review and small groups."
        lines += [f"Standard: {r.Standard}",f"Question #s: {r['Question #s']}",f"Why did students struggle?: {why}",f"Instructional Plan: {plan}",""]
    if reteach.empty: lines.append("No standard currently meets the 60% below-3.0 whole-group reteach rule.\n")

    # Individual major concerns = below 2.5 overall; include their weakest standards.
    lines += ["SECTION 3: INDIVIDUAL / SMALL GROUP REVIEW"]
    if not results.empty:
        low=results[pd.to_numeric(results["overall_score"],errors="coerce")<2.5]
        if low.empty:
            lines.append("Students of Major Concern: None below 2.5 overall.")
        else:
            c=conn()
            for _,rr in low.iterrows():
                weak=pd.read_sql_query("""SELECT standard_code,standard_score FROM interim_standard_scores
                                           WHERE assessment_id=? AND scholar_id=? AND standard_score<3 ORDER BY standard_score,standard_code""",c,params=[int(assessment_id),int(rr.scholar_id)])
                weak_list=", ".join(weak.standard_code.head(5).tolist()) if not weak.empty else "Review overall assessment performance"
                lines.append(f"{rr.Scholar}: score {float(rr.overall_score):g}; most help with {weak_list}; support during workshop/Do Now/small group.")
            c.close()
    lines += ["", "SECTION 4: DEFINING NEXT STEPS — SPIRAL REVIEW"]
    if not std.empty:
        for i,(_,r) in enumerate(std.head(4).iterrows(),1):
            lines.append(f"Week {i} Standards for Review: {r.Standard} — {str(r['Standard Text'])[:180]}")
            lines.append(f"Week {i} New Standards: [teacher enters upcoming standard]")
    else:
        for i in range(1,5): lines.append(f"Week {i} Standards for Review: [data will populate after import]")
    analysis="\n".join(lines)
    return goals,analysis

def render_interim_center(selected_scope_class=None):
    st.markdown("## 📝 Interims")
    st.caption("Upload the class interim PDF once. ChapLab reads each scholar page, saves the 0–4 overall score to Grades, keeps the standard/question data for analysis, and builds copy-ready text for the school goal and data-day documents.")
    classes=classes_df()
    if classes.empty:
        st.info("The Interim workspace is ready. Add a class and roster before importing a PDF so ChapLab can match scholar names/IDs and send overall scores to Grades.")
        c1,c2,c3=st.columns(3)
        c1.selectbox("Class",["No classes yet"],disabled=True,key="interim_no_class")
        c2.radio("Subject",["ELA","Math"],horizontal=True,key="interim_subject_empty")
        c3.radio("Interim",[1,2,3],horizontal=True,key="interim_num_empty")
        st.file_uploader("Upload interim results PDF",type=["pdf"],disabled=True,key="interim_pdf_empty")
        st.markdown("### Goals & Potential Risers")
        st.text_area("Copy into the Interim Goals document",value="Class proficiency, 2.5 potential risers, and next-interim goal will appear here.",height=130,disabled=True,key="interim_goals_empty")
        st.markdown("### Assessment Analysis & Action Plan")
        st.text_area("Copy into the Data Day Analysis document",value="Strengths, whole-group reteach standards, question numbers, major concerns, small-group needs, and four-week spiral standards will appear here.",height=260,disabled=True,key="interim_analysis_empty")
        return

    ids=list(classes.id.astype(int))
    default=selected_scope_class if selected_scope_class in ids else ids[0]
    top1,top2,top3=st.columns([2,1,1])
    class_id=top1.selectbox("Class",ids,index=ids.index(default),format_func=lambda x:classes[classes.id==x].iloc[0].class_name,key="interim_class")
    subject=top2.radio("Subject",["ELA","Math"],horizontal=True,key="interim_subject")
    interim_no=top3.radio("Interim",[1,2,3],horizontal=True,key="interim_number")

    st.markdown("### Import PDF Results")
    uploaded=st.file_uploader("Upload the class interim PDF",type=["pdf"],key="interim_pdf_upload")
    if uploaded:
        try:
            parsed=parse_interim_pdf(uploaded)
            roster=scholars_df(class_id)
            matched=match_interim_results_to_roster(parsed,roster)
            review=[]
            for r in matched:
                review.append({"Page":r["page"],"PDF Scholar":r["name"],"School ID":r["student_id"],"Match":r["match_method"],"Score":r["overall_score"],"MC":f"{_fmt_score(r['mc_earned'])}/{_fmt_score(r['mc_possible'])}","CR":f"{_fmt_score(r['cr_earned'])}/{_fmt_score(r['cr_possible'])}","Standards":len(r.get("standards",[]))})
            st.dataframe(pd.DataFrame(review),hide_index=True,use_container_width=True)
            unmatched=sum(1 for r in matched if not r.get("scholar_id"))
            if unmatched: st.warning(f"{unmatched} PDF scholar page(s) are unmatched. Check the roster names/School IDs before saving.")
            if st.button("Import Interim & Send Overall Scores to Grades",type="primary",key="save_interim_pdf"):
                saved,skipped,assessment_id=save_interim_import(class_id,subject,interim_no,uploaded.name,matched)
                st.success(f"Imported {saved} scholars. {skipped} unmatched/skipped. {subject} Interim {interim_no} overall scores were entered in Grades as points out of 4.")
                st.session_state["last_interim_assessment_id"]=assessment_id
                st.rerun()
        except Exception as e:
            st.error(f"Could not read the interim PDF: {e}")

    rec=interim_assessment_record(class_id,subject,interim_no)
    if not rec:
        st.info(f"No saved {subject} Interim {interim_no} results yet. The analysis areas below will populate after you import the PDF.")
        results=pd.DataFrame(); std=pd.DataFrame(); qdf=pd.DataFrame(); assessment_id=None
    else:
        assessment_id=int(rec["id"]); results=interim_results_df(assessment_id); std=interim_standard_summary_df(assessment_id); qdf=interim_question_summary_df(assessment_id)
        st.markdown(f"### {subject} Interim {interim_no} Class Snapshot")
        prof=interim_proficiency(results); risers=interim_potential_risers(results)
        m1,m2,m3,m4=st.columns(4)
        m1.metric("Scholars Imported",len(results))
        m2.metric("Proficiency (3.0+)",f"{prof:.0f}%" if prof is not None else "—")
        m3.metric("Potential Risers (2.5)",len(risers))
        m4.metric("Class Avg Score",f"{pd.to_numeric(results['overall_score'],errors='coerce').mean():.2f}" if not results.empty else "—")
        if risers: st.write("**Potential Risers:** "+", ".join(risers))
        if not std.empty:
            st.markdown("#### Standards Analysis")
            st.dataframe(std,hide_index=True,use_container_width=True)
        if not qdf.empty:
            st.markdown("#### Question Analysis")
            st.dataframe(qdf,hide_index=True,use_container_width=True)

    if interim_no<3:
        next_no=interim_no+1
        existing_goal=get_interim_goal(class_id,subject,next_no)
        goal=st.number_input(f"Interim {next_no} {subject} Proficiency Goal (%)",0.0,100.0,float(existing_goal) if existing_goal is not None else 50.0,1.0,key=f"goal_{subject}_{next_no}")
        if st.button(f"Save Interim {next_no} Goal",key=f"save_goal_{subject}_{next_no}"):
            save_interim_goal(class_id,subject,next_no,goal); st.success("Goal saved."); st.rerun()

    goals,analysis=interim_copy_blocks(class_id,subject,interim_no,assessment_id) if assessment_id else (f"{class_name_from_id(class_id)} | {subject} Interim {interim_no}\nProficiency %: [imports will populate]\nPotential Risers 2.5 (C): [imports will populate]", "Import the interim PDF to generate the Assessment Analysis & Action Plan text.")
    st.markdown("---")
    st.markdown("### 📋 Copy/Paste for Required School Documents")
    st.markdown("#### Interim 2 & 3 Goals Document")
    st.text_area("Copy this into the Goals document",value=goals,height=170,key="interim_goals_copy")
    st.markdown("#### Assessment Analysis & Action Plan")
    st.text_area("Copy this into the Data Day Analysis document",value=analysis,height=520,key="interim_analysis_copy")
# ---------- Assessment Data Center ----------
def assessment_roster(class_id=None):
    """Use selected class when available; otherwise use all active scholars."""
    return scholars_df(class_id if class_id else None)

def _safe_num(v):
    try:
        s=str(v or "").strip()
        return float(s) if s else None
    except:
        return None

def _fmt_score(v):
    n=_safe_num(v)
    if n is None:
        return "—"
    return f"{n:g}"

def iready_for_scholar(sid):
    c=conn()
    r=c.execute("SELECT * FROM iready_scores WHERE scholar_id=?",(int(sid),)).fetchone()
    c.close()
    return r

def assessment_scope_label(class_id):
    if class_id:
        name=class_name_from_id(int(class_id))
        return name or "Selected class"
    return "All scholars"

def _assessment_subject_keys(prefix, subject):
    side="reading" if subject=="Reading" else "math"
    return (
        f"{prefix}_fall_{side}" if prefix=="nwea" else f"fall_{side}",
        f"{prefix}_winter_{side}" if prefix=="nwea" else f"winter_{side}",
        f"{prefix}_spring_{side}" if prefix=="nwea" else f"spring_{side}",
        f"{prefix}_{side}_goal" if prefix=="nwea" else f"{side}_goal"
    )

def nwea_scope_dataframe(class_id, subject):
    roster=assessment_roster(class_id)
    side="reading" if subject=="Reading" else "math"
    rows=[]
    for _,sr in roster.iterrows():
        br=benchmark_for_scholar(int(sr.id))
        fall=winter=spring=goal=None
        if br:
            fall=_safe_num(br[f"nwea_fall_{side}"]) if f"nwea_fall_{side}" in br.keys() else None
            winter=_safe_num(br[f"nwea_winter_{side}"]) if f"nwea_winter_{side}" in br.keys() else None
            spring=_safe_num(br[f"nwea_spring_{side}"]) if f"nwea_spring_{side}" in br.keys() else None
            goal=_safe_num(br[f"nwea_{side}_goal"]) if f"nwea_{side}_goal" in br.keys() else None
        latest=spring if spring is not None else winter if winter is not None else fall
        growth=(spring-fall) if spring is not None and fall is not None else None
        rows.append({"Scholar":nm(sr),"Fall":fall,"Winter":winter,"Spring":spring,"Goal":goal,"Latest":latest,"Growth":growth})
    return pd.DataFrame(rows)

def iready_scope_dataframe(class_id, subject):
    roster=assessment_roster(class_id)
    side="reading" if subject=="Reading" else "math"
    rows=[]
    for _,sr in roster.iterrows():
        br=iready_for_scholar(int(sr.id))
        fall=winter=spring=goal=None
        if br:
            fall=_safe_num(br[f"fall_{side}"])
            winter=_safe_num(br[f"winter_{side}"])
            spring=_safe_num(br[f"spring_{side}"])
            goal=_safe_num(br[f"{side}_goal"])
        latest=spring if spring is not None else winter if winter is not None else fall
        growth=(spring-fall) if spring is not None and fall is not None else None
        rows.append({"Scholar":nm(sr),"Fall":fall,"Winter":winter,"Spring":spring,"Goal":goal,"Latest":latest,"Growth":growth})
    return pd.DataFrame(rows)

def assessment_high_low_growth(df):
    if df.empty:
        return None,None,None
    scored=df[df["Latest"].notna()]
    high=low=None
    if not scored.empty:
        hi=scored["Latest"].max()
        lo=scored["Latest"].min()
        high=(scored[scored["Latest"]==hi]["Scholar"].tolist(),hi)
        low=(scored[scored["Latest"]==lo]["Scholar"].tolist(),lo)
    grew=df[df["Growth"].notna()]
    growth=None
    if not grew.empty:
        mx=grew["Growth"].max()
        growth=(grew[grew["Growth"]==mx]["Scholar"].tolist(),mx)
    return high,low,growth

def render_assessment_insights(df, subject, label):
    st.markdown("### Insights")
    hi,lo,growth=assessment_high_low_growth(df)
    a,b,c=st.columns(3)
    with a:
        st.markdown("**⭐ Highest Score**")
        if hi:
            st.success(f"{', '.join(hi[0])} — {hi[1]:g}")
        else:
            st.caption("Will populate after scores are entered.")
    with b:
        st.markdown("**↘ Lowest Score**")
        if lo:
            st.warning(f"{', '.join(lo[0])} — {lo[1]:g}")
        else:
            st.caption("Will populate after scores are entered.")
    with c:
        st.markdown("**📈 Highest Growth (Fall → Spring)**")
        if growth:
            st.success(f"{', '.join(growth[0])} — +{growth[1]:g}")
        else:
            st.caption("Will populate after Fall and Spring scores are entered.")

    st.caption(f"{label} insights shown for {subject}. Highest/lowest use the latest available score (Spring, then Winter, then Fall).")
    show=df[["Scholar","Fall","Winter","Spring","Goal","Growth"]].copy() if not df.empty else pd.DataFrame(columns=["Scholar","Fall","Winter","Spring","Goal","Growth"])
    st.dataframe(show,hide_index=True,use_container_width=True)

def render_nwea_center(class_id=None):
    st.markdown("## 📊 NWEA MAP Growth")
    st.caption(f"Scope: **{assessment_scope_label(class_id)}** · Track Reading and Math RIT scores, goals, rankings, and Fall-to-Spring growth.")

    roster=assessment_roster(class_id)
    subject=st.radio("Subject",["Reading","Math"],horizontal=True,key="nwea_subject_separate")
    side="reading" if subject=="Reading" else "math"

    if roster.empty:
        st.selectbox("Scholar",["No scholars yet"],disabled=True,key="nwea_empty_scholar")
        g1,g2,g3,g4=st.columns(4)
        g1.text_input("Goal (End-of-Year RIT)",disabled=True,key="nwea_empty_goal")
        g2.text_input("Fall RIT",disabled=True,key="nwea_empty_fall")
        g3.text_input("Winter RIT",disabled=True,key="nwea_empty_winter")
        g4.text_input("Spring RIT",disabled=True,key="nwea_empty_spring")
        st.button("Save NWEA Scores",disabled=True,key="nwea_empty_save")
        render_assessment_insights(pd.DataFrame(columns=["Scholar","Fall","Winter","Spring","Goal","Latest","Growth"]),subject,"NWEA")
        st.info("Add scholars in the Scholars section. Their names will automatically appear in the Scholar dropdown.")
        return

    sid=st.selectbox("Scholar",list(roster.id.astype(int)),format_func=lambda x:nm(roster[roster.id==x].iloc[0]),key="nwea_scholar_separate")
    br=benchmark_for_scholar(sid)
    def bv(k):
        return str(br[k] or "") if br and k in br.keys() else ""

    with st.form("nwea_separate_form"):
        goal=st.text_input("Goal (End-of-Year RIT)",value=bv(f"nwea_{side}_goal"),placeholder="Example: 195")
        a,b,c=st.columns(3)
        fall=a.text_input("Fall RIT",value=bv(f"nwea_fall_{side}"),placeholder="RIT score")
        winter=b.text_input("Winter RIT",value=bv(f"nwea_winter_{side}"),placeholder="RIT score")
        spring=c.text_input("Spring RIT",value=bv(f"nwea_spring_{side}"),placeholder="RIT score")
        if st.form_submit_button("Save NWEA Scores"):
            other="math" if side=="reading" else "reading"
            existing=benchmark_for_scholar(sid)
            def ev(k):
                return str(existing[k] or "") if existing and k in existing.keys() else ""
            vals={
                "nwea_fall_reading":fall if side=="reading" else ev("nwea_fall_reading"),
                "nwea_winter_reading":winter if side=="reading" else ev("nwea_winter_reading"),
                "nwea_spring_reading":spring if side=="reading" else ev("nwea_spring_reading"),
                "nwea_reading_goal":goal if side=="reading" else ev("nwea_reading_goal"),
                "nwea_fall_math":fall if side=="math" else ev("nwea_fall_math"),
                "nwea_winter_math":winter if side=="math" else ev("nwea_winter_math"),
                "nwea_spring_math":spring if side=="math" else ev("nwea_spring_math"),
                "nwea_math_goal":goal if side=="math" else ev("nwea_math_goal"),
                "fp_fall_level":ev("fp_fall_level"),
                "fp_winter_level":ev("fp_winter_level"),
                "fp_spring_level":ev("fp_spring_level"),
                "fp_fall_word_list":ev("fp_fall_word_list"),
                "fp_winter_word_list":ev("fp_winter_word_list"),
                "fp_spring_word_list":ev("fp_spring_word_list"),
                "notes":ev("notes"),
            }
            cdb=conn()
            cdb.execute("""INSERT INTO benchmark_scores(
                scholar_id,nwea_fall_reading,nwea_winter_reading,nwea_spring_reading,
                nwea_fall_math,nwea_winter_math,nwea_spring_math,nwea_reading_goal,nwea_math_goal,
                fp_fall_level,fp_winter_level,fp_spring_level,fp_fall_word_list,fp_winter_word_list,fp_spring_word_list,notes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(scholar_id) DO UPDATE SET
                nwea_fall_reading=excluded.nwea_fall_reading,
                nwea_winter_reading=excluded.nwea_winter_reading,
                nwea_spring_reading=excluded.nwea_spring_reading,
                nwea_fall_math=excluded.nwea_fall_math,
                nwea_winter_math=excluded.nwea_winter_math,
                nwea_spring_math=excluded.nwea_spring_math,
                nwea_reading_goal=excluded.nwea_reading_goal,
                nwea_math_goal=excluded.nwea_math_goal,
                fp_fall_level=excluded.fp_fall_level,
                fp_winter_level=excluded.fp_winter_level,
                fp_spring_level=excluded.fp_spring_level,
                fp_fall_word_list=excluded.fp_fall_word_list,
                fp_winter_word_list=excluded.fp_winter_word_list,
                fp_spring_word_list=excluded.fp_spring_word_list,
                notes=excluded.notes""",
                (sid,vals["nwea_fall_reading"],vals["nwea_winter_reading"],vals["nwea_spring_reading"],
                 vals["nwea_fall_math"],vals["nwea_winter_math"],vals["nwea_spring_math"],vals["nwea_reading_goal"],vals["nwea_math_goal"],
                 vals["fp_fall_level"],vals["fp_winter_level"],vals["fp_spring_level"],vals["fp_fall_word_list"],vals["fp_winter_word_list"],vals["fp_spring_word_list"],vals["notes"]))
            cdb.commit(); cdb.close()
            st.success("NWEA scores saved.")
            st.rerun()

    render_assessment_insights(nwea_scope_dataframe(class_id,subject),subject,"NWEA")

def render_fp_center(class_id=None):
    st.markdown("## 📚 F&P Reading Levels")
    st.caption(f"Scope: **{assessment_scope_label(class_id)}** · F&P is separate from NWEA and can be entered at any time.")

    roster=assessment_roster(class_id)
    if roster.empty:
        st.selectbox("Scholar",["No scholars yet"],disabled=True,key="fp_empty_scholar")
        a,b,c=st.columns(3)
        a.text_input("Fall F&P Level",disabled=True,key="fp_empty_fall")
        b.text_input("Winter / Midyear F&P Level",disabled=True,key="fp_empty_winter")
        c.text_input("Spring F&P Level",disabled=True,key="fp_empty_spring")
        w1,w2,w3=st.columns(3)
        w1.text_input("Fall Word List Level / Score",disabled=True,key="fp_empty_word1")
        w2.text_input("Winter Word List Level / Score",disabled=True,key="fp_empty_word2")
        w3.text_input("Spring Word List Level / Score",disabled=True,key="fp_empty_word3")
        st.button("Save F&P Data",disabled=True,key="fp_empty_save")
        st.info("Add scholars in the Scholars section. Their names will automatically populate here.")
        return

    sid=st.selectbox("Scholar",list(roster.id.astype(int)),format_func=lambda x:nm(roster[roster.id==x].iloc[0]),key="fp_scholar_separate")
    br=benchmark_for_scholar(sid)
    def bv(k):
        return str(br[k] or "") if br and k in br.keys() else ""

    with st.form("fp_separate_form"):
        a,b,c=st.columns(3)
        fall=a.text_input("Fall F&P Level",value=bv("fp_fall_level"),placeholder="Example: L")
        winter=b.text_input("Winter / Midyear F&P Level",value=bv("fp_winter_level"),placeholder="Example: M")
        spring=c.text_input("Spring F&P Level",value=bv("fp_spring_level"),placeholder="Example: N")
        w1,w2,w3=st.columns(3)
        wf=w1.text_input("Fall Word List Level / Score",value=bv("fp_fall_word_list"))
        ww=w2.text_input("Winter Word List Level / Score",value=bv("fp_winter_word_list"))
        ws=w3.text_input("Spring Word List Level / Score",value=bv("fp_spring_word_list"))
        notes=st.text_area("F&P Notes",value=bv("notes"))
        if st.form_submit_button("Save F&P Data"):
            cdb=conn()
            existing=benchmark_for_scholar(sid)
            def ev(k):
                return str(existing[k] or "") if existing and k in existing.keys() else ""
            cdb.execute("""INSERT INTO benchmark_scores(
                scholar_id,nwea_fall_reading,nwea_winter_reading,nwea_spring_reading,
                nwea_fall_math,nwea_winter_math,nwea_spring_math,nwea_reading_goal,nwea_math_goal,
                fp_fall_level,fp_winter_level,fp_spring_level,fp_fall_word_list,fp_winter_word_list,fp_spring_word_list,notes)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(scholar_id) DO UPDATE SET
                fp_fall_level=excluded.fp_fall_level,
                fp_winter_level=excluded.fp_winter_level,
                fp_spring_level=excluded.fp_spring_level,
                fp_fall_word_list=excluded.fp_fall_word_list,
                fp_winter_word_list=excluded.fp_winter_word_list,
                fp_spring_word_list=excluded.fp_spring_word_list,
                notes=excluded.notes""",
                (sid,ev("nwea_fall_reading"),ev("nwea_winter_reading"),ev("nwea_spring_reading"),
                 ev("nwea_fall_math"),ev("nwea_winter_math"),ev("nwea_spring_math"),ev("nwea_reading_goal"),ev("nwea_math_goal"),
                 fall,winter,spring,wf,ww,ws,notes))
            cdb.commit(); cdb.close()
            st.success("F&P data saved.")
            st.rerun()

    # F&P class/all-scholar snapshot.
    rows=[]
    for _,sr in roster.iterrows():
        r=benchmark_for_scholar(int(sr.id))
        rows.append({
            "Scholar":nm(sr),
            "Fall":str(r["fp_fall_level"] or "") if r else "",
            "Winter / Midyear":str(r["fp_winter_level"] or "") if r and "fp_winter_level" in r.keys() else "",
            "Spring":str(r["fp_spring_level"] or "") if r else "",
        })
    st.markdown("### F&P Snapshot")
    st.dataframe(pd.DataFrame(rows),hide_index=True,use_container_width=True)

def render_iready_center(class_id=None):
    st.markdown("## 🟣 I‑Ready Diagnostics")
    st.caption(f"Scope: **{assessment_scope_label(class_id)}** · Track Reading and Math diagnostic scale scores, goals, rankings, and growth.")

    roster=assessment_roster(class_id)
    subject=st.radio("Subject",["Reading","Math"],horizontal=True,key="iready_subject_separate")
    side="reading" if subject=="Reading" else "math"

    if roster.empty:
        st.selectbox("Scholar",["No scholars yet"],disabled=True,key="iready_empty_scholar")
        g1,g2,g3,g4=st.columns(4)
        g1.text_input("Goal (End-of-Year Scale Score)",disabled=True,key="iready_empty_goal")
        g2.text_input("Fall Diagnostic",disabled=True,key="iready_empty_fall")
        g3.text_input("Winter / Midyear Diagnostic",disabled=True,key="iready_empty_winter")
        g4.text_input("Spring Diagnostic",disabled=True,key="iready_empty_spring")
        st.button("Save I‑Ready Scores",disabled=True,key="iready_empty_save")
        render_assessment_insights(pd.DataFrame(columns=["Scholar","Fall","Winter","Spring","Goal","Latest","Growth"]),subject,"I‑Ready")
        st.info("Add scholars in the Scholars section. Their names will automatically populate here.")
        return

    sid=st.selectbox("Scholar",list(roster.id.astype(int)),format_func=lambda x:nm(roster[roster.id==x].iloc[0]),key="iready_scholar_separate")
    br=iready_for_scholar(sid)
    def bv(k):
        return str(br[k] or "") if br and k in br.keys() else ""

    with st.form("iready_separate_form"):
        goal=st.text_input("Goal (End-of-Year Scale Score)",value=bv(f"{side}_goal"),placeholder="Target scale score")
        a,b,c=st.columns(3)
        fall=a.text_input("Fall Diagnostic",value=bv(f"fall_{side}"),placeholder="Scale score")
        winter=b.text_input("Winter / Midyear Diagnostic",value=bv(f"winter_{side}"),placeholder="Scale score")
        spring=c.text_input("Spring Diagnostic",value=bv(f"spring_{side}"),placeholder="Scale score")
        notes=st.text_area("I‑Ready Notes",value=bv("notes"))
        if st.form_submit_button("Save I‑Ready Scores"):
            existing=iready_for_scholar(sid)
            def ev(k):
                return str(existing[k] or "") if existing and k in existing.keys() else ""
            vals={
                "fall_reading":fall if side=="reading" else ev("fall_reading"),
                "winter_reading":winter if side=="reading" else ev("winter_reading"),
                "spring_reading":spring if side=="reading" else ev("spring_reading"),
                "reading_goal":goal if side=="reading" else ev("reading_goal"),
                "fall_math":fall if side=="math" else ev("fall_math"),
                "winter_math":winter if side=="math" else ev("winter_math"),
                "spring_math":spring if side=="math" else ev("spring_math"),
                "math_goal":goal if side=="math" else ev("math_goal"),
            }
            cdb=conn()
            cdb.execute("""INSERT INTO iready_scores(
                scholar_id,fall_reading,winter_reading,spring_reading,reading_goal,
                fall_math,winter_math,spring_math,math_goal,notes)
                VALUES (?,?,?,?,?,?,?,?,?,?)
                ON CONFLICT(scholar_id) DO UPDATE SET
                fall_reading=excluded.fall_reading,
                winter_reading=excluded.winter_reading,
                spring_reading=excluded.spring_reading,
                reading_goal=excluded.reading_goal,
                fall_math=excluded.fall_math,
                winter_math=excluded.winter_math,
                spring_math=excluded.spring_math,
                math_goal=excluded.math_goal,
                notes=excluded.notes""",
                (sid,vals["fall_reading"],vals["winter_reading"],vals["spring_reading"],vals["reading_goal"],
                 vals["fall_math"],vals["winter_math"],vals["spring_math"],vals["math_goal"],notes))
            cdb.commit(); cdb.close()
            st.success("I‑Ready diagnostic data saved.")
            st.rerun()

    render_assessment_insights(iready_scope_dataframe(class_id,subject),subject,"I‑Ready")


# ---------- Newsletter Hub ----------
def current_author_key():
    return str(st.session_state.get("chaplab_username") or (auth_config() or {}).get("username") or "local_teacher").strip().lower()

def current_author_name():
    return teacher_dashboard_info().get("display_name") or get_teacher_name() or "Teacher"

def newsletter_is_lead(author_key=None):
    author_key=author_key or current_author_key()
    c=conn()
    r=c.execute("SELECT is_newsletter_lead FROM newsletter_roles WHERE author_key=?",(author_key,)).fetchone()
    c.close()
    return bool(r["is_newsletter_lead"]) if r else False

def set_newsletter_lead(value, author_key=None):
    author_key=author_key or current_author_key()
    c=conn()
    c.execute("""INSERT INTO newsletter_roles(author_key,is_newsletter_lead) VALUES (?,?)
                 ON CONFLICT(author_key) DO UPDATE SET is_newsletter_lead=excluded.is_newsletter_lead""",
              (author_key,1 if value else 0))
    c.commit(); c.close()

def save_newsletter_version(blurb_id, action, snapshot, author_key=None, author_name=None):
    c=conn()
    c.execute("""INSERT INTO newsletter_blurb_versions(
        blurb_id,author_key,author_name,saved_at,action,content_snapshot)
        VALUES (?,?,?,?,?,?)""",
        (blurb_id,author_key or current_author_key(),author_name or current_author_name(),
         datetime.now().isoformat(timespec="minutes"),action,snapshot))
    c.commit(); c.close()

def newsletter_rows(period=None):
    c=conn()
    q="SELECT * FROM newsletter_blurbs WHERE 1=1"
    p=[]
    if period:
        q+=" AND newsletter_period=?"
        p.append(period)
    q+=" ORDER BY updated_at DESC,id DESC"
    df=pd.read_sql_query(q,c,params=p)
    c.close()
    return df

def generate_subject_newsletter_blurb(subject,recent,family,next_up,prepare):
    pieces=[]
    if recent.strip():
        pieces.append(f"In {subject}, scholars have recently been learning about {recent.strip().rstrip('.')}.")
    if family.strip():
        pieces.append(f"At home, families can help strengthen these skills by {family.strip().rstrip('.')}.")
    if next_up.strip():
        pieces.append(f"Coming up next, scholars will begin {next_up.strip().rstrip('.')}.")
    if prepare.strip():
        pieces.append(f"To help your scholar prepare, {prepare.strip().rstrip('.')}.")
    return " ".join(pieces)

def generate_event_newsletter_blurb(event_type,event_details):
    et=str(event_type or "Upcoming Event").strip()
    details=str(event_details or "").strip()
    return f"Upcoming reminder — {et}: {details}" if details else f"Upcoming reminder: {et}."

def assistant_scholar_context(roster,key_prefix):
    sid=st.selectbox(
        "Scholar",
        list(roster.id.astype(int)),
        format_func=lambda x:nm(roster[roster.id==x].iloc[0]),
        key=f"{key_prefix}_scholar"
    )
    scholar=roster[roster.id==sid].iloc[0]
    name=nm(scholar)
    pro=scholar_pronouns(sid)
    return sid,scholar,name,pro,pro["subject"],pro["possessive"]


# ---------- Creator Rollout Controls ----------
def is_creator_account():
    auth=auth_config() or {}
    # Current single-account build treats the configured ChapLab account as Creator/Admin.
    return current_author_key()==str(auth.get("username","")).strip().lower()

def rollout_setting(feature_key):
    c=conn()
    r=c.execute("SELECT * FROM feature_rollout WHERE feature_key=?",(feature_key,)).fetchone()
    c.close()
    if r:
        return dict(r)
    return {"feature_key":feature_key,"enabled":0,"rollout_stage":"Creator Only"}

def save_rollout_setting(feature_key,enabled,stage):
    c=conn()
    c.execute("""INSERT INTO feature_rollout(feature_key,enabled,rollout_stage,updated_at,updated_by)
                 VALUES (?,?,?,?,?)
                 ON CONFLICT(feature_key) DO UPDATE SET
                   enabled=excluded.enabled,
                   rollout_stage=excluded.rollout_stage,
                   updated_at=excluded.updated_at,
                   updated_by=excluded.updated_by""",
              (feature_key,1 if enabled else 0,stage,datetime.now().isoformat(timespec="minutes"),current_author_name()))
    c.commit(); c.close()

def newsletter_request_rows(active_only=True):
    c=conn()
    q="SELECT * FROM newsletter_requests"
    if active_only:
        q+=" WHERE active=1"
    q+=" ORDER BY COALESCE(due_date,'') ASC,id DESC"
    df=pd.read_sql_query(q,c)
    c.close()
    return df

def newsletter_submission_status(request_row,author_key,subjects):
    period=str(request_row.get("newsletter_period") or "")
    rows=newsletter_rows(period if period else None)
    if rows.empty:
        return []
    rows=rows[rows.author_key==author_key]
    statuses=[]
    for subj in subjects:
        sr=rows[(rows.subject==subj) & (rows.status.isin(["Submitted","Finalized"]))]
        statuses.append((subj,"Submitted" if not sr.empty else "Not Submitted"))
    return statuses

def applicable_newsletter_requests(author_key,profile):
    df=newsletter_request_rows(True)
    if df.empty:
        return df
    grade=str(profile.get("grade_title") or "").lower()
    homeroom=str(profile.get("homeroom") or "").lower()
    subjects=[str(s).lower() for s in profile.get("subjects",[])]
    keep=[]
    for _,r in df.iterrows():
        at=str(r["audience_type"] or "Everyone")
        av=str(r["audience_value"] or "").lower()
        ok=False
        if at=="Everyone":
            ok=True
        elif at=="Grade / Team":
            ok=(av in grade) or (av in homeroom)
        elif at=="Subject Team":
            ok=any(av==s or av in s for s in subjects)
        elif at=="Selected People":
            vals=[x.strip().lower() for x in av.split(",") if x.strip()]
            ok=(author_key.lower() in vals) or (current_author_name().lower() in vals)
        if ok:
            keep.append(r)
    return pd.DataFrame(keep) if keep else pd.DataFrame(columns=df.columns)


# ---------- Demo Data + Team Role Management ----------
DEMO_CLASS_NAME="ChapLab Demo Class"

def approved_staff_df():
    c=conn()
    df=pd.read_sql_query(
        """SELECT * FROM staff_accounts
           WHERE active=1 AND approval_status='Approved'
           ORDER BY display_name,email""",c
    )
    c.close()
    return df

def role_assignments_df(active_only=True):
    c=conn()
    q="SELECT * FROM staff_role_assignments"
    if active_only:
        q+=" WHERE active=1"
    q+=" ORDER BY grade_band,role_name,staff_name"
    df=pd.read_sql_query(q,c)
    c.close()
    return df

def assign_team_role(grade_band,role_name,staff_email,staff_name):
    c=conn()
    # One active Grade Team Leader per grade band; one Newsletter Lead per grade band.
    c.execute(
        """UPDATE staff_role_assignments SET active=0
           WHERE grade_band=? AND role_name=? AND active=1""",
        (grade_band,role_name)
    )
    c.execute(
        """INSERT INTO staff_role_assignments(
           grade_band,role_name,staff_email,staff_name,assigned_by,assigned_at,active)
           VALUES (?,?,?,?,?,?,1)""",
        (grade_band,role_name,staff_email,staff_name,current_author_name(),
         datetime.now().isoformat(timespec="minutes"))
    )
    c.commit(); c.close()

def demo_setting():
    c=conn()
    rows=c.execute(
        "SELECT key,value FROM settings WHERE key IN ('demo_enabled','demo_grade')"
    ).fetchall()
    c.close()
    out={"enabled":False,"grade":"Grade 3"}
    for r in rows:
        try:
            val=json.loads(r["value"])
        except:
            val=r["value"]
        if r["key"]=="demo_enabled":
            out["enabled"]=bool(val)
        elif r["key"]=="demo_grade":
            out["grade"]=str(val or "Grade 3")
    return out

def save_demo_setting(enabled,grade):
    c=conn()
    c.execute("INSERT OR REPLACE INTO settings(key,value) VALUES ('demo_enabled',?)",(json.dumps(bool(enabled)),))
    c.execute("INSERT OR REPLACE INTO settings(key,value) VALUES ('demo_grade',?)",(json.dumps(str(grade)),))
    c.commit(); c.close()

def ensure_demo_data(grade_label="Grade 3"):
    """Create/update one isolated demo class with 3 fake scholars and 4 assignments."""
    grade_num=re.sub(r'[^0-9Kk]','',str(grade_label)).upper() or "3"
    class_label=f"{DEMO_CLASS_NAME} — {grade_label}"

    c=conn()
    row=c.execute("SELECT id FROM classes WHERE is_demo=1 ORDER BY id LIMIT 1").fetchone()
    if row:
        cid=int(row["id"])
        c.execute("UPDATE classes SET class_name=?,subject_note=?,active=1,is_demo=1 WHERE id=?",
                  (class_label,"Demo data only",cid))
    else:
        cur=c.execute(
            "INSERT INTO classes(class_name,subject_note,active,is_demo) VALUES (?,?,1,1)",
            (class_label,"Demo data only")
        )
        cid=int(cur.lastrowid)

    # Clear prior demo children/assignments for a deterministic reset.
    old_sids=[int(r["id"]) for r in c.execute("SELECT id FROM scholars WHERE is_demo=1").fetchall()]
    old_aids=[int(r["id"]) for r in c.execute("SELECT id FROM assignments WHERE is_demo=1").fetchall()]
    if old_sids:
        marks=",".join(["?"]*len(old_sids))
        c.execute(f"DELETE FROM grades WHERE scholar_id IN ({marks})",old_sids)
        c.execute(f"DELETE FROM benchmark_scores WHERE scholar_id IN ({marks})",old_sids)
        c.execute(f"DELETE FROM iready_scores WHERE scholar_id IN ({marks})",old_sids)
        c.execute(f"DELETE FROM scholars WHERE id IN ({marks})",old_sids)
    if old_aids:
        marks=",".join(["?"]*len(old_aids))
        c.execute(f"DELETE FROM grades WHERE assignment_id IN ({marks})",old_aids)
        c.execute(f"DELETE FROM assignments WHERE id IN ({marks})",old_aids)

    demo_students=[
        ("Maya","Johnson","she/her"),
        ("Noah","Williams","he/him"),
        ("Jordan","Taylor","they/them"),
    ]
    sids=[]
    for first,last,pron in demo_students:
        cur=c.execute(
            """INSERT INTO scholars(first_name,last_name,class_name,active,class_id,
               school_name,academic_year,grade_level,student_id,pronouns,is_demo)
               VALUES (?,?,?,1,?,?,?,?,?,?,1)""",
            (first,last,class_label,cid,"ChapLab Demo School",current_academic_year(),
             grade_label,f"DEMO-{len(sids)+1:03d}",pron)
        )
        sids.append(int(cur.lastrowid))

    # Four cross-subject assignments.
    assignment_specs=[
        ("Reading Response","ELA","Classwork","3R3",20.0),
        ("Math Skills Check","Math","Assessment","NY-3.OA.1",20.0),
        ("Science Investigation","Science","Classwork","3-PS2-2",20.0),
        ("Community Exit Ticket","Social Studies","Quiz","SS3-COMM",20.0),
    ]
    aids=[]
    today=date.today()
    for i,(title,subject,category,std,pts) in enumerate(assignment_specs):
        adate=str(today-timedelta(days=(3-i)*3))
        cur=c.execute(
            """INSERT INTO assignments(
               title,subject,category,standard_code,points_possible,assignment_date,
               class_id,marking_period,include_in_average,is_demo)
               VALUES (?,?,?,?,?,?,?,?,1,1)""",
            (title,subject,category,std,pts,adate,cid,
             quarter_for_date(adate,current_academic_year()) or "Quarter 1")
        )
        aids.append(int(cur.lastrowid))

    # Scores intentionally varied so dashboards/gradebook show something meaningful.
    score_grid=[
        [19,18,20,17],  # Maya
        [14,16,15,13],  # Noah
        [17,12,18,16],  # Jordan
    ]
    for sid,row_scores in zip(sids,score_grid):
        for aid,score in zip(aids,row_scores):
            c.execute(
                "INSERT OR REPLACE INTO grades(scholar_id,assignment_id,points_earned) VALUES (?,?,?)",
                (sid,aid,float(score))
            )

    # Demonstration benchmark data.
    nwea_rows=[
        ("188","194","201","202","190","197","204","205","M","N","O"),
        ("176","181","187","194","179","184","190","196","K","L","M"),
        ("183","190","198","200","181","189","197","202","L","M","N"),
    ]
    for sid,vals in zip(sids,nwea_rows):
        c.execute("""INSERT INTO benchmark_scores(
            scholar_id,nwea_fall_reading,nwea_winter_reading,nwea_spring_reading,
            nwea_reading_goal,nwea_fall_math,nwea_winter_math,nwea_spring_math,nwea_math_goal,
            fp_fall_level,fp_winter_level,fp_spring_level)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            (sid,*vals)
        )

    c.commit(); c.close()
    return cid

def set_demo_mode(enabled,grade_label):
    save_demo_setting(enabled,grade_label)
    c=conn()
    if enabled:
        c.close()
        cid=ensure_demo_data(grade_label)
        st.session_state["selected_class"]=cid
        return cid
    c.execute("UPDATE classes SET active=0 WHERE is_demo=1")
    c.execute("UPDATE scholars SET active=0 WHERE is_demo=1")
    c.commit(); c.close()
    if st.session_state.get("selected_class"):
        c=conn()
        r=c.execute("SELECT is_demo FROM classes WHERE id=?",(int(st.session_state["selected_class"]),)).fetchone()
        c.close()
        if r and int(r["is_demo"] or 0)==1:
            st.session_state["selected_class"]=0
    return None


# ---------- Book Leveler: Book Finder + Scholar Fit ----------
def scholar_current_reading_levels(sid):
    fp=""
    lx=""
    try:
        br=benchmark_for_scholar(int(sid))
    except Exception:
        br=None
    if br:
        for key in ("fp_spring_level","fp_winter_level","fp_fall_level"):
            try:
                value=(br[key] or "").strip().upper()
            except Exception:
                value=""
            if value:
                fp=value
                break
        for key in ("lexile_spring","lexile_winter","lexile_fall","lexile"):
            try:
                value=(br[key] or "").strip().upper()
            except Exception:
                value=""
            if value:
                lx=value
                break
    return fp,lx

def book_comparison_levels(book):
    book=book or {}
    sch=book.get("scholastic") or {}
    levels=sch.get("levels") or {}
    direct_fp=(levels.get("fountas_pinnell") or levels.get("guided_reading") or "").strip().upper()
    lexile=(book.get("lexile_measure") or levels.get("lexile") or "").strip().upper()
    est_fp=((book.get("estimated_fp_from_lexile") or {}).get("estimate") or "").strip().upper()
    return direct_fp or est_fp,lexile,direct_fp,est_fp

def classify_roster_for_book(book,roster_df):
    groups={
        "Likely to Struggle":[],
        "Okay / With Support":[],
        "Little to No Problems":[],
        "Need More Data":[]
    }
    if roster_df is None or roster_df.empty:
        return groups
    book_fp,book_lx,_,_=book_comparison_levels(book)
    for _,row in roster_df.iterrows():
        sid=int(row["id"])
        name=nm(row)
        sfp,slx=scholar_current_reading_levels(sid)
        fit,use,why=evaluate_book_fit(
            book_fp=book_fp,book_lexile=book_lx,
            scholar_fp=sfp,scholar_lexile=slx
        )
        item={"name":name,"fp":sfp,"lexile":slx,"fit":fit,"use":use,"why":why}
        if fit=="Too Difficult Right Now":
            groups["Likely to Struggle"].append(item)
        elif fit=="Slightly Challenging":
            groups["Okay / With Support"].append(item)
        elif fit in ("Good Fit","Likely Too Easy"):
            groups["Little to No Problems"].append(item)
        else:
            groups["Need More Data"].append(item)
    return groups

def enrich_book_result(book):
    if not isinstance(book,dict):
        return book
    isbn=re.sub(r"[^0-9Xx]","",str(book.get("isbn") or ""))
    if len(isbn) in (10,13):
        try:
            researched=internet_book_lookup_isbn(isbn)
            if researched:
                return researched
        except Exception:
            pass
    return book

def catalog_recommendations_for_scholar(sid,limit=20):
    sfp,slx=scholar_current_reading_levels(sid)
    c=conn()
    try:
        cat=pd.read_sql_query("SELECT * FROM book_catalog ORDER BY title",c)
    except Exception:
        cat=pd.DataFrame()
    c.close()
    recs=[]
    if cat.empty:
        return recs,sfp,slx
    for _,r in cat.iterrows():
        book_fp=str(r.get("fp_level") or "").strip().upper()
        fit,use,why=evaluate_book_fit(
            book_fp=book_fp,book_lexile="",
            scholar_fp=sfp,scholar_lexile=slx
        )
        if fit in ("Good Fit","Slightly Challenging"):
            recs.append({
                "title":str(r.get("title") or "Untitled"),
                "author":str(r.get("author") or ""),
                "fp":book_fp,"fit":fit,"use":use,"why":why
            })
    rank={"Good Fit":0,"Slightly Challenging":1}
    recs=sorted(recs,key=lambda x:(rank.get(x["fit"],9),x["title"].lower()))
    return recs[:limit],sfp,slx

# ---------- App Shell ----------
# Honor saved Demo Mode without touching real data.
_demo_state=demo_setting()
if not _demo_state["enabled"]:
    _c=conn()
    _c.execute("UPDATE classes SET active=0 WHERE is_demo=1")
    _c.execute("UPDATE scholars SET active=0 WHERE is_demo=1")
    _c.commit(); _c.close()

cdf=classes_df()
folder={0:"All Classes", **{int(r.id):r.class_name for _,r in cdf.iterrows()}}

if "selected_class" not in st.session_state:
    st.session_state["selected_class"]=0

preferred_names=["3-207","3-208","3-212"]
if not st.session_state["selected_class"] and not cdf.empty:
    preferred_ids=[int(cdf[cdf.class_name==n].iloc[0].id) for n in preferred_names if not cdf[cdf.class_name==n].empty]
    st.session_state["selected_class"]=preferred_ids[0] if preferred_ids else int(cdf.iloc[0].id)

selected_class=st.session_state["selected_class"]

internal_pages=["Home Page","Class Dashboard","Scholars","Scholar Profile","Scholar Binder","Book Leveler","Student Grouping","Report Card Comments","Little Assistant","Bulletin Board","Communication Log","Web & Backup"]
if "nav_page" not in st.session_state or st.session_state["nav_page"] not in internal_pages:
    st.session_state["nav_page"]="Home Page"

home_action=st.query_params.get("home")
home_class=st.query_params.get("class")
if home_class:
    try:
        st.session_state["selected_class"]=int(home_class)
        selected_class=int(home_class)
    except:
        pass
    st.query_params.clear()
if home_action:
    home_map={
        "dashboard":("Home Page",None,None,False,False),
        "classes":("Home Page",None,None,True,False),
        "scholars":("Scholars",None,None,False,False),
        "assignments":("Scholar Binder","Add Assignment",None,False,False),
        "grades":("Scholar Binder","Overview",None,False,False),
        "messages":("Little Assistant",None,"Parent Message",False,False),
        "planner":("Home Page",None,None,False,True),
    }
    if home_action in home_map:
        p,bt,at,sc,sq=home_map[home_action]
        st.session_state["nav_page"]=p
        if bt is not None:
            st.session_state["class_binder_tool"]=bt
        if at is not None:
            st.session_state["assistant_tool"]=at
        if sc:
            st.session_state["show_home_class_settings"]=True
        if sq:
            st.session_state["show_quarter_settings"]=True
    st.query_params.clear()

quick=st.query_params.get("binder")
if quick:
    quick_map={
        "planning":("Home Page",None,None),
        "students":("Scholars",None,None),
        "data":("Student Grouping",None,None),
        "lessons":("Scholar Binder","Add Assignment",None),
        "resources":("Scholar Binder","Work Samples",None),
        "meetings":("Communication Log",None,None),
        "misc":("Little Assistant",None,"IEP / Student Support"),
    }
    if quick in quick_map:
        p,bt,at=quick_map[quick]
        st.session_state["nav_page"]=p
        if bt: st.session_state["class_binder_tool"]=bt
        if at: st.session_state["assistant_tool"]=at
    st.query_params.clear()

page=st.session_state["nav_page"]

teacher=get_teacher_name() or "Ms. Chapman"

# Persistent teacher-hub navigation
active_classes=classes_df()
preferred_names=["3-207","3-208","3-212"]
records=[]
if not active_classes.empty:
    records=active_classes.to_dict("records")
    order={n:i for i,n in enumerate(preferred_names)}
    records=sorted(records,key=lambda r:(order.get(r["class_name"],99),r["class_name"]))

page=st.session_state["nav_page"]
selected_class=st.session_state.get("selected_class",0)

teacher=get_teacher_name() or "Ms. Chapman"
initials="".join([p[0] for p in teacher.replace(".","").split() if p])[:2].upper() or "MC"

nav_items=[
    ("🏠","Main Dashboard","Home Page"),
    ("🎓","Scholars","Scholars"),
    ("Ⓐ","Grades","Scholar Binder"),
    ("📚","Book Leveler","Book Leveler"),
    ("👥","Student Grouping","Student Grouping"),
    ("📝","Report Card Comments","Report Card Comments"),
    ("✨","Little Assistant","Little Assistant"),
    ("📌","Bulletin Board","Bulletin Board"),
    ("💬","Communication Log","Communication Log"),
    ("⚙️","Web & Backup","Web & Backup"),
]

def _chaplab_nav_to(target):
    st.session_state["nav_page"]=target
    if target=="Scholar Binder":
        st.session_state["class_binder_tool"]="Overview"

# ---------- Top Teacher Hub Shell ----------
st.markdown("**Classes**")
if records:
    class_cols=st.columns(len(records))
    for col,item in zip(class_cols,records):
        cid=int(item["id"])
        cname=item["class_name"]
        with col:
            if st.button(cname,key=f"chaplab_class_{cid}",use_container_width=True,
                         type="primary" if cid==selected_class else "secondary"):
                st.session_state["selected_class"]=cid
                st.session_state["nav_page"]="Class Dashboard"
                st.session_state["class_binder_tool"]="Overview"
                st.rerun()

profile=teacher_dashboard_info()
display_name=profile.get("display_name") or get_teacher_name() or "Ms. Chapman"
bits=[]
if profile.get("show_grade_title") and profile.get("grade_title"):
    bits.append(f"🍎 {html.escape(profile['grade_title'])}")
if profile.get("show_homeroom") and profile.get("homeroom"):
    bits.append(f"🏫 Homeroom {html.escape(profile['homeroom'])}")
if profile.get("show_school") and profile.get("school"):
    bits.append(f"🏛️ {html.escape(profile['school'])}")
if profile.get("show_subjects") and profile.get("subjects"):
    bits.append("📚 "+html.escape(" · ".join(profile["subjects"])))
if profile.get("show_tagline") and profile.get("tagline"):
    bits.append("💜 "+html.escape(profile["tagline"]))
detail=" &nbsp;&nbsp;|&nbsp;&nbsp; ".join(bits) or "Your personalized teacher workspace"

with st.container(border=True):
    b1,b2=st.columns([5,1])
    with b1:
        st.markdown(
            f"<div class='teacher-hub-banner'><div class='teacher-hub-banner-title'>{html.escape(display_name)}'s Teacher Hub</div><div class='teacher-hub-banner-details'>{detail}</div></div>",
            unsafe_allow_html=True
        )
    with b2:
        st.write("")
        if st.button("⚙️ Profile Settings",key="profile_settings_toggle",use_container_width=True):
            st.session_state["show_teacher_profile_settings"]=not st.session_state.get("show_teacher_profile_settings",False)
            st.rerun()
    if st.session_state.get("show_teacher_profile_settings",False):
        st.markdown("#### Customize Your Teacher Hub Banner")
        p1,p2=st.columns(2)
        pn=p1.text_input("Display name",value=profile.get("display_name",""),key="prof_name")
        pg=p2.text_input("Grade / teacher title",value=profile.get("grade_title",""),placeholder="Example: 3rd Grade Teacher",key="prof_grade")
        p3,p4=st.columns(2)
        ph=p3.text_input("Homeroom",value=profile.get("homeroom",""),placeholder="Example: 3-208",key="prof_home")
        ps=p4.text_input("School",value=profile.get("school",""),key="prof_school")
        psub=st.multiselect("Subjects",["ELA","Math","Science","Social Studies","Grammar","Writing"],
                            default=[s for s in profile.get("subjects",[]) if s in ["ELA","Math","Science","Social Studies","Grammar","Writing"]],key="prof_subjects")
        pt=st.text_input("Custom line / tagline",value=profile.get("tagline",""),key="prof_tagline")
        st.markdown("**Show in Teacher Hub Banner**")
        q1,q2,q3,q4,q5=st.columns(5)
        sg=q1.checkbox("Grade/title",value=profile.get("show_grade_title",True),key="prof_show_grade")
        sh=q2.checkbox("Homeroom",value=profile.get("show_homeroom",True),key="prof_show_home")
        ss=q3.checkbox("School",value=profile.get("show_school",True),key="prof_show_school")
        su=q4.checkbox("Subjects",value=profile.get("show_subjects",True),key="prof_show_subjects")
        stg=q5.checkbox("Tagline",value=profile.get("show_tagline",True),key="prof_show_tagline")
        c1,c2=st.columns(2)
        if c1.button("💾 Save Profile",type="primary",key="prof_save",use_container_width=True):
            save_teacher_dashboard_info(pn,pg,ph,ps,psub,pt,sg,sh,ss,su,stg)
            st.session_state["show_teacher_profile_settings"]=False
            st.rerun()
        if c2.button("Cancel",key="prof_cancel",use_container_width=True):
            st.session_state["show_teacher_profile_settings"]=False
            st.rerun()

        st.markdown("---")
        st.markdown("#### 🔑 Login Username")
        user_state=auth_username_state()
        st.caption(f"Current username: **{user_state['username']}**")

        CREATOR_ADMIN_TITLE="ChapLab App Creator & Administrator"

        if not user_state["initial_choice_used"]:
            st.caption(
                "You have one self-service username change available. "
                "Your current username will remain unchanged unless you choose to change it."
            )

            @st.dialog("Change Username")
            def _first_username_change_dialog():
                current_state=auth_username_state()
                st.warning(
                    f"**This is your one-time self-service username change.**\n\n"
                    f"After you save a new username, any future username changes must be submitted "
                    f"to the **{CREATOR_ADMIN_TITLE}** for approval."
                )
                st.write(f"Current username: **{current_state['username']}**")
                proposed=st.text_input(
                    "New username",
                    placeholder="3–30 characters: letters, numbers, . _ -",
                    key="settings_first_username_new"
                )
                cleaned=_username_clean(proposed)

                if proposed and not _username_valid(cleaned):
                    st.warning(
                        "Use 3–30 characters with no spaces. "
                        "Letters, numbers, periods, underscores, and hyphens are allowed."
                    )

                if st.button(
                    "Save One-Time Username Change",
                    type="primary",
                    use_container_width=True,
                    key="settings_save_first_username"
                ):
                    if not _username_valid(cleaned):
                        st.error("Enter a valid username first.")
                    elif cleaned==current_state["username"]:
                        st.info("That is already your current username.")
                    else:
                        save_initial_username_choice(cleaned)
                        st.success(f"Username changed to **{cleaned}**.")
                        st.rerun()

            if st.button(
                "Change Username",
                key="settings_open_first_username_change",
                use_container_width=True
            ):
                _first_username_change_dialog()

        else:
            st.caption(
                f"Your one-time self-service username change has been used. "
                f"Future changes require approval from the **{CREATOR_ADMIN_TITLE}**."
            )

            @st.dialog("Request Username Change")
            def _username_change_request_dialog():
                current_state=auth_username_state()
                st.info(
                    f"This request will be forwarded to the **{CREATOR_ADMIN_TITLE}** for approval. "
                    "Your current username will continue to work unless the request is approved."
                )
                st.write(f"Current username: **{current_state['username']}**")

                requested=st.text_input(
                    "Requested username",
                    placeholder="3–30 characters: letters, numbers, . _ -",
                    key="settings_requested_username"
                )
                reason=st.text_area(
                    "Reason for change (optional)",
                    placeholder="Explain why you would like your username changed.",
                    height=85,
                    key="settings_username_request_reason"
                )

                if st.button(
                    "Send Request for Approval",
                    type="primary",
                    use_container_width=True,
                    key="settings_send_username_request"
                ):
                    cleaned=_username_clean(requested)
                    if not _username_valid(cleaned):
                        st.error(
                            "Use 3–30 characters with no spaces. "
                            "Letters, numbers, periods, underscores, and hyphens are allowed."
                        )
                    elif cleaned==current_state["username"]:
                        st.info("That is already your current username.")
                    else:
                        submit_username_change_request(cleaned,reason)
                        st.success(
                            f"Username change request sent to the **{CREATOR_ADMIN_TITLE}**."
                        )

            if st.button(
                "Change Username",
                key="settings_open_username_request",
                use_container_width=True
            ):
                _username_change_request_dialog()

        st.markdown("---")
        st.markdown("#### Newsletter Role")
        current_lead=newsletter_is_lead()
        lead_choice=st.checkbox(
            "Newsletter Lead — can edit/finalize all submitted blurbs",
            value=current_lead,
            key="prof_newsletter_lead"
        )
        if st.button("Save Newsletter Role",key="prof_save_newsletter_role"):
            set_newsletter_lead(lead_choice)
            st.success("Newsletter role saved.")
            st.rerun()

        if is_creator_account():
            st.markdown("---")
            st.markdown("#### 👤 Account Management")
            st.caption(
                "View everyone registered for ChapLab, confirm who is active, see recent sign-in activity, "
                "and deactivate accounts that should no longer have access."
            )
            st.info(
                "School staff accounts must use an email beginning with **79.** and ending with "
                "**@nhaschools.com**. After approval, first-time staff use the case-insensitive access code "
                "**bscs** to enter password setup. Once their personal password is created, the access code stops "
                "working for that account. The ChapLab App Creator & Administrator recovery login is the only exception."
            )

            _accounts=all_staff_accounts_df()
            if _accounts.empty:
                st.caption("No staff accounts are registered yet.")
            else:
                _display=_accounts[[
                    "display_name","email","role_type","grade_band",
                    "approval_status","active","password_setup_required",
                    "last_login","last_seen","login_count"
                ]].copy()
                _display["password_setup_required"]=_display["password_setup_required"].apply(
                    lambda x:"Needs Setup" if bool(x) else "Password Set"
                )
                _display.columns=[
                    "Name","Email","Role","Grade/Team","Approval",
                    "Active","Password","Last Login","Last Seen","Logins"
                ]
                st.dataframe(_display,use_container_width=True,hide_index=True)

                _account_emails=list(_accounts["email"])
                _manage_email=st.selectbox(
                    "Manage account",
                    _account_emails,
                    format_func=lambda e:(
                        f"{_accounts[_accounts.email==e].iloc[0]['display_name']} — {e}"
                    ),
                    key="creator_manage_staff_account"
                )
                _row=_accounts[_accounts.email==_manage_email].iloc[0]

                am1,am2,am3=st.columns(3)
                am1.metric("Status","Active" if bool(_row["active"]) else "Deactivated")
                am2.metric("Approval",str(_row["approval_status"]))
                am3.metric("Login count",int(_row["login_count"] or 0))

                if not valid_school_staff_email(_manage_email):
                    st.error(
                        "⚠️ This staff account does not match the required school email format "
                        "`79.…@nhaschools.com`. It should not be approved for staff access."
                    )

                _reason=st.text_input(
                    "Deactivation / account note (optional)",
                    key="creator_account_action_reason"
                )
                ac1,ac2,ac3,ac4=st.columns(4)

                if bool(_row["active"]):
                    if ac1.button("🚫 Deactivate Account",key="creator_deactivate_account",use_container_width=True):
                        set_staff_account_active(_manage_email,False,_reason)
                        st.success(f"{_manage_email} has been deactivated.")
                        st.rerun()
                else:
                    if ac1.button("✅ Reactivate Account",key="creator_reactivate_account",use_container_width=True):
                        if valid_school_staff_email(_manage_email):
                            set_staff_account_active(_manage_email,True,_reason)
                            st.success(f"{_manage_email} has been reactivated.")
                            st.rerun()
                        else:
                            st.error("This email does not meet the required 79.* NHA school format.")

                if str(_row["approval_status"])!="Approved":
                    if ac2.button("Approve Account",key="creator_approve_account",use_container_width=True):
                        if set_staff_approval(_manage_email,"Approved"):
                            st.success("Account approved.")
                            st.rerun()
                        else:
                            st.error("Only valid 79.*@nhaschools.com staff emails can be approved.")
                else:
                    if ac2.button("Set to Pending",key="creator_pending_account",use_container_width=True):
                        set_staff_approval(_manage_email,"Pending")
                        st.success("Account moved to Pending.")
                        st.rerun()

                if ac3.button("🔑 Reset Password Setup",key="creator_reset_password_setup",use_container_width=True):
                    reset_staff_password_setup(_manage_email)
                    st.success(
                        f"{_manage_email} can now sign in with the initial access code and create a new password."
                    )
                    st.rerun()

                if ac4.button("Refresh Account List",key="creator_refresh_accounts",use_container_width=True):
                    st.rerun()

                with st.expander("Recent account activity"):
                    c=conn()
                    _activity=pd.read_sql_query(
                        """SELECT staff_email,logged_in_at,event_type,details
                           FROM staff_login_activity
                           ORDER BY id DESC LIMIT 50""",c
                    )
                    c.close()
                    if _activity.empty:
                        st.caption("No staff login activity recorded yet.")
                    else:
                        st.dataframe(_activity,use_container_width=True,hide_index=True)

            st.markdown("---")
            st.markdown("#### 🔐 Creator Rollout Controls")
            st.caption("Keep unfinished roles/features hidden until you move them through testing.")

            rollout_stages=[
                "Creator Only",
                "Creator + Grade Team",
                "Creator + Grade Team + Dean",
                "School Pilot",
                "Released"
            ]
            feature_defs=[
                ("self_signup","Staff Self-Signup / Pending Approval"),
                ("dean_role","Dean Role & Dashboard"),
                ("sped_dean_role","SPED Dean Role"),
                ("staff_roles","Expanded Staff Roles / Specials"),
                ("cross_program","Cross-Program Collaboration"),
            ]
            changed=[]
            for fkey,label in feature_defs:
                cur=rollout_setting(fkey)
                rc1,rc2=st.columns([1,2])
                enabled=rc1.toggle(label,value=bool(cur.get("enabled",0)),key=f"rollout_enable_{fkey}")
                current_stage=cur.get("rollout_stage","Creator Only")
                if current_stage not in rollout_stages:
                    current_stage="Creator Only"
                stage=rc2.selectbox(
                    f"{label} testing stage",
                    rollout_stages,
                    index=rollout_stages.index(current_stage),
                    key=f"rollout_stage_{fkey}"
                )
                changed.append((fkey,enabled,stage))

            if st.button("💾 Save Rollout Controls",key="save_rollout_controls"):
                for fkey,enabled,stage in changed:
                    save_rollout_setting(fkey,enabled,stage)
                st.success("Rollout controls saved.")
                st.rerun()

            st.info(
                "Current plan: Creator → Creator + Grade Team → Creator + Grade Team + Dean. "
                "Dean/SPED Dean/cross-program features remain hidden until you turn them on."
            )


            st.markdown("---")
            st.markdown("#### 🔑 Username Change Requests — App Administration")
            pending_names=pending_username_requests()
            if pending_names.empty:
                st.caption("No username change requests are waiting.")
            else:
                for _,ur in pending_names.iterrows():
                    rid=int(ur["id"])
                    with st.container(border=True):
                        st.write(
                            f"**{ur['requester_name'] or ur['requester_key']}** wants to change "
                            f"`{ur['current_username']}` → `{ur['requested_username']}`"
                        )
                        if ur["reason"]:
                            st.caption("Reason: "+str(ur["reason"]))
                        review_note=st.text_input(
                            "Review note (optional)",
                            key=f"username_review_note_{rid}"
                        )
                        ua1,ua2=st.columns(2)
                        if ua1.button("✅ Approve",key=f"approve_username_{rid}",use_container_width=True):
                            review_username_request(rid,True,review_note)
                            st.success("Username change approved.")
                            st.rerun()
                        if ua2.button("❌ Deny",key=f"deny_username_{rid}",use_container_width=True):
                            review_username_request(rid,False,review_note)
                            st.success("Username change denied.")
                            st.rerun()

            st.markdown("---")
            st.markdown("#### 🎭 Demo Class")
            st.caption(
                "Turn on a safe fake class for demonstrations. Demo data is tagged separately "
                "and can be hidden again without touching real classes or scholars."
            )
            demo=demo_setting()
            dm1,dm2=st.columns(2)
            demo_enabled=dm1.toggle("Show Demo Class",value=bool(demo["enabled"]),key="creator_demo_enabled")
            demo_grade=dm2.selectbox(
                "Demo grade",
                ["Kindergarten","Grade 1","Grade 2","Grade 3","Grade 4","Grade 5","Grade 6","Grade 7","Grade 8"],
                index=(["Kindergarten","Grade 1","Grade 2","Grade 3","Grade 4","Grade 5","Grade 6","Grade 7","Grade 8"].index(demo["grade"])
                       if demo["grade"] in ["Kindergarten","Grade 1","Grade 2","Grade 3","Grade 4","Grade 5","Grade 6","Grade 7","Grade 8"] else 3),
                key="creator_demo_grade"
            )
            if st.button("Apply Demo Class Setting",key="apply_demo_setting",use_container_width=True):
                cid=set_demo_mode(demo_enabled,demo_grade)
                st.success(
                    f"Demo Class is {'ON' if demo_enabled else 'OFF'}."
                    + (f" Opened {demo_grade} demo data." if demo_enabled else "")
                )
                st.rerun()

            if demo_enabled:
                st.info(
                    "Demo contains 3 fake scholars, 4 assignments across ELA/Math/Science/Social Studies, "
                    "sample grades, and sample reading/assessment data."
                )

            st.markdown("---")
            st.markdown("#### 🧪 Grade 3 Pilot Team")
            st.caption(
                "These teachers are pre-approved for the Grade 3 testing stage. "
                "No Grade Team Leader or Newsletter Lead has been assigned yet."
            )
            _pilot=approved_staff_df()
            if not _pilot.empty:
                _pilot=_pilot[_pilot["email"].str.lower().isin([
                    "79.jschroeder@nhaschools.com",
                    "79.ncampbell@nhaschools.com",
                    "79.adavidson@nhaschools.com",
                ])]
                for _,_p in _pilot.iterrows():
                    st.write(f"✅ **{_p['display_name']}** — {_p['email']} — Grade 3 Teacher")

            st.markdown("---")
            st.markdown("#### 👥 Team & Role Management")
            st.caption(
                "Until the Dean role is released, the ChapLab App Creator & Administrator controls Grade Team Leader and Newsletter Lead assignments. "
                "When Dean is enabled, this same structure will move to the appropriate Dean permissions."
            )
            staff=approved_staff_df()
            if staff.empty:
                st.info(
                    "Role framework is ready. There are no approved staff accounts yet because self-signup is still OFF. "
                    "Once approved accounts exist, they will appear here automatically."
                )
            else:
                grade_band=st.selectbox(
                    "Grade band / team",
                    ["Kindergarten","Grade 1","Grade 2","Grade 3","Grade 4","Grade 5","Grade 6","Grade 7","Grade 8",
                     "Special Areas","Intervention / Student Support","Other / Custom"],
                    key="role_grade_band"
                )
                custom_band=st.text_input("Other / custom grade band",key="role_grade_band_other")
                final_band=custom_band.strip() if grade_band=="Other / Custom" and custom_band.strip() else grade_band
                staff_ids=list(staff.index)
                selected_idx=st.selectbox(
                    "Approved staff member",
                    staff_ids,
                    format_func=lambda i:f"{staff.loc[i,'display_name']} ({staff.loc[i,'email']})",
                    key="role_staff_member"
                )
                staff_row=staff.loc[selected_idx]
                rr1,rr2=st.columns(2)
                if rr1.button("Assign Grade Team Leader",key="assign_grade_team_leader",use_container_width=True):
                    assign_team_role(final_band,"Grade Team Leader",staff_row["email"],staff_row["display_name"])
                    st.success(f"{staff_row['display_name']} assigned as {final_band} Grade Team Leader.")
                    st.rerun()
                if rr2.button("Assign Newsletter Lead",key="assign_newsletter_lead",use_container_width=True):
                    assign_team_role(final_band,"Newsletter Lead",staff_row["email"],staff_row["display_name"])
                    st.success(f"{staff_row['display_name']} assigned as {final_band} Newsletter Lead.")
                    st.rerun()

            roles=role_assignments_df(True)
            if not roles.empty:
                st.markdown("##### Current Team Roles")
                st.dataframe(
                    roles[["grade_band","role_name","staff_name","staff_email","assigned_by","assigned_at"]],
                    hide_index=True,use_container_width=True
                )

_nav=[("🏠 Dashboard","Home Page"),("🎓 Scholars","Scholars"),("Ⓐ Grades","Scholar Binder"),("📚 Book Leveler","Book Leveler"),("👥 Student Grouping","Student Grouping"),("📝 Report Comments","Report Card Comments"),("✨ Little Assistant","Little Assistant"),("📌 Bulletin Board","Bulletin Board"),("💬 Communication","Communication Log"),("⚙️ Web & Backup","Web & Backup")]
cols=st.columns(10)
for col,(label,target) in zip(cols,_nav):
    with col:
        if st.button(label,key=f"topnav_{target}",use_container_width=True,type="primary" if page==target else "secondary"):
            st.session_state["nav_page"]=target
            if target=="Scholar Binder":
                st.session_state["class_binder_tool"]="Overview"
            st.rerun()
st.markdown("---")

# ---------- Pages ----------
if page=="Home Page":
    all_classes=classes_df()
    total_all=len(scholars_df())

    class_summaries=[]
    for _,cr in all_classes.iterrows():
        cid=int(cr.id)
        counts,_=scholar_status_summary(cid)
        class_summaries.append({
            "name":cr.class_name,
            "count":len(scholars_df(cid)),
            "on":counts["On Track"],
            "approach":counts["Approaching"],
            "risk":counts["At Risk"],
            "nodata":counts["No Data"]
        })

    recent=recent_assignments_df(1)
    latest="No assignments yet"
    if not recent.empty:
        rr=recent.iloc[0]
        latest=f"{rr.title} • {rr.subject}"

    anns=home_announcements()
    next_due="No saved reminder"
    if anns:
        k,l,d=anns[0]
        next_due=f"{k}: {l}" + (f" • {d}" if d else "")

    st.markdown('<div class="accent-page-label">Main Dashboard</div>',unsafe_allow_html=True)

    posts=[]
    colors=["s-pink","s-orange","s-yellow"]
    for i,info in enumerate(class_summaries[:3]):
        posts.append(
            f'<div class="sticky-card {colors[i]}">'
            f'<strong>{html.escape(info["name"])}</strong>'
            f'{info["count"]} Scholars'
            f'<small>{info["on"]} on track • {info["approach"]+info["risk"]} need attention</small>'
            f'</div>'
        )
    posts.extend([
        f'<div class="sticky-card s-green"><strong>{total_all}</strong>Total Scholars<small>Across all active classes</small></div>',
        f'<div class="sticky-card s-blue"><strong>Latest</strong>{html.escape(latest)}<small>Most recent assignment entered</small></div>',
        f'<div class="sticky-card s-purple"><strong>Reminder</strong>{html.escape(next_due)}<small>Next saved deadline/update</small></div>',
    ])
    st.markdown('<div class="sticky-grid">'+"".join(posts)+'</div>',unsafe_allow_html=True)

    st.markdown('<div class="page-title">All Classes Overview</div>',unsafe_allow_html=True)
    overview_rows=[]
    for info in class_summaries:
        overview_rows.append({
            "Class":info["name"],
            "Scholars":info["count"],
            "On Track":info["on"],
            "Approaching":info["approach"],
            "At Risk":info["risk"],
            "No Data":info["nodata"]
        })
    if overview_rows:
        st.dataframe(pd.DataFrame(overview_rows),hide_index=True,use_container_width=True)

    left,right=st.columns(2)
    with left:
        st.markdown("### Recent Assignments")
        recent5=recent_assignments_df(5)
        if recent5.empty:
            st.caption("No assignments yet.")
        else:
            cols=[c for c in ["assignment_date","title","subject","class_name"] if c in recent5.columns]
            st.dataframe(recent5[cols],hide_index=True,use_container_width=True)
    with right:
        st.markdown("### Reminders & Deadlines")
        anns=home_announcements()
        if not anns:
            st.caption("No saved reminders or report-card deadlines.")
        else:
            st.dataframe(pd.DataFrame(anns,columns=["Type","Reminder","Due"]),hide_index=True,use_container_width=True)


    # Newsletter deadlines assigned by the Newsletter Lead / future Dean.
    profile_now=teacher_dashboard_info()
    assigned_requests=applicable_newsletter_requests(current_author_key(),profile_now)
    if not assigned_requests.empty:
        st.markdown("### ⏰ Newsletter Deadlines")
        for _,req in assigned_requests.iterrows():
            req_subjects=[x.strip() for x in str(req["required_subjects"] or "").split(",") if x.strip()]
            due=req["due_date"] or "No due date"
            with st.container(border=True):
                st.markdown(f"**{req['newsletter_period'] or 'Newsletter Submission'}** — Due **{due}**")
                st.caption(f"For: {req['audience_type']} {req['audience_value'] or ''}")
                if req["instructions"]:
                    st.write(req["instructions"])
                if req_subjects:
                    st.write("Required blurbs: "+", ".join(req_subjects))
                    status_pairs=newsletter_submission_status(req,current_author_key(),req_subjects)
                    st.write(" • ".join([f"{s}: {status}" for s,status in status_pairs]))

    st.markdown("---")
    st.markdown('<div class="page-title">📰 Newsletter Hub</div>',unsafe_allow_html=True)
    st.caption(
        "Create family-friendly subject blurbs, submit them to the newsletter lead, "
        "and keep a dated history of every saved version."
    )

    author_key=current_author_key()
    author_name=current_author_name()
    is_lead=newsletter_is_lead(author_key)
    teacher_subjects=teacher_dashboard_info().get("subjects",[]) or ["ELA","Math","Science","Social Studies"]
    ay=current_academic_year()

    n1,n2=st.columns([2,1])
    newsletter_period=n1.text_input(
        "Newsletter issue / week",
        value=st.session_state.get("newsletter_period",""),
        placeholder="Example: Week of September 14",
        key="newsletter_period_input"
    )
    st.session_state["newsletter_period"]=newsletter_period
    n2.info("Newsletter Lead" if is_lead else "Teacher Contributor")

    tab_names=["✍️ Create My Blurb","📬 Submitted Blurbs","🧾 My Archive"]
    if is_lead:
        tab_names.append("🗞️ Newsletter Lead")
    tabs=st.tabs(tab_names)

    with tabs[0]:
        blurb_type=st.radio("Blurb type",["Subject Update","School / Upcoming Event"],horizontal=True,key="newsletter_blurb_type")

        if blurb_type=="Subject Update":
            subject=st.selectbox(
                "Subject",
                list(dict.fromkeys(teacher_subjects+["ELA","Math","Science","Social Studies","Other / Custom"])),
                key="newsletter_subject"
            )
            subject_custom=st.text_input("Other / custom subject",key="newsletter_subject_custom")
            subject_final=subject_custom.strip() if subject=="Other / Custom" and subject_custom.strip() else subject
            title=st.text_input("Optional blurb title",placeholder=f"{subject_final} Update",key="newsletter_title")
            recent=st.text_area("What have scholars recently learned?",height=90,key="newsletter_recent")
            family=st.text_area("How can families help scholars improve/practice these skills?",height=90,key="newsletter_family")
            next_up=st.text_area("What's coming up next?",height=80,key="newsletter_next")
            prepare=st.text_area("How can families help their child prepare?",height=80,key="newsletter_prepare")
            if st.button("✨ Generate Subject Blurb",key="newsletter_generate_subject"):
                st.session_state["newsletter_generated"]=generate_subject_newsletter_blurb(subject_final,recent,family,next_up,prepare)
            event_type=""
            event_details=""
            record_type="Subject"
        else:
            subject_final=""
            title=st.text_input("Optional event heading",placeholder="Upcoming Dates & Reminders",key="newsletter_event_title")
            event_choice=st.selectbox(
                "Event / announcement type",
                ["Dress-Down Day","Field Trip","School Event","Assessment / Testing",
                 "Half Day / Schedule Change","Family Event","Deadline / Form Due",
                 "School Closure","Fundraiser","Other / Custom"],
                key="newsletter_event_type"
            )
            event_custom=st.text_input("Other / custom event type",key="newsletter_event_type_other")
            event_type=event_custom.strip() if event_choice=="Other / Custom" and event_custom.strip() else event_choice
            event_details=st.text_area(
                "Event details",
                placeholder="Include date, time, clothing/items needed, permission slip deadline, cost, location, or anything families should know.",
                height=135,key="newsletter_event_details"
            )
            if st.button("✨ Generate Event Blurb",key="newsletter_generate_event"):
                st.session_state["newsletter_generated"]=generate_event_newsletter_blurb(event_type,event_details)
            recent=family=next_up=prepare=""
            record_type="Event"

        generated=st.text_area(
            "Edit / finalize your blurb",
            value=st.session_state.get("newsletter_generated",""),
            height=180,key="newsletter_generated_editor"
        )
        save1,save2=st.columns(2)
        if save1.button("💾 Save Draft",key="newsletter_save_draft",use_container_width=True):
            now=datetime.now().isoformat(timespec="minutes")
            c=conn()
            cur=c.execute("""INSERT INTO newsletter_blurbs(
                author_key,author_name,created_at,updated_at,academic_year,newsletter_period,
                blurb_type,subject,title,recently_taught,family_help,coming_next,prepare_next,
                event_type,event_details,generated_blurb,status)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (author_key,author_name,now,now,ay,newsletter_period,record_type,subject_final,title,
                 recent,family,next_up,prepare,event_type,event_details,generated,"Draft"))
            bid=int(cur.lastrowid); c.commit(); c.close()
            save_newsletter_version(bid,"Draft Saved",generated,author_key,author_name)
            st.success("Draft saved to your Newsletter archive.")

        if save2.button("📨 Save & Submit to Newsletter",type="primary",key="newsletter_submit",use_container_width=True):
            if not generated.strip():
                st.warning("Generate or type the blurb before submitting.")
            else:
                now=datetime.now().isoformat(timespec="minutes")
                c=conn()
                cur=c.execute("""INSERT INTO newsletter_blurbs(
                    author_key,author_name,created_at,updated_at,academic_year,newsletter_period,
                    blurb_type,subject,title,recently_taught,family_help,coming_next,prepare_next,
                    event_type,event_details,generated_blurb,status,submitted_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (author_key,author_name,now,now,ay,newsletter_period,record_type,subject_final,title,
                     recent,family,next_up,prepare,event_type,event_details,generated,"Submitted",now))
                bid=int(cur.lastrowid); c.commit(); c.close()
                save_newsletter_version(bid,"Submitted",generated,author_key,author_name)
                st.success("Submitted. Everyone can now see this blurb in Submitted Blurbs.")

        st.caption("Teaching multiple subjects? Save/submit this one, choose your next subject, and create the next blurb.")

    with tabs[1]:
        submitted=newsletter_rows(newsletter_period if newsletter_period.strip() else None)
        submitted=submitted[submitted.status.isin(["Submitted","Finalized"])] if not submitted.empty else submitted
        if submitted.empty:
            st.caption("No submitted blurbs yet.")
        else:
            for _,r in submitted.iterrows():
                heading=r["title"] or r["subject"] or r["event_type"] or "Newsletter Blurb"
                with st.expander(f"{r['status']} • {heading} • {r['author_name']}",expanded=False):
                    st.caption(f"{r['newsletter_period'] or 'No issue set'} • Updated {r['updated_at']}")
                    st.text_area("Submitted blurb",value=r["generated_blurb"] or "",height=150,disabled=True,key=f"shared_newsletter_{int(r['id'])}")

    with tabs[2]:
        mine=newsletter_rows()
        mine=mine[mine.author_key==author_key] if not mine.empty else mine
        if mine.empty:
            st.caption("You have not saved any newsletter blurbs yet.")
        else:
            for _,r in mine.iterrows():
                bid=int(r["id"])
                heading=r["title"] or r["subject"] or r["event_type"] or "Newsletter Blurb"
                with st.expander(f"{r['status']} • {heading} • {r['updated_at']}",expanded=False):
                    edit_text=st.text_area("Your saved blurb",value=r["generated_blurb"] or "",height=155,key=f"mine_newsletter_edit_{bid}")
                    ec1,ec2=st.columns(2)
                    if ec1.button("Save My Changes",key=f"mine_newsletter_save_{bid}",use_container_width=True):
                        now=datetime.now().isoformat(timespec="minutes")
                        c=conn()
                        c.execute("UPDATE newsletter_blurbs SET generated_blurb=?,updated_at=? WHERE id=? AND author_key=?",
                                  (edit_text,now,bid,author_key))
                        c.commit(); c.close()
                        save_newsletter_version(bid,"Author Edited",edit_text,author_key,author_name)
                        st.rerun()
                    if r["status"]=="Draft":
                        if ec2.button("Submit This Draft",key=f"mine_newsletter_submit_{bid}",use_container_width=True):
                            now=datetime.now().isoformat(timespec="minutes")
                            c=conn()
                            c.execute("""UPDATE newsletter_blurbs SET status='Submitted',
                                       submitted_at=?,updated_at=? WHERE id=? AND author_key=?""",
                                      (now,now,bid,author_key))
                            c.commit(); c.close()
                            save_newsletter_version(bid,"Submitted",edit_text,author_key,author_name)
                            st.rerun()
                    c=conn()
                    versions=pd.read_sql_query(
                        """SELECT saved_at,action,author_name
                           FROM newsletter_blurb_versions WHERE blurb_id=?
                           ORDER BY id DESC""",c,params=[bid]
                    )
                    c.close()
                    if not versions.empty:
                        st.caption(f"{len(versions)} saved version(s) kept as receipts.")
                        st.dataframe(versions,hide_index=True,use_container_width=True)

    if is_lead:
        with tabs[3]:
            st.markdown("### Newsletter Lead Workspace")
            st.caption("Newsletter Lead can set submission deadlines, assign who needs to submit, edit submitted blurbs, and finalize newsletter copy.")
            st.caption("Role path: Creator/Admin assigns the Grade Team Leader until Dean is released. The Grade Team Leader may serve as Newsletter Lead or assign another approved staff member in that grade band.")

            st.markdown("#### 📅 Create Newsletter Request / Due Date")
            with st.form("newsletter_request_form",clear_on_submit=True):
                rq1,rq2=st.columns(2)
                rq_period=rq1.text_input("Newsletter issue / week",value=newsletter_period,placeholder="Example: Week of September 14")
                rq_due=rq2.date_input("Due date",value=date.today())
                rq3,rq4=st.columns(2)
                rq_audience=rq3.selectbox("Who needs to submit?",["Everyone","Grade / Team","Subject Team","Selected People"])
                rq_value=rq4.text_input(
                    "Grade/team/subject/people",
                    placeholder="Examples: Grade 3 | Science | chapman, smith"
                )
                rq_subjects=st.multiselect(
                    "Required subject blurbs",
                    ["ELA","Math","Science","Social Studies","Art","Music","Physical Education / Gym","Technology","Library / Media","STEM","Health","World Language","Intervention / AIS","ENL / ELL","Special Education","Other / Custom"]
                )
                rq_subject_other=st.text_input("Other / custom required subject")
                rq_instructions=st.text_area("Instructions / notes",placeholder="Anything staff should include in this newsletter issue.")
                if st.form_submit_button("Create Newsletter Request",use_container_width=True):
                    subjects=list(rq_subjects)
                    if rq_subject_other.strip():
                        subjects.append(rq_subject_other.strip())
                    c=conn()
                    c.execute("""INSERT INTO newsletter_requests(
                        created_by_key,created_by_name,created_at,newsletter_period,due_date,
                        audience_type,audience_value,required_subjects,instructions,active)
                        VALUES (?,?,?,?,?,?,?,?,?,1)""",
                        (current_author_key(),current_author_name(),datetime.now().isoformat(timespec="minutes"),
                         rq_period.strip(),str(rq_due),rq_audience,rq_value.strip(),
                         ",".join(subjects),rq_instructions.strip()))
                    c.commit(); c.close()
                    st.success("Newsletter request created.")
                    st.rerun()

            requests=newsletter_request_rows(True)
            if not requests.empty:
                st.markdown("#### Active Newsletter Requests")
                for _,req in requests.iterrows():
                    rid=int(req["id"])
                    with st.expander(f"{req['newsletter_period'] or 'Newsletter Request'} • Due {req['due_date'] or 'TBD'}",expanded=False):
                        st.write(f"Audience: **{req['audience_type']}** {req['audience_value'] or ''}")
                        st.write(f"Required blurbs: {req['required_subjects'] or 'Not specified'}")
                        if req["instructions"]:
                            st.write(req["instructions"])
                        if st.button("Close Request",key=f"close_news_req_{rid}"):
                            c=conn()
                            c.execute("UPDATE newsletter_requests SET active=0 WHERE id=?",(rid,))
                            c.commit(); c.close()
                            st.rerun()

            st.markdown("---")
            st.markdown("#### Submitted Newsletter Blurbs")
            all_sub=newsletter_rows(newsletter_period if newsletter_period.strip() else None)
            all_sub=all_sub[all_sub.status.isin(["Submitted","Finalized"])] if not all_sub.empty else all_sub
            if all_sub.empty:
                st.caption("No submitted blurbs are waiting.")
            else:
                for _,r in all_sub.iterrows():
                    bid=int(r["id"])
                    heading=r["title"] or r["subject"] or r["event_type"] or "Newsletter Blurb"
                    with st.expander(f"{r['status']} • {r['author_name']} • {heading}",expanded=False):
                        lead_edit=st.text_area("Newsletter copy",value=r["generated_blurb"] or "",height=160,key=f"lead_newsletter_edit_{bid}")
                        l1,l2=st.columns(2)
                        if l1.button("Save Lead Edit",key=f"lead_newsletter_save_{bid}",use_container_width=True):
                            now=datetime.now().isoformat(timespec="minutes")
                            c=conn()
                            c.execute("UPDATE newsletter_blurbs SET generated_blurb=?,updated_at=? WHERE id=?",(lead_edit,now,bid))
                            c.commit(); c.close()
                            save_newsletter_version(bid,"Newsletter Lead Edited",lead_edit,author_key,author_name)
                            st.rerun()
                        if l2.button("✅ Finalize Blurb",key=f"lead_newsletter_finalize_{bid}",use_container_width=True):
                            now=datetime.now().isoformat(timespec="minutes")
                            c=conn()
                            c.execute("""UPDATE newsletter_blurbs SET generated_blurb=?,status='Finalized',
                                       finalized_at=?,updated_at=? WHERE id=?""",(lead_edit,now,now,bid))
                            c.commit(); c.close()
                            save_newsletter_version(bid,"Finalized",lead_edit,author_key,author_name)
                            st.rerun()

            if not all_sub.empty:
                final_copy=[]
                for _,r in all_sub.iterrows():
                    heading=r["title"] or r["subject"] or r["event_type"] or "Update"
                    final_copy.append(f"{heading}\n{r['generated_blurb']}\n— {r['author_name']}")
                st.text_area("All submitted blurbs — copy into newsletter",value="\n\n".join(final_copy),height=360,key="newsletter_all_copy")

elif page=="Class Dashboard":
    if not selected_class:
        st.warning("Choose a class from the top class bar.")
        st.stop()

    selected_name=class_name_from_id(int(selected_class)) or "Selected Class"
    counts,detail=scholar_status_summary(int(selected_class))
    roster=scholars_df(int(selected_class))
    total=len(roster)

    c=conn()
    assignment_count=c.execute("SELECT COUNT(*) n FROM assignments WHERE class_id=?",(int(selected_class),)).fetchone()["n"]
    contact_count=c.execute("""SELECT COUNT(*) n FROM communications
                               WHERE scholar_id IN (SELECT id FROM scholars WHERE class_id=?)""",
                            (int(selected_class),)).fetchone()["n"]
    c.close()

    recent=class_assignments_df(int(selected_class),"All Subjects","Newest → Oldest")
    latest="No assignments yet"
    if not recent.empty:
        rr=recent.iloc[0]
        latest=f"{rr.title} • {rr.subject}"

    hero=f"""
    <div class="hero-card">
      <h1>{html.escape(selected_name)} Dashboard</h1>
      <div class="accent">Class View</div>
    </div>
    """
    if hasattr(st,"html"): st.html(hero)
    else: st.markdown(hero,unsafe_allow_html=True)

    posts=[
        f'<div class="sticky-card s-pink"><strong>{total}</strong>Scholars<small>Active roster</small></div>',
        f'<div class="sticky-card s-orange"><strong>{counts["On Track"]}</strong>On Track<small>{counts["Approaching"]} approaching • {counts["At Risk"]} at risk</small></div>',
        f'<div class="sticky-card s-yellow"><strong>{assignment_count}</strong>Assignments<small>{html.escape(latest)}</small></div>',
        f'<div class="sticky-card s-green"><strong>{contact_count}</strong>Family Contacts<small>Logged communication</small></div>',
        f'<div class="sticky-card s-blue"><strong>{counts["No Data"]}</strong>No Data<small>Needs more grade evidence</small></div>',
        f'<div class="sticky-card s-purple"><strong>{counts["Approaching"]+counts["At Risk"]}</strong>Need Attention<small>Approaching + at risk</small></div>',
    ]
    st.markdown('<div class="sticky-grid">'+"".join(posts)+'</div>',unsafe_allow_html=True)

    left,right=st.columns(2)
    with left:
        st.markdown("### Scholars to Watch")
        if detail.empty:
            st.caption("No scholar data yet.")
        else:
            watch=detail[detail.Status.isin(["Approaching","At Risk"])].copy()
            if watch.empty:
                st.caption("No scholars are currently flagged.")
            else:
                st.dataframe(watch,hide_index=True,use_container_width=True)

    with right:
        st.markdown("### Recent Assignments")
        if recent.empty:
            st.caption("No assignments yet.")
        else:
            cols=[c for c in ["assignment_date","title","subject","standard_code"] if c in recent.columns]
            st.dataframe(recent.head(6)[cols],hide_index=True,use_container_width=True)

elif page=="Scholars":
    st.markdown('<div class="page-title">Scholars</div><div class="page-subtitle">Manage rosters and open scholar profiles.</div>',unsafe_allow_html=True)

    st.markdown("### Add Scholars")
    add_mode=st.radio(
        "How would you like to add scholars?",
        ["Manual Entry","Roster Spreadsheet"],
        horizontal=True,
        key="scholar_add_mode"
    )

    if add_mode=="Manual Entry":
        class_choices=classes_df()
        if class_choices.empty:
            st.info("Create a class from Home Page → Settings before adding scholars manually.")
        else:
            class_ids=list(class_choices.id.astype(int))
            default_class=selected_class if selected_class in class_ids else class_ids[0]
            with st.form("manual_add_scholar",clear_on_submit=True):
                r1c1,r1c2,r1c3=st.columns(3)
                first=r1c1.text_input("First Name")
                last=r1c2.text_input("Last Name")
                class_id=r1c3.selectbox(
                    "Class",
                    class_ids,
                    format_func=lambda x:class_choices[class_choices.id==x].iloc[0].class_name,
                    index=class_ids.index(default_class)
                )

                r2c1,r2c2,r2c3=st.columns(3)
                student_id=r2c1.text_input("School ID")
                grade_level=r2c2.text_input("Grade Level",value="3")
                academic_year=r2c3.text_input("Academic Year",value=current_academic_year())

                school_name=st.text_input("School Name")
                address=st.text_input("Address")
                a1,a2,a3=st.columns(3)
                city=a1.text_input("City")
                state=a2.text_input("State")
                zip_code=a3.text_input("ZIP Code")
                residency=st.text_input("Residency")

                if st.form_submit_button("Add Scholar"):
                    if not first.strip() or not last.strip():
                        st.error("First and last name are required.")
                    else:
                        cname=class_name_from_id(class_id)
                        c=conn()
                        duplicate=None
                        if student_id.strip():
                            duplicate=c.execute(
                                "SELECT id FROM scholars WHERE student_id=? AND student_id<>''",
                                (student_id.strip(),)
                            ).fetchone()
                        if not duplicate:
                            duplicate=c.execute(
                                """SELECT id FROM scholars
                                   WHERE lower(first_name)=lower(?)
                                     AND lower(last_name)=lower(?)
                                     AND class_id=?""",
                                (first.strip(),last.strip(),int(class_id))
                            ).fetchone()

                        if duplicate:
                            c.close()
                            st.warning("A matching scholar already exists in the roster.")
                        else:
                            c.execute("""INSERT INTO scholars(
                                first_name,last_name,class_name,class_id,school_name,academic_year,
                                grade_level,student_id,address,city,state_code,zip_code,residency,active)
                                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,1)""",
                                (first.strip(),last.strip(),cname,int(class_id),school_name.strip(),
                                 academic_year.strip(),grade_level.strip(),student_id.strip(),
                                 address.strip(),city.strip(),state.strip(),zip_code.strip(),residency.strip()))
                            c.commit(); c.close()
                            st.success(f"{first.strip()} {last.strip()} added to {cname}.")
                            st.rerun()

    else:
        if st.button("📥 Import Roster Spreadsheet", key="show_roster_import"):
            st.session_state["show_roster_importer"]=not st.session_state.get("show_roster_importer",False)
        if st.session_state.get("show_roster_importer",False):
            st.markdown("### Roster Spreadsheet Importer")
            st.write("Upload an Excel or CSV roster. The importer can create scholar profiles and attach parent/guardian contact information.")
            up=st.file_uploader("Roster file",type=["xlsx","xls","csv"])
            if up:
                try:
                    if up.name.lower().endswith(".csv"):
                        df=pd.read_csv(up,dtype=str)
                    else:
                        df=pd.read_excel(up,dtype=str)
                    df=df.fillna("")
                    st.success(f"Loaded {len(df)} rows and {len(df.columns)} columns.")
                    st.dataframe(df.head(10),use_container_width=True,hide_index=True)

                    detected=auto_mapping(df.columns)
                    st.markdown("### Column matching")
                    st.caption("I matched the likely columns automatically. Change any dropdown that is incorrect.")
                    none="— Not in this file —"
                    opts=[none]+list(df.columns)
                    labels=[
                        ("school_name","School Name"),("academic_year","Academic Year"),("grade_level","Grade Level"),
                        ("class_name","Class / Course / Section"),("student_id","Student ID"),
                        ("student_first","Scholar First Name"),("student_last","Scholar Last Name"),
                        ("address","Address"),("city","City"),("state_code","State"),("zip_code","Zip Code"),
                        ("residency","Residency"),("guardian_first","Guardian First Name"),("guardian_last","Guardian Last Name"),
                        ("relationship","Relationship"),("home_phone","Home Phone"),("work_phone","Work Phone"),
                        ("cell_phone","Cell Phone"),("email","Guardian Email")
                    ]
                    chosen={}
                    left,right=st.columns(2)
                    for i,(key,label) in enumerate(labels):
                        container=left if i%2==0 else right
                        default=detected.get(key,none)
                        idx=opts.index(default) if default in opts else 0
                        chosen[key]=container.selectbox(label,opts,index=idx,key=f"map_{key}")

                    fallback_class=st.text_input("If class/section is missing, place scholars in this class folder",value="")
                    update_existing=st.checkbox("Update matching existing scholar profiles",value=True)
                    st.info("Duplicate guardian rows are combined under the same scholar. A scholar is matched primarily by Student ID when available; otherwise by first name + last name + class.")

                    if st.button("Import Scholar Profiles"):
                        if chosen["student_first"]==none or chosen["student_last"]==none:
                            st.error("Choose the Scholar First Name and Scholar Last Name columns.")
                        else:
                            c=conn(); cur=c.cursor()
                            scholar_count=0; guardian_count=0; updated=0
                            for _,r in df.iterrows():
                                def val(key):
                                    col=chosen[key]
                                    return "" if col==none else clean(r[col])
                                sf,sl=val("student_first"),val("student_last")
                                if not sf or not sl: continue
                                sid_text=val("student_id")
                                class_name=val("class_name") or fallback_class.strip()
                                class_id=None
                                if class_name:
                                    cur.execute("INSERT OR IGNORE INTO classes(class_name) VALUES (?)",(class_name,))
                                    class_id=cur.execute("SELECT id FROM classes WHERE class_name=?",(class_name,)).fetchone()["id"]

                                existing=None
                                if sid_text:
                                    existing=cur.execute("SELECT * FROM scholars WHERE student_id=? AND student_id<>''",(sid_text,)).fetchone()
                                if not existing:
                                    if class_id:
                                        existing=cur.execute("""SELECT * FROM scholars WHERE lower(first_name)=lower(?) AND lower(last_name)=lower(?) AND class_id=?""",
                                                             (sf,sl,class_id)).fetchone()
                                    else:
                                        existing=cur.execute("""SELECT * FROM scholars WHERE lower(first_name)=lower(?) AND lower(last_name)=lower(?)""",(sf,sl)).fetchone()

                                fields=(sf,sl,class_name,class_id,val("school_name"),val("academic_year"),val("grade_level"),
                                        sid_text,val("address"),val("city"),val("state_code"),val("zip_code"),val("residency"))
                                if existing:
                                    scholar_id=existing["id"]
                                    if update_existing:
                                        cur.execute("""UPDATE scholars SET first_name=?,last_name=?,class_name=?,class_id=?,
                                          school_name=?,academic_year=?,grade_level=?,student_id=?,address=?,city=?,state_code=?,zip_code=?,residency=?,active=1
                                          WHERE id=?""",fields+(scholar_id,))
                                        updated+=1
                                else:
                                    cur.execute("""INSERT INTO scholars(first_name,last_name,class_name,class_id,school_name,academic_year,grade_level,
                                      student_id,address,city,state_code,zip_code,residency) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",fields)
                                    scholar_id=cur.lastrowid; scholar_count+=1

                                gf,gl,rel=val("guardian_first"),val("guardian_last"),val("relationship")
                                hp,wp,cp,em=val("home_phone"),val("work_phone"),val("cell_phone"),val("email")
                                if gf or gl or hp or wp or cp or em:
                                    cur.execute("""INSERT OR IGNORE INTO guardians
                                      (scholar_id,first_name,last_name,relationship,home_phone,work_phone,cell_phone,email)
                                      VALUES (?,?,?,?,?,?,?,?)""",(scholar_id,gf,gl,rel,hp,wp,cp,em))
                                    # update contact fields if same guardian already exists
                                    cur.execute("""UPDATE guardians SET home_phone=?,work_phone=?,cell_phone=?,email=?
                                      WHERE scholar_id=? AND first_name=? AND last_name=? AND relationship=?""",
                                      (hp,wp,cp,em,scholar_id,gf,gl,rel))
                                    guardian_count+=1
                            c.commit(); c.close()
                            st.success(f"Import complete: {scholar_count} new scholars, {updated} profiles updated, and guardian/contact rows processed.")
                            st.rerun()
                except Exception as e:
                    st.error(f"I could not read that spreadsheet: {e}")

    roster=scholars_df(selected_class or None)
    st.dataframe(roster[["first_name","last_name","grade_level","display_class","student_id"]],hide_index=True,use_container_width=True)

    if not roster.empty:
        st.markdown("### Scholars")
        st.caption("Click a scholar's name to open the profile.")
        cols_per_row=3
        for i in range(0,len(roster),cols_per_row):
            rowcols=st.columns(cols_per_row)
            chunk=roster.iloc[i:i+cols_per_row]
            for j,(_,sr) in enumerate(chunk.iterrows()):
                rowcols[j].button(
                    nm(sr),
                    key=f"open_profile_{int(sr.id)}",
                    use_container_width=True,
                    on_click=open_scholar_profile,
                    args=(int(sr.id),)
                )

    if not roster.empty:
        st.markdown("### Remove from roster")
        rid=st.selectbox("Scholar",list(roster.id.astype(int)),format_func=lambda x:nm(roster[roster.id==x].iloc[0]))
        a,b=st.columns(2)
        if a.button("Archive Scholar"):
            c=conn(); c.execute("UPDATE scholars SET active=0 WHERE id=?",(rid,)); c.commit(); c.close(); st.rerun()
        ok=b.checkbox("I understand permanent delete cannot be undone")
        if b.button("Delete Permanently",disabled=not ok):
            c=conn()
            for t in ["grades","communications","work_samples","support_notes","report_comments","guardians"]:
                c.execute(f"DELETE FROM {t} WHERE scholar_id=?",(rid,))
            c.execute("DELETE FROM scholars WHERE id=?",(rid,)); c.commit(); c.close(); st.rerun()

elif page=="Scholar Profile":
    sid=st.session_state.get("selected_profile_scholar")
    if not sid:
        st.warning("Choose a scholar from the Scholars page.")
        st.button("Back to Scholars",on_click=return_to_scholars)
    else:
        s=scholar_full_profile(sid)
        if not s:
            st.warning("That scholar profile could not be found.")
            st.button("Back to Scholars",on_click=return_to_scholars)
        else:
            st.button("← Back to Scholars",key="back_to_scholars_profile",on_click=return_to_scholars)
            initials=((s["first_name"][:1] if s["first_name"] else "")+(s["last_name"][:1] if s["last_name"] else "")).upper()

            st.markdown('<div class="profile-cover"></div>',unsafe_allow_html=True)
            profile_html=(
                '<div class="profile-header">'
                f'<div class="profile-avatar">{initials}</div>'
                f'<h2 style="margin:12px 0 2px 0">{html.escape(s["first_name"])} {html.escape(s["last_name"])}</h2>'
                f'<div>School ID: {html.escape(str(s["student_id"] or "—"))} | Grade {html.escape(str(s["grade_level"] or "—"))} | {html.escape(str(s["display_class"] or "No Class"))}</div>'
                f'<div style="margin-top:5px">{html.escape(str(s["school_name"] or ""))}</div>'
                f'<div style="margin-top:5px;color:#667085;">Gender: {html.escape(str(s["gender"] or "—"))} | Pronouns: {html.escape(str(s["pronouns"] or "—"))}</div>'
                '</div>'
            )
            st.markdown(profile_html,unsafe_allow_html=True)

            # Scholar Information
            st.markdown('<div class="profile-section"><h3>Scholar Information</h3></div>',unsafe_allow_html=True)
            class_choices=classes_df()
            class_ids=[0]+list(class_choices.id.astype(int))
            class_map={0:"Unassigned",**{int(r.id):r.class_name for _,r in class_choices.iterrows()}}
            current_class=int(s["class_id"]) if s["class_id"] and int(s["class_id"]) in class_ids else 0
            with st.form("profile_edit_continuous"):
                a,b=st.columns(2)
                ef=a.text_input("First Name",value=s["first_name"] or "")
                el=b.text_input("Last Name",value=s["last_name"] or "")
                c1,c2,c3=st.columns(3)
                eid=c1.text_input("School ID",value=s["student_id"] or "")
                egrade=c2.text_input("Grade Level",value=s["grade_level"] or "")
                eclass=c3.selectbox("Class",class_ids,format_func=lambda x:class_map[x],index=class_ids.index(current_class))

                g1,g2=st.columns(2)
                gender_options=["","Female","Male","Nonbinary","Other / Prefer to self-describe","Prefer not to say"]
                current_gender=str(s["gender"] or "")
                gender_index=gender_options.index(current_gender) if current_gender in gender_options else 0
                egender=g1.selectbox("Gender",gender_options,index=gender_index)
                pronoun_options=["","she/her","he/him","they/them","Custom"]
                current_pronouns=str(s["pronouns"] or "")
                pronoun_index=pronoun_options.index(current_pronouns) if current_pronouns in pronoun_options else (4 if current_pronouns else 0)
                epronoun_choice=g2.selectbox("Pronouns",pronoun_options,index=pronoun_index)
                custom_pronouns=st.text_input(
                    "Custom pronouns",
                    value=current_pronouns if epronoun_choice=="Custom" else "",
                    placeholder="Example: ze/hir/hir"
                )
                epronouns=custom_pronouns.strip() if epronoun_choice=="Custom" else epronoun_choice

                school=st.text_input("School Name",value=s["school_name"] or "")
                year=st.text_input("Academic Year",value=s["academic_year"] or "")
                if st.form_submit_button("Save Scholar Changes"):
                    c=conn()
                    cname=class_map[eclass] if eclass else ""
                    c.execute("""UPDATE scholars SET first_name=?,last_name=?,student_id=?,grade_level=?,
                                 class_id=?,class_name=?,school_name=?,academic_year=?,gender=?,pronouns=? WHERE id=?""",
                              (ef.strip(),el.strip(),eid.strip(),egrade.strip(),
                               eclass if eclass else None,cname,school.strip(),year.strip(),egender,epronouns,int(sid)))
                    c.commit(); c.close()
                    st.success("Scholar information updated.")
                    st.rerun()

            # Benchmarks
            br=benchmark_for_scholar(sid)
            st.markdown("### Benchmark Snapshot")
            b1,b2,b3=st.columns(3)
            b1.metric("F&P", (br["fp_spring_level"] or br["fp_fall_level"] or "—") if br else "—")
            b2.metric("NWEA Reading", (br["nwea_spring_reading"] or br["nwea_winter_reading"] or br["nwea_fall_reading"] or "—") if br else "—")
            b3.metric("NWEA Math", (br["nwea_spring_math"] or br["nwea_winter_math"] or br["nwea_fall_math"] or "—") if br else "—")

            # Grades
            st.markdown("### Current Grades")
            cols=st.columns(4)
            for i,subj in enumerate(["ELA","Math","Science","Social Studies"]):
                av,_=summary(sid,subj)
                cols[i].metric(subj,"No grades" if av is None else f"{av:.1f}% ({letter(av)})")

            # Academic summary
            strengths,needs,teacher_actions=academic_summary_for_scholar(sid)
            st.markdown("### Academic Summary")
            left,right=st.columns(2)
            with left:
                st.markdown("**Strengths**")
                if strengths:
                    for subj,skill,pct in strengths:
                        st.write(f"- {subj}: {skill} — {pct:.0f}%")
                else:
                    st.caption("Not enough skill data yet.")
            with right:
                st.markdown("**Needs Support**")
                if needs:
                    for subj,skill,pct in needs:
                        st.write(f"- {subj}: {skill} — {pct:.0f}%")
                else:
                    st.caption("No below-mastery skill pattern identified.")
            st.markdown("**Teacher Support Plan**")
            for action in teacher_actions:
                st.write(f"- {action}")

            # Parent contacts
            st.markdown("### Parent / Guardian Contacts")
            c=conn()
            g=pd.read_sql_query("SELECT * FROM guardians WHERE scholar_id=? ORDER BY relationship,last_name,first_name",c,params=[sid])
            c.close()
            if g.empty:
                st.caption("No parent/guardian contacts saved yet.")
            else:
                display_g=g.copy()
                display_g["Name"]=(display_g["first_name"].fillna("")+" "+display_g["last_name"].fillna("")).str.strip()
                display_g["Phone"]=display_g["cell_phone"].where(display_g["cell_phone"].fillna("")!="",display_g["home_phone"])
                st.dataframe(display_g[["Name","relationship","Phone","email"]],hide_index=True,use_container_width=True)

            with st.expander("Add Parent / Guardian"):
                with st.form("profile_add_guardian_cont",clear_on_submit=True):
                    a,b=st.columns(2)
                    gf=a.text_input("First Name")
                    gl=b.text_input("Last Name")
                    c1,c2=st.columns(2)
                    rel=c1.text_input("Relationship")
                    phone=c2.text_input("Phone Number")
                    email=st.text_input("Email")
                    if st.form_submit_button("Add Contact"):
                        if not gf.strip() and not gl.strip():
                            st.error("Enter at least a first or last name.")
                        else:
                            c=conn()
                            c.execute("""INSERT OR IGNORE INTO guardians
                                (scholar_id,first_name,last_name,relationship,cell_phone,email)
                                VALUES (?,?,?,?,?,?)""",
                                (int(sid),gf.strip(),gl.strip(),rel.strip(),phone.strip(),email.strip()))
                            c.commit(); c.close(); st.rerun()

            if not g.empty:
                with st.expander("Edit / Delete Parent Contact"):
                    gid=st.selectbox("Choose Contact",list(g.id.astype(int)),
                                     format_func=lambda x:(f"{g[g.id==x].iloc[0].first_name} {g[g.id==x].iloc[0].last_name}".strip()),
                                     key="profile_guardian_select_cont")
                    gr=g[g.id==gid].iloc[0]
                    with st.form("profile_edit_guardian_cont"):
                        a,b=st.columns(2)
                        egf=a.text_input("First Name",value=gr.first_name or "")
                        egl=b.text_input("Last Name",value=gr.last_name or "")
                        c1,c2=st.columns(2)
                        erel=c1.text_input("Relationship",value=gr.relationship or "")
                        ephone=c2.text_input("Phone Number",value=(gr.cell_phone or gr.home_phone or ""))
                        eemail=st.text_input("Email",value=gr.email or "")
                        if st.form_submit_button("Save Contact Changes"):
                            c=conn()
                            c.execute("""UPDATE guardians SET first_name=?,last_name=?,relationship=?,
                                         cell_phone=?,home_phone='',work_phone='',email=? WHERE id=?""",
                                      (egf.strip(),egl.strip(),erel.strip(),ephone.strip(),eemail.strip(),int(gid)))
                            c.commit(); c.close(); st.rerun()
                    confirm=st.checkbox("Delete this contact",key="profile_del_guardian_confirm")
                    if st.button("Delete Parent / Guardian",disabled=not confirm,key="profile_del_guardian"):
                        c=conn()
                        c.execute("UPDATE communications SET guardian_id=NULL WHERE guardian_id=?",(int(gid),))
                        c.execute("UPDATE contact_reminders SET guardian_id=NULL WHERE guardian_id=?",(int(gid),))
                        c.execute("UPDATE parent_update_preferences SET preferred_guardian_id=NULL WHERE preferred_guardian_id=?",(int(gid),))
                        c.execute("DELETE FROM guardians WHERE id=?",(int(gid),))
                        c.commit(); c.close(); st.rerun()

            # History / reports / support / reminders
            st.markdown("### Parent Contact History")
            c=conn()
            hist=pd.read_sql_query("""SELECT communications.created_at,
                 COALESCE(TRIM(guardians.first_name||' '||guardians.last_name),'') guardian,
                 guardians.relationship,communications.communication_type,
                 communications.subject,communications.reason,communications.details
                 FROM communications LEFT JOIN guardians ON guardians.id=communications.guardian_id
                 WHERE communications.scholar_id=? ORDER BY communications.id DESC""",c,params=[sid])
            rc=pd.read_sql_query("SELECT created_at,subject,marking_period,comment_text FROM report_comments WHERE scholar_id=? ORDER BY id DESC",c,params=[sid])
            sn=pd.read_sql_query("SELECT created_at,note_type,area,observation,intervention,response_to_intervention,impact FROM support_notes WHERE scholar_id=? ORDER BY id DESC",c,params=[sid])
            c.close()

            if hist.empty: st.caption("No parent contacts logged yet.")
            else: st.dataframe(hist,hide_index=True,use_container_width=True)

            st.markdown("### Saved Report Card Comments")
            if rc.empty: st.caption("No saved report card comments yet.")
            else: st.dataframe(rc,hide_index=True,use_container_width=True)

            st.markdown("### Support / IEP Evidence")
            if sn.empty: st.caption("No support notes saved yet.")
            else: st.dataframe(sn,hide_index=True,use_container_width=True)

            st.markdown("### Parent Update Reminders")
            rem=reminder_rows(sid)
            if rem.empty: st.caption("No reminders.")
            else: st.dataframe(rem[["due_date","guardian","reason","notes","completed"]],hide_index=True,use_container_width=True)

elif page=="Scholar Binder":
    st.markdown('<div class="page-title">Grades</div><div class="page-subtitle">Assignments, gradebook, standards, NWEA, F&P, I-Ready, Interims, and grade settings.</div>',unsafe_allow_html=True)

    # Grades is always visible, even before a class or scholar has been added.
    binder_tool=st.radio(
        "Gradebook Section",
        ["Overview","Add Assignment","Skills & Standards","Work Samples","NWEA","F&P","I-Ready","Interims","Grade Settings"],
        horizontal=True,
        key="class_binder_tool"
    )

    if not selected_class:
        st.caption("No class is selected yet. This section is ready now; class and scholar data will populate here automatically once you add them.")

        if binder_tool=="Overview":
            st.markdown("## Class Gradebook Overview")
            p1,p2=st.columns(2)
            p1.selectbox("Subject",["All Subjects","ELA","Math","Science","Social Studies"],key="preview_grid_subject",disabled=True)
            p2.selectbox("Sort assignments by",["Oldest → Newest","Newest → Oldest","Subject","Standard Code"],key="preview_grid_sort",disabled=True)

            preview_df=pd.DataFrame(columns=["Scholar"])
            st.dataframe(preview_df,hide_index=True,use_container_width=True,height=300)
            st.info("Scholar names and assignment columns will appear in this gradebook automatically after a class and roster are added.")

            st.markdown("### Printable / Uploadable Excel Grade Sheet")
            st.caption("The download and import tools will activate once a class exists.")
            st.button("⬇️ Download Excel Grade Sheet",disabled=True,key="preview_download_grade_sheet")

            st.markdown("---")
            st.markdown("### 📸 Import Grades from NHA Screenshot")
            st.caption("Screenshot grade import will activate once a class and at least one assignment exist.")
            st.file_uploader("Upload NHA gradebook screenshot(s)",type=["png","jpg","jpeg"],disabled=True,key="preview_nha_upload")

        elif binder_tool=="Add Assignment":
            st.markdown("## Add Assignment")
            st.caption("You can see the assignment setup now. Saving will activate once a class exists.")
            a,b=st.columns(2)
            a.text_input("Assignment title",disabled=True,key="preview_assignment_title")
            b.selectbox("Grade category",list(get_setting("weights").keys()),disabled=True,key="preview_assignment_category")
            c,d=st.columns(2)
            c.selectbox("Subject",["ELA","Math","Science","Social Studies"],disabled=True,key="preview_assignment_subject")
            d.number_input("Points Possible",min_value=1.0,value=100.0,disabled=True,key="preview_assignment_points")
            st.selectbox("Skill / Standard",["Select a class first"],disabled=True,key="preview_assignment_standard")
            st.date_input("Assignment Date",value=date.today(),disabled=True,key="preview_assignment_date")
            st.button("Create Assignment",disabled=True,key="preview_create_assignment")

        elif binder_tool=="Skills & Standards":
            st.markdown("## Skills & Standards")
            st.caption("Standards are available even before a roster is added.")
            preview_subject=st.selectbox(
                "Subject",
                ["Science","Social Studies"],
                key="preview_standard_subject"
            )
            sdf=standards_df(preview_subject)
            if sdf.empty:
                st.info("No standards are saved for this subject yet.")
            else:
                st.dataframe(
                    sdf[[c for c in ["code","skill","description"] if c in sdf.columns]],
                    hide_index=True,use_container_width=True
                )
            st.caption("Scholar skill performance will appear here after grades are entered.")

        elif binder_tool=="Work Samples":
            st.markdown("## Work Samples")
            st.file_uploader("Upload Scholar Work",disabled=True,key="preview_work_sample")
            st.info("The scholar selector and saved work samples will appear here after a roster is added.")

        elif binder_tool=="NWEA":
            render_nwea_center(None)

        elif binder_tool=="F&P":
            render_fp_center(None)

        elif binder_tool=="I-Ready":
            render_iready_center(None)

        elif binder_tool=="Interims":
            render_interim_center(None)

        elif binder_tool=="Grade Settings":
            st.markdown("## Grade Settings")
            st.caption("These settings are available before a class exists.")
            weights=get_setting("weights") or {}
            if weights:
                st.dataframe(
                    pd.DataFrame([{"Category":k,"Weight %":v} for k,v in weights.items()]),
                    hide_index=True,use_container_width=True
                )
            scale=get_setting("scale") or []
            if scale:
                st.markdown("### Grade Scale")
                st.dataframe(
                    pd.DataFrame(scale,columns=["Grade","Minimum","Maximum"]),
                    hide_index=True,use_container_width=True
                )

    else:
        class_name=folder.get(selected_class,"Selected Class")
        st.markdown(f"### 📌 {class_name}")

        if binder_tool=="Overview":
            st.markdown("## Class Gradebook Overview")
            f1,f2=st.columns(2)
            subject_filter=f1.selectbox("Subject",["All Subjects","ELA","Math","Science","Social Studies"],key="grid_subject_filter")
            sort_order=f2.selectbox("Sort assignments by",["Oldest → Newest","Newest → Oldest","Subject","Standard Code"],key="grid_sort")
            matrix,assignments=gradebook_matrix(selected_class,subject_filter,sort_order)
            if matrix.empty: st.info("No scholars are available for this class.")
            else: st.dataframe(matrix,hide_index=True,use_container_width=True,height=520)

            st.markdown("### Printable / Uploadable Excel Grade Sheet")
            st.write("Download the sheet, print it or type grades into it, then upload the completed Excel file to enter grades in bulk.")
            excel_bytes=make_grade_sheet_xlsx(selected_class,class_name,subject_filter,sort_order)
            st.download_button("⬇️ Download Excel Grade Sheet",data=excel_bytes,file_name=f"{class_name.replace(' ','_')}_Grade_Entry.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",key="download_grade_sheet")
            uploaded_grade_sheet=st.file_uploader("Upload completed Excel grade sheet",type=["xlsx"],key="upload_completed_grade_sheet")
            if uploaded_grade_sheet is not None and st.button("Import Grades from Excel",key="import_grades_xlsx"):
                try:
                    saved,skipped=import_grade_sheet_xlsx(uploaded_grade_sheet)
                    st.success(f"Imported {saved} grade entries. Skipped {skipped} invalid entries.")
                    st.rerun()
                except Exception as e:
                    st.error(f"Could not import the grade sheet: {e}")

            st.markdown("---")
            st.markdown("### 📸 Import Grades from NHA Screenshot")
            st.caption(
                "Upload a screenshot that shows scholar names and their entered grades. "
                "ChapLab reads the screenshot locally, then gives you an editable preview before saving anything."
            )

            screenshot_subject=st.selectbox(
                "Subject for screenshot",
                ["ELA","Math","Science","Social Studies"],
                key="nha_ss_subject"
            )
            screenshot_assignments=class_assignments_df(selected_class,screenshot_subject,"Oldest → Newest")

            if screenshot_assignments.empty:
                st.info("Create the assignment(s) in ChapLab first so the screenshot grades have somewhere to go.")
            else:
                assign_map={
                    int(r.id):f"{r.title} | {r.standard_code or 'No Standard'} | {r.assignment_date} | {r.points_possible:g} pts"
                    for _,r in screenshot_assignments.iterrows()
                }
                selected_aids=st.multiselect(
                    "Assignments visible in the screenshot — choose them in LEFT-TO-RIGHT order",
                    list(assign_map.keys()),
                    format_func=lambda x:assign_map[x],
                    key="nha_ss_assignments"
                )

                uploaded_screens=st.file_uploader(
                    "Upload NHA gradebook screenshot(s)",
                    type=["png","jpg","jpeg"],
                    accept_multiple_files=True,
                    key="nha_grade_screens"
                )

                st.info(
                    "For best results: crop the screenshot so the scholar names and selected grade columns are visible, "
                    "keep the text at normal zoom or larger, and avoid covering rows with pop-ups."
                )

                if uploaded_screens and selected_aids:
                    selected_rows=[]
                    for aid in selected_aids:
                        ar=screenshot_assignments[screenshot_assignments.id==aid].iloc[0]
                        selected_rows.append({
                            "id":int(aid),
                            "title":ar.title,
                            "points_possible":float(ar.points_possible)
                        })

                    if st.button("Read Screenshot Grades",key="read_nha_screenshot"):
                        all_found=[]
                        messages=[]
                        roster=scholars_df(selected_class)
                        for screen in uploaded_screens:
                            try:
                                found,msg=ocr_grade_screenshot(screen,roster,selected_rows)
                                messages.append(f"{screen.name}: {msg}")
                                if not found.empty:
                                    all_found.append(found)
                            except Exception as e:
                                messages.append(f"{screen.name}: {e}")
                        if all_found:
                            combined=pd.concat(all_found,ignore_index=True)
                            combined=combined.sort_values("OCR Match %",ascending=False).drop_duplicates("Scholar ID").sort_values("Scholar")
                            st.session_state["nha_ocr_preview"]=combined
                            st.session_state["nha_ocr_assignment_ids"]=selected_aids
                        st.session_state["nha_ocr_messages"]=messages

                for msg in st.session_state.get("nha_ocr_messages",[]):
                    st.write(f"- {msg}")

                if "nha_ocr_preview" in st.session_state:
                    preview=st.session_state["nha_ocr_preview"].copy()
                    st.markdown("#### Review Before Import")
                    st.warning("OCR can make mistakes. Check scholar names and every grade before importing.")
                    edited_preview=st.data_editor(
                        preview,
                        disabled=["Scholar ID","Scholar","OCR Match %"],
                        hide_index=True,
                        use_container_width=True,
                        key="nha_ocr_editor"
                    )

                    current_ids=st.session_state.get("nha_ocr_assignment_ids",[])
                    col_to_aid={}
                    for aid in current_ids:
                        match=screenshot_assignments[screenshot_assignments.id==aid]
                        if not match.empty:
                            ar=match.iloc[0]
                            col_to_aid[f"{ar.title} ({ar.points_possible:g} pts)"]=int(aid)

                    if st.button("Import Reviewed Screenshot Grades",key="import_nha_ocr_grades"):
                        saved,skipped=import_ocr_preview(edited_preview,col_to_aid)
                        st.success(f"Imported {saved} reviewed grade entries. Skipped {skipped} invalid entries.")
                        st.session_state.pop("nha_ocr_preview",None)
                        st.session_state.pop("nha_ocr_messages",None)
                        st.session_state.pop("nha_ocr_assignment_ids",None)
                        st.rerun()

            with st.expander("Local screenshot reader setup"):
                st.write(
                    "The screenshot reader uses Tesseract OCR installed on your Windows computer. "
                    "The image is processed locally and is not sent to an outside AI service."
                )
                ok,ocr_status=configure_local_tesseract()
                if ok:
                    st.success(f"Local OCR ready: {ocr_status}")
                else:
                    st.warning(
                        "Tesseract OCR is not installed yet. Install Tesseract OCR for Windows, "
                        "then restart ChapLab. The Excel import continues to work without it."
                    )

        elif binder_tool=="Add Assignment":
            subject=st.selectbox("Subject",["ELA","Math","Science","Social Studies"],key="assignment_subject")
            st.markdown("### Assignment Information")
            with st.form("new_assignment_binder"):
                a,b=st.columns(2)
                title=a.text_input("Assignment title")
                category=b.selectbox("Grade category",list(get_setting("weights").keys()))
                sdf=standards_df(subject)
                skill_choice=st.selectbox("Skill / Standard",[""]+[f"{r.code} — {r.skill}" for _,r in sdf.iterrows()])
                c1,c2=st.columns(2)
                points=c1.number_input("Points possible",1.0,1000.0,100.0)
                adate=c2.date_input("Date",date.today())
                detected_quarter=quarter_for_date(str(adate),current_academic_year())
                marking_period=st.selectbox("Marking Period",
                                            ["Quarter 1","Quarter 2","Quarter 3","Quarter 4"],
                                            index=["Quarter 1","Quarter 2","Quarter 3","Quarter 4"].index(detected_quarter) if detected_quarter in ["Quarter 1","Quarter 2","Quarter 3","Quarter 4"] else 0,
                                            help=f"Detected from saved date ranges: {detected_quarter or 'No match'}")
                if st.form_submit_button("Create Assignment") and title.strip():
                    qsettings=get_quarter_settings(current_academic_year())
                    qrow=qsettings[qsettings.quarter_name==marking_period] if not qsettings.empty else pd.DataFrame()
                    if not qrow.empty and bool(qrow.iloc[0].locked):
                        st.error(f"{marking_period} is locked. Unlock it in Home Page settings before adding assignments.")
                        st.stop()
                    code=skill_choice.split(" — ")[0] if skill_choice else ""
                    c=conn()
                    c.execute("""INSERT INTO assignments(title,subject,category,standard_code,points_possible,assignment_date,class_id,marking_period)
                                 VALUES (?,?,?,?,?,?,?,?)""",(title.strip(),subject,category,code,float(points),str(adate),selected_class,marking_period))
                    c.commit(); c.close(); st.success("Assignment created."); st.rerun()

            assignments=class_assignments_df(selected_class,subject,"Newest → Oldest")
            if not assignments.empty:
                amap={int(r.id):f"{r.title} | {r.marking_period or 'No Quarter'} | {r.standard_code or 'No Standard'} | {r.assignment_date} | {r.points_possible:g} pts" for _,r in assignments.iterrows()}
                aid=st.selectbox("Assignment to enter grades",list(amap),format_func=lambda x:amap[x],key="assignment_grade_entry")
                ar=assignments[assignments.id==aid].iloc[0]

                with st.expander("🗑️ Delete Selected Assignment"):
                    st.warning("Deleting an assignment will also delete every scholar grade attached to that assignment.")
                    confirm_delete=st.checkbox(
                        f"I understand and want to delete: {ar.title}",
                        key="confirm_delete_assignment"
                    )
                    if st.button(
                        "Delete Assignment Permanently",
                        disabled=not confirm_delete,
                        key="delete_assignment_button"
                    ):
                        c=conn()
                        c.execute("DELETE FROM grades WHERE assignment_id=?",(int(aid),))
                        c.execute("DELETE FROM assignments WHERE id=?",(int(aid),))
                        c.commit()
                        c.close()
                        st.success("Assignment and linked grades deleted.")
                        st.rerun()

                entry_mode=st.radio("Grade Entry",["Whole Class","Search One Scholar"],horizontal=True,key="entry_mode")
                roster=scholars_df(selected_class)
                if entry_mode=="Whole Class":
                    c=conn(); current=pd.read_sql_query("SELECT scholar_id,points_earned FROM grades WHERE assignment_id=?",c,params=[aid]); c.close()
                    cmap=dict(zip(current.scholar_id,current.points_earned))
                    grid=pd.DataFrame({"Scholar ID":roster.id.astype(int),"Scholar":[nm(r) for _,r in roster.iterrows()],"Points Earned":[cmap.get(int(r.id),None) for _,r in roster.iterrows()]})
                    edited=st.data_editor(grid,disabled=["Scholar ID","Scholar"],hide_index=True,use_container_width=True)
                    if st.button("Save Class Grades",key="save_full_assignment_grades"):
                        c=conn()
                        for _,r in edited.iterrows():
                            if pd.notna(r["Points Earned"]):
                                score=float(r["Points Earned"])
                                if 0<=score<=float(ar.points_possible):
                                    c.execute("INSERT OR REPLACE INTO grades(scholar_id,assignment_id,points_earned) VALUES (?,?,?)",(int(r["Scholar ID"]),int(aid),score))
                        c.commit(); c.close(); st.success("Class grades saved."); st.rerun()
                else:
                    q=st.text_input("Search scholar",placeholder="Type first or last name")
                    filtered=roster
                    if q.strip(): filtered=filtered[(filtered.first_name+" "+filtered.last_name).str.contains(q.strip(),case=False,na=False)]
                    if not filtered.empty:
                        sid=st.selectbox("Scholar",list(filtered.id.astype(int)),format_func=lambda x:nm(filtered[filtered.id==x].iloc[0]))
                        c=conn(); old=c.execute("SELECT points_earned FROM grades WHERE scholar_id=? AND assignment_id=?",(sid,aid)).fetchone(); c.close()
                        default=float(old["points_earned"]) if old and old["points_earned"] is not None else 0.0
                        score=st.number_input("Points earned",0.0,float(ar.points_possible),default,key="single_scholar_grade")
                        if st.button("Save Scholar Grade",key="save_single_scholar_grade"):
                            c=conn(); c.execute("INSERT OR REPLACE INTO grades(scholar_id,assignment_id,points_earned) VALUES (?,?,?)",(sid,aid,float(score))); c.commit(); c.close(); st.success("Grade saved."); st.rerun()

        elif binder_tool=="Skills & Standards":
            subj=st.selectbox("Subject",["ELA","Math","Science","Social Studies"],key="binder_std_subject_v9")
            sdf=standards_df(subj)
            st.dataframe(sdf[["id","code","skill","description"]],hide_index=True,use_container_width=True)

        elif binder_tool=="Work Samples":
            roster=scholars_df(selected_class)
            if not roster.empty:
                sid=st.selectbox("Scholar",list(roster.id.astype(int)),format_func=lambda x:nm(roster[roster.id==x].iloc[0]),key="ws_student_v9")
                subj=st.selectbox("Subject",["ELA","Math","Science","Social Studies","Other"],key="ws_subject_v9")
                title=st.text_input("Work Sample Title",key="ws_title_v9")
                up=st.file_uploader("Upload PDF, image, DOCX or TXT",type=["pdf","png","jpg","jpeg","docx","txt"],key="ws_upload_v9")
                obs=st.text_area("Teacher Observation",key="ws_obs_v9")
                x,y=st.columns(2); strengths=x.text_area("Strengths",key="ws_strength_v9"); needs=y.text_area("Needs / Patterns",key="ws_needs_v9")
                nxt=st.text_area("Next Steps",key="ws_next_v9")
                if st.button("Save Work Sample",key="ws_save_v9"):
                    fn=fp=""
                    if up:
                        safe_name=datetime.now().strftime("%Y%m%d_%H%M%S_")+re.sub(r'[^A-Za-z0-9._-]','_',up.name)
                        fn=up.name
                        if cloud_configured():
                            remote_path=f"work_samples/{int(sid)}/{safe_name}"
                            content_type=getattr(up,"type",None) or "application/octet-stream"
                            if cloud_upload_bytes(bytes(up.getbuffer()),remote_path,content_type):
                                fp=f"cloud://{remote_path}"
                            else:
                                st.error("The work-sample file could not be saved to cloud storage.")
                                st.stop()
                        else:
                            folderp=UPLOAD_DIR/f"{sid}"; folderp.mkdir(parents=True,exist_ok=True)
                            dest=folderp/safe_name
                            dest.write_bytes(up.getbuffer()); fp=str(dest)
                    c=conn(); c.execute("""INSERT INTO work_samples(scholar_id,uploaded_at,subject,title,file_name,file_path,teacher_observation,strengths,needs,next_steps)
                                         VALUES (?,?,?,?,?,?,?,?,?,?)""",(sid,datetime.now().isoformat(timespec="minutes"),subj,title,fn,fp,obs,strengths,needs,nxt)); c.commit(); c.close(); st.success("Saved."); st.rerun()

        elif binder_tool=="NWEA":
            render_nwea_center(selected_class)

        elif binder_tool=="F&P":
            render_fp_center(selected_class)

        elif binder_tool=="I-Ready":
            render_iready_center(selected_class)

        elif binder_tool=="Interims":
            render_interim_center(selected_class)

        elif binder_tool=="Grade Settings":
            w=pd.DataFrame([{"Category":k,"Weight %":v} for k,v in get_setting("weights").items()])
            e=st.data_editor(w,num_rows="dynamic",hide_index=True,use_container_width=True,key="weights_v9")
            if st.button("Save Weights",key="save_weights_v9"):
                nw={str(r["Category"]):float(r["Weight %"]) for _,r in e.iterrows() if str(r["Category"]).strip()}
                if abs(sum(nw.values())-100)>.01: st.error("Weights must total 100%.")
                else: save_setting("weights",nw); st.success("Saved."); st.rerun()
            sc=pd.DataFrame(get_setting("scale"),columns=["Letter","Minimum","Maximum"])
            se=st.data_editor(sc,num_rows="dynamic",hide_index=True,use_container_width=True,key="scale_v9")
            if st.button("Save Letter Scale",key="save_scale_v9"):
                save_setting("scale",[[str(r.Letter),float(r.Minimum),float(r.Maximum)] for _,r in se.iterrows()]); st.success("Saved.")
elif page=="Book Leveler":
    st.markdown(
        '<div class="page-title">Book Leveler</div>'
        '<div class="page-subtitle">Start with a book or start with a scholar.</div>',
        unsafe_allow_html=True
    )
    roster=scholars_df(selected_class or None)
    tabs=st.tabs(["🔎 Book Finder","👩🏽‍🎓 Scholar Fit","📚 My Book Catalog"])

    with tabs[0]:
        _pending_query=st.session_state.pop("_book_pending_query",None)
        if _pending_query is not None:
            st.session_state["book_online_query"]=str(_pending_query)

        _pending_lookup=st.session_state.pop("_book_pending_lookup_isbn",None)
        if _pending_lookup:
            with st.spinner("Book Finder is checking reading-level sources..."):
                researched=internet_book_lookup_isbn(_pending_lookup)
            st.session_state["book_online_results"]=[researched] if researched else []
            if researched:
                st.session_state["_book_scan_message"]=(
                    f"✅ ISBN **{_pending_lookup}** researched automatically. "
                    f"**{researched.get('title','Book')}** loaded."
                )

        st.markdown("### 🔎 Book Finder")
        st.caption(
            "Start with a book. Enter, upload, or scan the ISBN. "
            "ChapLab researches its reading level and shows which scholars may struggle, "
            "which scholars should be okay with support, and which should have little to no difficulty."
        )

        def clear_book_scanner():
            for k in (
                "book_camera_isbn","book_camera_manual_correction","book_online_query",
                "book_online_results","book_selected_result","book_selected_result_idx",
                "_book_scan_message","_book_scan_error","_last_camera_scan_token",
                "_last_live_scanned_value","_last_autolookup_isbn",
                "_book_pending_query","_book_pending_lookup_isbn","book_verified_fp_online",
                "_bookfinder_enriched_key","_bookfinder_enriched_book"
            ):
                st.session_state.pop(k,None)
            st.session_state["book_online_query"]=""
            st.session_state["book_camera_manual_correction"]=""
            st.session_state["book_camera_isbn"]=""
            st.session_state["book_online_results"]=[]
            st.session_state["book_online_search_mode"]="ISBN"
            st.session_state["book_live_scanner_nonce"]=int(st.session_state.get("book_live_scanner_nonce",0))+1

        identify_mode=st.radio(
            "Find book by",
            ["Upload Barcode Photo","Camera"],
            horizontal=True,key="book_identify_mode"
        )

        if identify_mode=="Camera":
            c1,c2=st.columns([3,1])
            c1.markdown("#### 📷 Automatic ISBN Scanner")
            if c2.button("🧹 Clear / New Book",key="clear_camera_book",use_container_width=True):
                clear_book_scanner(); st.rerun()
            nonce=int(st.session_state.get("book_live_scanner_nonce",0))
            scanned=chaplab_auto_barcode_scanner(key=f"chaplab_auto_barcode_{nonce}")
            if scanned:
                raw=str(scanned).strip()
                cleaned=re.sub(r"[^0-9Xx]","",raw)
                if raw!=st.session_state.get("_last_live_scanned_value"):
                    st.session_state["_last_live_scanned_value"]=raw
                    if len(cleaned) in (10,13):
                        st.session_state["book_camera_manual_correction"]=cleaned
                        st.session_state["book_online_search_mode"]="ISBN"
                        queue_book_isbn_lookup(cleaned)
                        st.rerun()
                    else:
                        st.session_state["_book_scan_error"]=f"`{raw}` was not read as a valid ISBN."
                        st.rerun()
            if st.session_state.get("_book_scan_error"):
                st.warning(st.session_state.pop("_book_scan_error"))
            if st.session_state.get("_book_scan_message"):
                st.success(st.session_state["_book_scan_message"])

            detected=st.session_state.get("book_camera_isbn","")
            if detected and st.session_state.get("book_camera_manual_correction")!=detected:
                st.session_state["book_camera_manual_correction"]=detected

            def _bf_manual_changed():
                cleaned=re.sub(r"[^0-9Xx]","",st.session_state.get("book_camera_manual_correction",""))
                if len(cleaned) in (10,13) and cleaned!=st.session_state.get("_last_autolookup_isbn"):
                    st.session_state["book_online_search_mode"]="ISBN"
                    queue_book_isbn_lookup(cleaned)

            st.text_input(
                "Detected ISBN / manual correction",
                placeholder="978...",key="book_camera_manual_correction",
                on_change=_bf_manual_changed
            )

        elif identify_mode=="Upload Barcode Photo":
            upload=st.file_uploader(
                "Upload a clear ISBN barcode photo",
                type=["png","jpg","jpeg","webp"],key="book_barcode_photo_upload"
            )
            if upload is not None:
                st.image(upload,width=350)
                if st.button("Read Barcode & Find Book",key="read_uploaded_book_barcode"):
                    detected,error=decode_isbn_barcode(upload)
                    if detected:
                        st.session_state["book_camera_manual_correction"]=detected
                        st.session_state["book_online_search_mode"]="ISBN"
                        queue_book_isbn_lookup(detected)
                        st.rerun()
                    else:
                        st.warning(error)

        st.markdown("#### 🔢 Enter ISBN")
        q=st.text_input("ISBN",placeholder="Example: 9780064400558",key="book_online_query")
        if st.button("🔎 Find Book & Reading Levels",key="book_search_internet"):
            cleaned=re.sub(r"[^0-9Xx]","",q or "")
            if len(cleaned) not in (10,13):
                st.warning("Enter a valid 10- or 13-digit ISBN.")
            else:
                try:
                    with st.spinner("Book Finder is researching this ISBN..."):
                        found=internet_book_lookup_isbn(cleaned)
                    st.session_state["book_online_results"]=[found] if found else []
                    if not found: st.warning("No matching book was found for that ISBN.")
                except Exception as e:
                    st.error(str(e))

        results=st.session_state.get("book_online_results",[])
        if results:
            labels=[
                f"{b.get('title') or 'Untitled'} — {b.get('author') or 'Unknown'}"
                +(f" | ISBN {b.get('isbn')}" if b.get("isbn") else "")
                for b in results
            ]
            idx=st.selectbox(
                "Select the exact book",list(range(len(results))),
                format_func=lambda i:labels[i],key="book_online_result_select"
            )
            raw_book=results[idx]
            enrich_key=f"{raw_book.get('isbn','')}|{raw_book.get('title','')}|{raw_book.get('author','')}"
            if st.session_state.get("_bookfinder_enriched_key")!=enrich_key:
                with st.spinner("Researching this edition's reading levels..."):
                    book=enrich_book_result(raw_book)
                st.session_state["_bookfinder_enriched_key"]=enrich_key
                st.session_state["_bookfinder_enriched_book"]=book
            else:
                book=st.session_state.get("_bookfinder_enriched_book") or raw_book

            cover=book.get("cover_url") or openlibrary_cover_url(book)
            a,b=st.columns([1,2])
            if cover:
                a.image(cover,width=180)
            b.markdown(f"**{book.get('title') or 'Untitled'}**")
            b.write(f"Author: {book.get('author') or 'Unknown'}")
            if book.get("isbn"):
                b.write(f"ISBN: {book.get('isbn')}")

            sch=book.get("scholastic") or {}
            levels=sch.get("levels") or {}
            lexile=(book.get("lexile_measure") or levels.get("lexile") or "").strip().upper()
            fp_est=book.get("estimated_fp_from_lexile") or {}
            direct_fp=(levels.get("fountas_pinnell") or levels.get("guided_reading") or "").strip().upper()

            st.markdown("#### Reading Level")
            if direct_fp:
                st.success(f"Published Guided Reading/F&P: **{direct_fp}**")
            if lexile:
                st.success(f"Lexile: **{lexile}**")
            if levels.get("grade_level"):
                st.write(f"Grade level: **{levels['grade_level']}**")
            if not direct_fp and fp_est.get("estimate"):
                st.info(f"Estimated F&P from Lexile: **{fp_est['estimate']}**")

            st.markdown("---")
            st.markdown("### 👥 Class Book Fit")
            if roster.empty:
                st.info("Add scholars to this class to compare the book with the roster.")
            else:
                groups=classify_roster_for_book(book,roster)
                col1,col2,col3=st.columns(3)
                specs=[
                    (col1,"🔴 Likely to Struggle","Likely to Struggle"),
                    (col2,"🟡 Okay / With Support","Okay / With Support"),
                    (col3,"🟢 Little to No Problems","Little to No Problems"),
                ]
                for col,title,key in specs:
                    with col:
                        st.markdown(f"#### {title}")
                        if not groups[key]:
                            st.caption("None")
                        for item in groups[key]:
                            st.markdown(f"**{item['name']}**")
                            st.caption(f"F&P {item['fp']}" if item["fp"] else "Reading level not saved")
                if groups["Need More Data"]:
                    st.caption("Need more data: "+", ".join(x["name"] for x in groups["Need More Data"]))

            verified_fp=st.text_input(
                "Verified F&P level (optional)",value=direct_fp,
                placeholder="Example: M",key="book_verified_fp_online"
            ).strip().upper()
            notes=st.text_area("Book notes",height=70,key="book_online_notes")

            if st.button("Save to My Book Catalog",key="save_online_book"):
                c=conn()
                cols=[r["name"] for r in c.execute("PRAGMA table_info(book_catalog)").fetchall()]
                title=(book.get("title") or "Untitled").strip()
                author=(book.get("author") or "").strip()
                isbn=(book.get("isbn") or "").strip()
                existing=None
                if isbn and "isbn" in cols:
                    existing=c.execute("SELECT id FROM book_catalog WHERE isbn=? LIMIT 1",(isbn,)).fetchone()
                if not existing:
                    existing=c.execute(
                        "SELECT id FROM book_catalog WHERE lower(title)=lower(?) AND lower(COALESCE(author,''))=lower(?) LIMIT 1",
                        (title,author)
                    ).fetchone()
                if existing:
                    sets=[]; vals=[]
                    if "fp_level" in cols:
                        sets.append("fp_level=?"); vals.append(verified_fp)
                    if "isbn" in cols:
                        sets.append("isbn=?"); vals.append(isbn)
                    if "notes" in cols:
                        sets.append("notes=?"); vals.append(notes.strip())
                    if sets:
                        vals.append(int(existing["id"]))
                        c.execute(f"UPDATE book_catalog SET {','.join(sets)} WHERE id=?",vals)
                else:
                    ic=["title"]; iv=[title]
                    if "author" in cols: ic.append("author"); iv.append(author)
                    if "fp_level" in cols: ic.append("fp_level"); iv.append(verified_fp)
                    if "isbn" in cols: ic.append("isbn"); iv.append(isbn)
                    if "notes" in cols: ic.append("notes"); iv.append(notes.strip())
                    c.execute(
                        f"INSERT INTO book_catalog({','.join(ic)}) VALUES ({','.join(['?']*len(ic))})",iv
                    )
                c.commit(); c.close()
                st.success("Book saved to My Book Catalog.")

    with tabs[1]:
        st.markdown("### 👩🏽‍🎓 Scholar Fit")
        st.caption("Scholar Fit stays available before your roster is added. Once scholars are entered, their names and saved reading levels populate here automatically.")

        if roster.empty:
            st.info("No scholars are in this class yet. You can still see and use Scholar Fit below.")
            scholar_name=st.text_input("Scholar name",placeholder="Type a name for this check",key="scholar_fit_manual_name").strip() or "Scholar"
            m1,m2=st.columns(2)
            scholar_fp=m1.text_input("Current F&P / Guided Reading",placeholder="Example: L",key="scholar_fit_manual_fp").strip().upper()
            scholar_lx=m2.text_input("Current Lexile",placeholder="Example: 520L",key="scholar_fit_manual_lexile").strip().upper()
            sid=None
        else:
            sid=st.selectbox("Scholar",list(roster.id.astype(int)),format_func=lambda x:nm(roster[roster.id==x].iloc[0]),key="scholar_fit_selected")
            scholar_name=nm(roster[roster.id==sid].iloc[0])
            scholar_fp,scholar_lx=scholar_current_reading_levels(sid)
            m1,m2=st.columns(2)
            m1.metric("Current F&P / Guided Reading",scholar_fp or "Not saved")
            m2.metric("Current Lexile",scholar_lx or "Not saved")

        st.markdown("#### 📚 Recommended From My Book Catalog")
        if sid is None:
            c=conn()
            try: _cat=pd.read_sql_query("SELECT * FROM book_catalog ORDER BY title",c)
            except Exception: _cat=pd.DataFrame()
            c.close(); recs=[]
            if not _cat.empty:
                for _,r in _cat.iterrows():
                    _fp=str(r.get("fp_level") or "").strip().upper()
                    _fit,_use,_why=evaluate_book_fit(book_fp=_fp,book_lexile="",scholar_fp=scholar_fp,scholar_lexile=scholar_lx)
                    if _fit in ("Good Fit","Slightly Challenging"):
                        recs.append({"title":str(r.get("title") or "Untitled"),"author":str(r.get("author") or ""),"fp":_fp,"fit":_fit,"use":_use})
        else:
            recs,_,_=catalog_recommendations_for_scholar(sid)
        if not recs:
            st.caption("No matching saved books yet. You can still check a specific book by ISBN below.")
        else:
            for r in recs:
                icon="✅" if r["fit"]=="Good Fit" else "🟡"
                line=f"{icon} **{r['title']}**"
                if r["author"]: line+=f" — {r['author']}"
                if r["fp"]: line+=f" | F&P {r['fp']}"
                st.markdown(line); st.caption(f"{r['fit']} • {r['use']}")

        st.markdown("---")
        st.markdown("#### 🔎 Check a Specific Book")
        sq=st.text_input("ISBN",placeholder="Example: 9780689818769",key="scholar_fit_book_query")
        if st.button("Check This Book for Scholar",key="scholar_fit_search"):
            cleaned=re.sub(r"[^0-9Xx]","",sq or "")
            if len(cleaned) not in (10,13):
                st.warning("Enter a valid 10- or 13-digit ISBN.")
            else:
                try:
                    with st.spinner("Researching the book and checking scholar fit..."):
                        found=internet_book_lookup_isbn(cleaned)
                    st.session_state["scholar_fit_results"]=[found] if found else []
                    if not found: st.warning("No matching book was found for that ISBN.")
                except Exception as ex: st.error(str(ex))

        sf_results=st.session_state.get("scholar_fit_results",[])
        if sf_results:
            sf_book=sf_results[0]
            sf_fp,sf_lx,sf_direct,sf_est=book_comparison_levels(sf_book)
            fit,use,why=evaluate_book_fit(book_fp=sf_fp,book_lexile=sf_lx,scholar_fp=scholar_fp,scholar_lexile=scholar_lx)
            st.markdown(f"**{sf_book.get('title') or 'Untitled'}** — {sf_book.get('author') or 'Unknown'}")
            if sf_direct: st.write(f"Published Guided Reading/F&P: **{sf_direct}**")
            elif sf_est: st.write(f"Estimated F&P: **{sf_est}**")
            if sf_lx: st.write(f"Lexile: **{sf_lx}**")
            if fit=="Good Fit": st.success(f"### ✅ Good Fit\\n**{use}**")
            elif fit=="Slightly Challenging": st.warning(f"### 🟡 Slightly Challenging\\n**{use}**")
            elif fit=="Too Difficult Right Now": st.error(f"### 🔴 Too Difficult Right Now\\n**{use}**")
            elif fit=="Likely Too Easy": st.info(f"### 🔵 Likely Too Easy\\n**{use}**")
            else: st.info(f"### ⚪ Need More Data\\n**{use}**")
            st.write(why)

    with tabs[2]:
        st.markdown("### My Book Catalog")
        c=conn()
        catalog=pd.read_sql_query("SELECT * FROM book_catalog ORDER BY title",c)
        c.close()

        if catalog.empty:
            st.caption("Your book catalog is empty. Use Search Internet to add books.")
        else:
            search_local=st.text_input(
                "Search saved books",
                placeholder="Type title or author",
                key="book_catalog_search_v32"
            )
            filtered=catalog.copy()
            if search_local.strip():
                mask=filtered["title"].astype(str).str.contains(search_local,case=False,na=False)
                if "author" in filtered.columns:
                    mask=mask | filtered["author"].astype(str).str.contains(search_local,case=False,na=False)
                filtered=filtered[mask]

            if filtered.empty:
                st.warning("No saved books match that search.")
            else:
                bid=st.selectbox(
                    "Saved book",
                    list(filtered.id.astype(int)),
                    format_func=lambda x:(
                        f"{filtered[filtered.id==x].iloc[0].title}"
                        + (
                            f" — {filtered[filtered.id==x].iloc[0].author}"
                            if "author" in filtered.columns and str(filtered[filtered.id==x].iloc[0].author or "").strip()
                            else ""
                        )
                    ),
                    key="saved_book_select_v32"
                )
                row=filtered[filtered.id==bid].iloc[0]
                saved_fp=str(row.fp_level or "").strip().upper() if "fp_level" in filtered.columns else ""

                st.write(f"**Title:** {row.title}")
                if "author" in filtered.columns and str(row.author or "").strip():
                    st.write(f"**Author:** {row.author}")
                if "isbn" in filtered.columns and str(row.isbn or "").strip():
                    st.write(f"**ISBN:** {row.isbn}")
                st.write(f"**Verified F&P:** {saved_fp or 'Not saved'}")

                if saved_fp and current_fp:
                    relation=fp_level_relation(current_fp,saved_fp)
                    if relation:
                        label,diff=relation
                        if diff==0:
                            st.success(f"On {scholar_name}'s current F&P level.")
                        elif diff<0:
                            st.info(f"{label}: book {saved_fp}, scholar {current_fp}.")
                        else:
                            st.warning(f"{label}: book {saved_fp}, scholar {current_fp}.")
                elif not saved_fp:
                    st.info("Add a verified F&P level to compare this book automatically.")

                st.markdown("#### Edit Saved Book")
                with st.form("edit_saved_book_v32"):
                    etitle=st.text_input("Title",value=str(row.title or ""))
                    eauthor=st.text_input(
                        "Author",
                        value=str(row.author or "") if "author" in filtered.columns else ""
                    )
                    eisbn=st.text_input(
                        "ISBN",
                        value=str(row.isbn or "") if "isbn" in filtered.columns else ""
                    )
                    efp=st.text_input("Verified F&P",value=saved_fp)
                    enotes=st.text_area(
                        "Notes",
                        value=str(row.notes or "") if "notes" in filtered.columns else ""
                    )
                    if st.form_submit_button("Save Book Changes"):
                        c=conn()
                        cols=[r["name"] for r in c.execute("PRAGMA table_info(book_catalog)").fetchall()]
                        sets=["title=?"]; vals=[etitle.strip()]
                        if "author" in cols:
                            sets.append("author=?"); vals.append(eauthor.strip())
                        if "isbn" in cols:
                            sets.append("isbn=?"); vals.append(eisbn.strip())
                        if "fp_level" in cols:
                            sets.append("fp_level=?"); vals.append(efp.strip().upper())
                        if "notes" in cols:
                            sets.append("notes=?"); vals.append(enotes.strip())
                        vals.append(int(bid))
                        c.execute(f"UPDATE book_catalog SET {','.join(sets)} WHERE id=?",vals)
                        c.commit(); c.close()
                        st.success("Saved book updated.")
                        st.rerun()

                confirm=st.checkbox("Delete this saved book",key="delete_saved_book_confirm_v32")
                if st.button("Delete Book",disabled=not confirm,key="delete_saved_book_v32"):
                    c=conn()
                    c.execute("DELETE FROM book_catalog WHERE id=?",(int(bid),))
                    c.commit(); c.close()
                    st.success("Book deleted from catalog.")
                    st.rerun()


elif page=="Student Grouping":
    st.markdown('<div class="page-title">Student Grouping</div><div class="page-subtitle">Group scholars by subject performance or skill.</div>',unsafe_allow_html=True)
    st.write("Create Low, Mid, and High groups using either overall subject grades or a specific skill/standard.")

    active_class=selected_class if selected_class else None
    low_default,mid_default=get_grouping_cutoffs(active_class)

    st.markdown("### Group Cutoffs")
    c1,c2=st.columns(2)
    low_max=c1.number_input("Low group: score at or below",min_value=0.0,max_value=100.0,value=float(low_default),step=1.0)
    mid_max=c2.number_input("Mid group: score at or below",min_value=0.0,max_value=100.0,value=float(mid_default),step=1.0)

    st.caption(f"Current rule: Low ≤ {low_max:g}%, Mid = above {low_max:g}% through {mid_max:g}%, High > {mid_max:g}%.")

    if mid_max<=low_max:
        st.error("The Mid cutoff must be higher than the Low cutoff.")
    else:
        if st.button("Save Group Cutoffs"):
            save_grouping_cutoffs(active_class,low_max,mid_max)
            st.success("Grouping cutoffs saved.")
            st.rerun()

        mode=st.radio("Group students by",["Overall Subject Grade","Specific Skill / Standard"],horizontal=True)
        subject=st.selectbox("Subject",["ELA","Math","Science","Social Studies"],key="group_subject")

        if mode=="Overall Subject Grade":
            df=subject_grouping_df(active_class,subject,low_max,mid_max)
            if "Group" not in df.columns:
                df["Group"]="No Data"

            a,b,c=st.columns(3)
            a.metric("Low",int((df["Group"]=="Low").sum()))
            b.metric("Mid",int((df["Group"]=="Mid").sum()))
            c.metric("High",int((df["Group"]=="High").sum()))

            st.markdown(f"### {subject} Groups")
            tabs=st.tabs(["Low","Mid","High","No Data","All"])

            for tab,label in zip(tabs,["Low","Mid","High","No Data",None]):
                with tab:
                    show=df if label is None else df[df["Group"]==label]
                    if show.empty:
                        st.caption("No scholars in this group.")
                    else:
                        st.dataframe(show,hide_index=True,use_container_width=True)

        else:
            sdf=standards_df(subject)
            if sdf.empty:
                st.info("No skills/standards are available for this subject.")
            else:
                skill_option=st.selectbox(
                    "Skill / Standard",
                    [f"{r.code} — {r.skill}" for _,r in sdf.iterrows()]
                )
                skill_code=skill_option.split(" — ")[0]
                skill_name=skill_option.split(" — ",1)[1] if " — " in skill_option else skill_code

                df=skill_grouping_df(active_class,subject,skill_code,low_max,mid_max)
                if "Group" not in df.columns:
                    df["Group"]="No Data"

                a,b,c=st.columns(3)
                a.metric("Low",int((df["Group"]=="Low").sum()))
                b.metric("Mid",int((df["Group"]=="Mid").sum()))
                c.metric("High",int((df["Group"]=="High").sum()))

                st.markdown(f"### {subject}: {skill_name}")
                tabs=st.tabs(["Low","Mid","High","No Data","All"])
                for tab,label in zip(tabs,["Low","Mid","High","No Data",None]):
                    with tab:
                        show=df if label is None else df[df["Group"]==label]
                        if show.empty:
                            st.caption("No scholars in this group.")
                        else:
                            st.dataframe(show,hide_index=True,use_container_width=True)

                st.info("Use these groups for reteach, on-level practice, enrichment, centers, or flexible small groups. Re-group whenever new grades are entered.")

elif page=="Report Card Comments":
    st.markdown('<div class="page-title">Report Card Comments</div><div class="page-subtitle">Generate parent-friendly comments using quarter data.</div>',unsafe_allow_html=True)
    roster=scholars_df(selected_class or None)
    academic_year=current_academic_year()

    if roster.empty:
        st.info("No scholars yet. The Report Card Comment workspace is ready below and will populate with scholar data automatically once you add a roster.")

        r0,r1,r2=st.columns(3)
        r0.text_input("Find scholar",placeholder="Scholar names will appear here",disabled=True,key="rc_preview_find")
        r1.selectbox("Subject",["ELA","Math","Science","Social Studies"],key="rc_preview_subject")
        r2.selectbox("Marking Period",["Quarter 1","Quarter 2","Quarter 3","Quarter 4"],key="rc_preview_mp")

        st.markdown("### Quarter Grade Evidence")
        st.dataframe(
            pd.DataFrame(columns=["Assignment Date","Assignment","Category","Standard","Grade"]),
            hide_index=True,use_container_width=True
        )
        st.caption("Quarter averages, assignment evidence, growth, and missing-grade checks will populate here.")

        st.markdown("### Upcoming Skills")
        preview_subj=st.session_state.get("rc_preview_subject","ELA")
        preview_skills=[r.skill for _,r in standards_df(preview_subj).iterrows()]
        st.multiselect("Upcoming skills to include",preview_skills,key="rc_preview_next_skills")

        st.markdown("### Comment Controls")
        c1,c2=st.columns(2)
        c1.selectbox("Comment Length",["Short","Standard","Detailed"],index=1,key="rc_preview_length")
        c2.number_input("Maximum Characters",min_value=100,max_value=5000,value=1000,step=50,key="rc_preview_chars")

        ck1,ck2=st.columns(2)
        with ck1:
            st.checkbox("Include overall quarter grade",value=True,key="rc_preview_overall")
            st.checkbox("Include quarter-to-quarter growth",value=True,key="rc_preview_growth")
            st.checkbox("Include strongest skill",value=True,key="rc_preview_strength")
            st.checkbox("Include area for improvement",value=True,key="rc_preview_need")
        with ck2:
            st.checkbox("Include assignment examples",value=True,key="rc_preview_assign")
            st.checkbox("Include how family can help",value=True,key="rc_preview_home")
            st.checkbox("Include what teacher will do",value=True,key="rc_preview_teacher")
            st.checkbox("Include upcoming skills",value=True,key="rc_preview_upcoming")

        st.button("Generate Quarter-Based Comment",disabled=True,key="rc_preview_generate")
        st.text_area(
            "Edit / finalize / copy",
            value="Generated scholar-specific comment will appear here.",
            height=220,disabled=True,key="rc_preview_text"
        )
        st.button("Save Final Comment",disabled=True,key="rc_preview_save")

        st.markdown("---")
        st.markdown("## Quarter Closeout")
        q1,q2,q3,q4=st.columns(4)
        q1.checkbox("Grades complete/reviewed",disabled=True,key="rc_preview_close1")
        q2.checkbox("Data reviewed",disabled=True,key="rc_preview_close2")
        q3.checkbox("Comment generated",disabled=True,key="rc_preview_close3")
        q4.checkbox("Comment finalized",disabled=True,key="rc_preview_close4")
        st.caption("Closeout status will become scholar-specific once a roster exists.")
    else:
        q=st.text_input("Find scholar",placeholder="Type first or last name",key="rc_find_scholar")
        filtered=roster
        if q.strip():
            filtered=filtered[(filtered.first_name+" "+filtered.last_name).str.contains(q.strip(),case=False,na=False)]
        if filtered.empty:
            st.warning("No matching scholar.")
        else:
            sid=st.selectbox("Scholar",list(filtered.id.astype(int)),
                             format_func=lambda x:nm(filtered[filtered.id==x].iloc[0]),
                             key="rc_scholar")
            name=nm(filtered[filtered.id==sid].iloc[0])
            subj=st.selectbox("Subject",["ELA","Math","Science","Social Studies"],key="rc_subject")
            mp=st.selectbox("Marking Period",["Quarter 1","Quarter 2","Quarter 3","Quarter 4"],key="rc_marking_period")

            due=report_deadline(academic_year,mp)
            if due:
                st.info(f"📅 {mp} report-card comments are due: {due}")

            qdf=quarter_grade_data(sid,subj,mp)
            missing,total_assignments=quarter_missing_grade_check(sid,subj,mp,selected_class)

            if total_assignments:
                complete_count=total_assignments-len(missing)
                st.write(f"**Grade completion:** {complete_count} of {total_assignments} assignments entered.")
                if missing:
                    st.warning("Missing grades: " + ", ".join(missing[:8]) + ("..." if len(missing)>8 else ""))
            else:
                st.caption("No assignments are currently tagged to this quarter/subject.")

            if not qdf.empty:
                qavg=weighted_average_from_grade_rows(qdf)
                a,b,c=st.columns(3)
                a.metric(f"{mp} Average",f"{qavg:.1f}% ({letter(qavg)})")
                b.metric("Graded Assignments",len(qdf))
                c.metric("Score Range",f"{qdf['pct'].min():.0f}%–{qdf['pct'].max():.0f}%")

                st.markdown("### Quarter Grade Evidence")
                display=qdf.copy()
                display["Grade"]=display["pct"].map(lambda x:f"{x:.1f}%")
                st.dataframe(display[["assignment_date","title","category","standard_code","Grade"]],
                             hide_index=True,use_container_width=True)

                # quarter-to-quarter growth
                quarter_order=["Quarter 1","Quarter 2","Quarter 3","Quarter 4"]
                mp_index=quarter_order.index(mp)
                if mp_index>0:
                    prev=quarter_order[mp_index-1]
                    prevdf=quarter_grade_data(sid,subj,prev)
                    if not prevdf.empty:
                        prevavg=weighted_average_from_grade_rows(prevdf)
                        delta=qavg-prevavg
                        st.write(f"**Quarter-to-quarter change:** {prev} {prevavg:.1f}% → {mp} {qavg:.1f}% ({delta:+.1f} points)")

            skill_options=[r.skill for _,r in standards_df(subj).iterrows()]
            selected_next_skills=st.multiselect("Upcoming skills to include",skill_options,key="rc_next_skills")
            next_skill_text=", ".join(selected_next_skills)

            st.markdown("### Comment Controls")
            c1,c2=st.columns(2)
            comment_length=c1.selectbox("Comment Length",["Short","Standard","Detailed"],index=1,key="rc_length")
            max_chars=c2.number_input("Maximum Characters",min_value=100,max_value=5000,value=1000,step=50,key="rc_max_chars")

            include_overall=st.checkbox("Include overall quarter grade",value=True,key="rc_include_overall")
            include_growth=st.checkbox("Include quarter-to-quarter growth",value=True,key="rc_include_growth")
            include_strength=st.checkbox("Include strongest skill",value=True,key="rc_include_strength")
            include_need=st.checkbox("Include area for improvement",value=True,key="rc_include_need")
            include_assignment=st.checkbox("Include assignment examples",value=True,key="rc_include_assignment")
            include_home=st.checkbox("Include how family can help",value=True,key="rc_include_home")
            include_teacher=st.checkbox("Include what teacher will do",value=True,key="rc_include_teacher")
            include_next=st.checkbox("Include upcoming skills",value=True,key="rc_include_next")

            if st.button("Generate Quarter-Based Comment",key="rc_generate"):
                base=quarter_report_comment(name,subj,mp,qdf,next_skill_text if include_next else "",sid=sid)
                # Simple trimming/removal based on controls
                comment=base
                if not include_assignment:
                    comment=re.sub(r' For example,.*?helpful\.', '', comment)
                if not include_home:
                    comment=re.sub(r' At home,.*?(?= In class,| Our next|$)', '', comment)
                if not include_teacher:
                    comment=re.sub(r' In class,.*?(?= Our next|$)', '', comment)
                if not include_next:
                    comment=re.sub(r' Our next instructional focus.*$', '', comment)
                if not include_growth:
                    comment=comment.replace(" improved over the course of the marking period","")
                    comment=comment.replace(" performed fairly consistently across the marking period","")
                    comment=comment.replace(" showed some decline later in the marking period","")
                if not include_overall:
                    comment=re.sub(r'During .*? graded assignments?\.', f"During {mp}, {name} completed graded {subj.lower()} work.", comment, count=1)
                if comment_length=="Short":
                    sentences=re.split(r'(?<=[.!?])\s+',comment)
                    comment=" ".join(sentences[:4])
                elif comment_length=="Standard":
                    sentences=re.split(r'(?<=[.!?])\s+',comment)
                    comment=" ".join(sentences[:6])
                if len(comment)>max_chars:
                    comment=comment[:max_chars].rsplit(" ",1)[0].rstrip(" ,;:")+"."
                st.session_state["comment"]=comment
                st.session_state["rc_comment_text"]=comment

            if "rc_comment_text" not in st.session_state:
                st.session_state["rc_comment_text"]=st.session_state.get("comment","")
            txt=st.text_area("Edit / finalize / copy",height=300,key="rc_comment_text")
            st.caption(f"Character count: {len(txt)} / {int(max_chars)}")

            if st.button("Save Final Comment",key="rc_save_final"):
                c=conn()
                c.execute("""INSERT INTO report_comments(scholar_id,created_at,subject,marking_period,comment_text)
                             VALUES (?,?,?,?,?)""",
                          (sid,datetime.now().isoformat(timespec="minutes"),subj,mp,txt))
                c.commit(); c.close()
                row=closeout_row(sid,mp,academic_year)
                save_closeout(sid,mp,academic_year,
                              bool(row["grades_reviewed"]) if row else False,
                              bool(row["data_reviewed"]) if row else False,
                              True,True)
                st.success("Saved to the scholar profile and marked finalized.")
                st.rerun()

            st.markdown("---")
            st.markdown("## Quarter Closeout")
            close=closeout_row(sid,mp,academic_year)
            g=bool(close["grades_reviewed"]) if close else False
            d=bool(close["data_reviewed"]) if close else False
            cg=bool(close["comment_generated"]) if close else False
            cf=bool(close["comment_finalized"]) if close else False

            q1,q2,q3,q4=st.columns(4)
            g_new=q1.checkbox("Grades complete/reviewed",value=g,key="close_grades")
            d_new=q2.checkbox("Data reviewed",value=d,key="close_data")
            cg_new=q3.checkbox("Comment generated",value=cg or bool(txt),key="close_generated")
            cf_new=q4.checkbox("Comment finalized",value=cf,key="close_finalized")
            if st.button("Save Closeout Status",key="save_closeout_status"):
                save_closeout(sid,mp,academic_year,g_new,d_new,cg_new,cf_new)
                st.success("Closeout status saved.")
                st.rerun()

            c=conn()
            saved_comments=pd.read_sql_query("SELECT * FROM report_comments WHERE scholar_id=? ORDER BY created_at DESC",c,params=[sid])
            c.close()
            if not saved_comments.empty:
                st.markdown("### Saved Report Card Comments")
                st.dataframe(saved_comments[["id","created_at","subject","marking_period","comment_text"]],
                             hide_index=True,use_container_width=True)

elif page=="Little Assistant":
    st.markdown('<div class="page-title">Little Assistant</div><div class="page-subtitle">Parent messages, call scripts, IAT referrals, and student-support tools.</div>',unsafe_allow_html=True)
    st.caption("Support tools for teacher paperwork and family communication.")
    tool=st.radio(
        "Choose a tool",
        ["IEP / Student Support","IAT Referral","Parent Message","Phone Call Script"],
        horizontal=True,
        key="assistant_tool"
    )

    roster=scholars_df(selected_class or None)
    if roster.empty:
        st.info("No scholars yet. You can still see the complete Little Assistant workspace; scholar information will prefill these tools automatically after a roster is added.")

        if tool=="IEP / Student Support":
            st.markdown("### Profile & Data Prefill")
            st.text_area("Existing scholar data",value="Scholar profile, assessment data, grades, work samples, and saved support notes will prefill here.",height=150,disabled=True,key="la_preview_prefill")
            p1,p2=st.columns(2)
            p1.selectbox("Documentation type",["Academic","Reading","Writing","Math","Attention / task completion","Behavior / self-management","Communication","Social interaction","Intervention response","Other"],key="la_preview_note_type")
            p2.selectbox("Area of concern",["Academic performance","Reading fluency","Reading comprehension","Written expression","Math computation","Attention / focus","Task completion","Behavior / self-management","Other / Custom"],key="la_preview_concern")
            st.text_area("Objective observation",placeholder="Describe only what you observed.",key="la_preview_observation")
            p3,p4=st.columns(2)
            p3.selectbox("Frequency / pattern",["Rarely","1–2 times per week","3–4 times per week","Daily","Multiple times per day","Across settings"],key="la_preview_frequency")
            p4.selectbox("Primary support/intervention tried",["Small-group reteach","1:1 teacher conference","Repeated directions","Visual reminder / checklist","Extended time","Frequent check-ins","Other / Custom"],key="la_preview_intervention")
            st.text_area("Additional intervention details",key="la_preview_intervention_details")
            p5,p6=st.columns(2)
            p5.selectbox("Response to support",["Improved with support","Some improvement","Improvement was temporary","No noticeable improvement yet","Performance remained inconsistent"],key="la_preview_response")
            p6.selectbox("Educational impact",["Reduces accuracy","Slows work completion","Makes independent work difficult","Affects participation","Affects comprehension","Affects written output"],key="la_preview_impact")
            st.button("Save Evidence Note",disabled=True,key="la_preview_save_evidence")
            st.button("Generate Teacher Support Summary",disabled=True,key="la_preview_support_summary")
            st.text_area("Editable teacher support summary",height=180,disabled=True,key="la_preview_support_text")

        elif tool=="IAT Referral":
            st.markdown("### IAT Referral Writer")
            st.caption("Choose from guided options or use Other / Custom when you need your own wording.")
            p1,p2=st.columns(2)
            p1.selectbox("Primary reason for referral",[
                "Academic progress below expectations","Reading difficulty","Writing difficulty","Math difficulty",
                "Attention / focus concerns","Task completion / work habits","Behavior / self-management",
                "Social / peer interaction","Communication concerns","Attendance / missed instruction",
                "Multiple areas of concern","Other / Custom"
            ],key="la_preview_iat_reason")
            p2.text_input("Other / custom reason",placeholder="Type your own reason",key="la_preview_iat_reason_other")
            st.multiselect("Areas of concern",[
                "Reading fluency","Reading comprehension","Phonics / decoding","Vocabulary","Written expression",
                "Spelling / conventions","Math computation","Math problem solving","Following directions",
                "Attention / focus","Task initiation","Task completion","Organization","Behavior regulation",
                "Peer interaction","Participation","Independence","Communication","Other / Custom"
            ],key="la_preview_iat_areas")
            st.text_input("Other / custom area of concern",key="la_preview_iat_area_other")
            p3,p4=st.columns(2)
            p3.selectbox("How long has this been observed?",[
                "Less than 1 month","1–2 months","3–4 months","Most of the school year",
                "Since the beginning of the school year","Other / Custom"
            ],key="la_preview_iat_duration")
            p4.text_input("Other / custom duration",key="la_preview_iat_duration_other")
            st.multiselect("Classroom supports already tried",[
                "Small-group reteach","1:1 teacher conference","Preferential seating","Directions broken into steps",
                "Visual checklist / reminder","Frequent check-ins","Extended time","Reduced task length",
                "Graphic organizer","Model/example provided","Repeated practice","Read directions aloud",
                "Peer support","Positive reinforcement","Behavior reminder / redirection","Home practice requested",
                "Family contacted","Progress monitoring","Other / Custom"
            ],key="la_preview_iat_supports")
            st.text_area("Other / custom support or additional details",height=85,key="la_preview_iat_support_other")
            p5,p6=st.columns(2)
            p5.selectbox("Response to support",[
                "Improved with support","Some improvement, but concern remains","Improvement was temporary",
                "Performance remains inconsistent","Minimal or no improvement","More data is needed","Other / Custom"
            ],key="la_preview_iat_response")
            p6.text_input("Other / custom response",key="la_preview_iat_response_other")
            st.text_area("Objective description / examples",height=100,key="la_preview_iat_objective")
            st.button("Generate IAT Referral",disabled=True,key="la_preview_iat_generate")
            st.text_area("Editable IAT referral",height=220,disabled=True,key="la_preview_iat_text")

        elif tool=="Parent Message":
            st.markdown("### Parent Message")
            st.caption("Use the dropdowns for quick setup, or choose Other / Custom and type exactly what you need.")
            p1,p2=st.columns(2)
            p1.selectbox("Message type",[
                "Positive update","Progress update","Growth update","Academic concern","Behavior concern",
                "Missing work","Homework reminder","Assessment / test","Conference request",
                "Injury / classroom incident","General reminder","Other / Custom"
            ],key="la_preview_parent_type")
            p2.selectbox("Subject",["General","ELA","Math","Science","Social Studies","Behavior","Attendance","Other / Custom"],key="la_preview_parent_subject")
            c1,c2=st.columns(2)
            c1.text_input("Other / custom message type",key="la_preview_parent_type_other")
            c2.text_input("Other / custom subject",key="la_preview_parent_subject_other")
            st.selectbox("Tone",["Warm & supportive","Positive & celebratory","Direct & professional","Concerned but supportive","Brief reminder","Other / Custom"],key="la_preview_parent_tone")
            st.text_input("Other / custom tone",key="la_preview_parent_tone_other")
            st.text_area("What happened / what should the family know?",height=110,key="la_preview_parent_details")
            st.selectbox("Requested next step",[
                "No action needed — informational only","Please discuss with scholar","Please complete/return work",
                "Please practice at home","Please reply to this message","Please schedule a conference",
                "I will follow up again","Other / Custom"
            ],key="la_preview_parent_next_choice")
            st.text_area("Other / custom next step or additional request",height=80,key="la_preview_parent_next")
            st.button("Generate Parent Message",disabled=True,key="la_preview_parent_generate")
            st.text_area("Editable parent message",height=180,disabled=True,key="la_preview_parent_text")

        elif tool=="Phone Call Script":
            st.markdown("### Phone Call Script")
            st.caption("Choose the reason, subject, tone, and goal of the call. Other / Custom lets you enter your own.")
            p1,p2=st.columns(2)
            p1.selectbox("Reason for call",[
                "Positive update","Progress update","Academic concern","Behavior concern","Missing work",
                "Attendance / lateness","Assessment / test","Injury / classroom incident",
                "Conference request","Follow-up from previous contact","Other / Custom"
            ],key="la_preview_call_reason")
            p2.selectbox("Subject",["General","ELA","Math","Science","Social Studies","Behavior","Attendance","Other / Custom"],key="la_preview_call_subject")
            c1,c2=st.columns(2)
            c1.text_input("Other / custom reason",key="la_preview_call_reason_other")
            c2.text_input("Other / custom subject",key="la_preview_call_subject_other")
            st.selectbox("Tone",["Warm & conversational","Positive & celebratory","Direct & professional","Concerned but supportive","Other / Custom"],key="la_preview_call_tone")
            st.text_input("Other / custom tone",key="la_preview_call_tone_other")
            st.text_area("Key details to discuss",height=110,key="la_preview_call_details")
            st.selectbox("Desired outcome / next step",[
                "Inform family only","Ask family to reinforce at home","Create a school-home plan",
                "Schedule a conference","Ask family for input","Follow up again later","Other / Custom"
            ],key="la_preview_call_next_choice")
            st.text_area("Other / custom next step or additional details",height=90,key="la_preview_call_next")
            st.button("Generate Phone Call Script",disabled=True,key="la_preview_call_generate")
            st.text_area("Editable phone call script",height=180,disabled=True,key="la_preview_call_text")
    else:
        if tool=="IEP / Student Support":
            sid,scholar,name,pro,subj_pr,poss_pr=assistant_scholar_context(roster,"assistant_iep")
            st.markdown("### Profile & Data Prefill")
            prefill=iep_prefill_summary(sid)
            st.text_area("Existing scholar data",value=prefill,height=260,disabled=True)

            concern_choices=[
                "Academic performance","Reading fluency","Reading comprehension","Written expression",
                "Spelling / conventions","Math computation","Math problem solving","Attention / focus",
                "Task completion","Following directions","Organization / materials","Behavior / self-management",
                "Peer interaction","Communication","Participation","Independence","Other / Custom"
            ]
            frequency_choices=[
                "Rarely","1–2 times per week","3–4 times per week","Daily","Multiple times per day",
                "Only during independent work","Only during whole group","Only during small group",
                "Across settings","Other / Custom"
            ]
            intervention_choices=[
                "Small-group reteach","1:1 teacher conference","Repeated directions","Directions broken into steps",
                "Visual reminder / checklist","Graphic organizer","Preferential seating","Extended time",
                "Reduced task length","Frequent check-ins","Model/example provided","Read directions aloud",
                "Peer support","Positive reinforcement","Behavior reminder / redirection",
                "Practice with corrected work","Home practice requested","Other / Custom"
            ]
            response_choices=[
                "Improved with support","Some improvement","Improvement was temporary","No noticeable improvement yet",
                "Needed repeated prompting","Completed with adult support","Completed independently after support",
                "Performance remained inconsistent","Other / Custom"
            ]
            impact_choices=[
                "Reduces accuracy","Slows work completion","Makes independent work difficult","Affects participation",
                "Affects comprehension","Affects written output","Affects retention / recall",
                "Affects ability to follow multi-step directions","Affects peer/classroom interactions",
                "Affects consistent academic progress","Other / Custom"
            ]

            with st.form("assistant_iep_form",clear_on_submit=True):
                note_type=st.selectbox("Documentation type",[
                    "Academic","Reading","Writing","Math","Attention / task completion",
                    "Behavior / self-management","Communication","Social interaction",
                    "Intervention response","Other"
                ])
                concern=st.selectbox("Area of concern",concern_choices)
                custom_concern=st.text_input("Custom concern (if needed)")
                area=custom_concern.strip() if concern=="Other / Custom" and custom_concern.strip() else concern
                observation=st.text_area("Objective observation")
                freq_choice=st.selectbox("Frequency / pattern",frequency_choices)
                custom_freq=st.text_input("Custom frequency/pattern")
                frequency=custom_freq.strip() if freq_choice=="Other / Custom" and custom_freq.strip() else freq_choice
                intervention_choice=st.selectbox("Primary support/intervention tried",intervention_choices)
                custom_intervention=st.text_area("Additional intervention details")
                intervention=(intervention_choice if intervention_choice!="Other / Custom" else "")
                if custom_intervention.strip():
                    intervention=(intervention+(": " if intervention else "")+custom_intervention.strip())
                response_choice=st.selectbox("Response to support",response_choices)
                response_details=st.text_area("Additional response details")
                response=(response_choice if response_choice!="Other / Custom" else "")
                if response_details.strip():
                    response=(response+(": " if response else "")+response_details.strip())
                impact_choice=st.selectbox("Educational impact",impact_choices)
                impact_details=st.text_area("Additional impact details")
                impact=(impact_choice if impact_choice!="Other / Custom" else "")
                if impact_details.strip():
                    impact=(impact+(": " if impact else "")+impact_details.strip())

                if st.form_submit_button("Save Evidence Note"):
                    c=conn()
                    c.execute("""INSERT INTO support_notes(
                        scholar_id,created_at,note_type,area,observation,frequency,intervention,
                        response_to_intervention,impact,concern_category,frequency_choice,
                        intervention_choice,response_choice,impact_choice)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (sid,datetime.now().isoformat(timespec="minutes"),note_type,area,observation,
                         frequency,intervention,response,impact,concern,freq_choice,
                         intervention_choice,response_choice,impact_choice))
                    c.commit(); c.close()
                    st.success("Evidence note saved.")
                    st.rerun()

            c=conn()
            notes=pd.read_sql_query("SELECT * FROM support_notes WHERE scholar_id=? ORDER BY created_at DESC",c,params=[sid])
            ws=pd.read_sql_query("SELECT * FROM work_samples WHERE scholar_id=? ORDER BY uploaded_at DESC",c,params=[sid])
            c.close()
            if st.button("Generate Teacher Support Summary",key="assistant_generate_support"):
                parts=[prefill,"\nTeacher Observations / Interventions:"]
                for _,r in notes.head(8).iterrows():
                    parts.append(f"- {r.area}: {r.observation} Frequency: {r.frequency}. Support: {r.intervention}. Response: {r.response_to_intervention}. Impact: {r.impact}.")
                if not ws.empty:
                    parts.append("\nWork-Sample Evidence:")
                    for _,r in ws.head(5).iterrows():
                        parts.append(f"- {r.title or r.file_name}: strengths {r.strengths or '—'}; needs {r.needs or '—'}; next steps {r.next_steps or '—'}.")
                parts.append("\nThis summary documents classroom evidence and teacher observations; it is not a diagnosis.")
                st.session_state["assistant_support_summary"]="\n".join(parts)
            st.text_area("Editable teacher support summary",value=st.session_state.get("assistant_support_summary",""),height=380)

        elif tool=="IAT Referral":
            sid,scholar,name,pro,subj_pr,poss_pr=assistant_scholar_context(roster,"assistant_iat")
            st.markdown("### IAT Referral Writer")
            st.caption(
                "Build a classroom-based referral using the scholar's existing academic data, "
                "intervention evidence, work samples, and family-contact history."
            )

            # Existing scholar data
            prefill=iep_prefill_summary(sid)
            strengths,needs,teacher_actions=academic_summary_for_scholar(sid)
            br=benchmark_for_scholar(sid)

            c=conn()
            support_history=pd.read_sql_query(
                """SELECT created_at,note_type,area,observation,frequency,intervention,
                          response_to_intervention,impact
                   FROM support_notes
                   WHERE scholar_id=?
                   ORDER BY created_at DESC""",
                c,params=[sid]
            )
            work_history=pd.read_sql_query(
                """SELECT uploaded_at,title,file_name,strengths,needs,next_steps
                   FROM work_samples
                   WHERE scholar_id=?
                   ORDER BY uploaded_at DESC""",
                c,params=[sid]
            )
            contact_history=pd.read_sql_query(
                """SELECT communications.created_at,communications.communication_type,
                          communications.reason,communications.details,
                          COALESCE(TRIM(guardians.first_name||' '||guardians.last_name),'') guardian
                   FROM communications
                   LEFT JOIN guardians ON guardians.id=communications.guardian_id
                   WHERE communications.scholar_id=?
                   ORDER BY communications.id DESC""",
                c,params=[sid]
            )
            c.close()

            # Snapshot shown to teacher before writing referral
            with st.expander("Scholar Data Being Used",expanded=False):
                st.text_area("Profile / academic prefill",value=prefill,height=230,disabled=True)

                if br:
                    b1,b2,b3=st.columns(3)
                    b1.metric("F&P",br["fp_spring_level"] or br["fp_fall_level"] or "—")
                    b2.metric("NWEA Reading",br["nwea_spring_reading"] or br["nwea_winter_reading"] or br["nwea_fall_reading"] or "—")
                    b3.metric("NWEA Math",br["nwea_spring_math"] or br["nwea_winter_math"] or br["nwea_fall_math"] or "—")

                if strengths:
                    st.markdown("**Current strengths**")
                    for subj,skill,pct in strengths[:5]:
                        st.write(f"- {subj}: {skill} — {pct:.0f}%")

                if needs:
                    st.markdown("**Current areas of need**")
                    for subj,skill,pct in needs[:5]:
                        st.write(f"- {subj}: {skill} — {pct:.0f}%")

            referral_reason_choices=[
                "Academic progress below expectations",
                "Reading difficulty",
                "Writing difficulty",
                "Math difficulty",
                "Attention / focus concerns",
                "Task completion / work habits",
                "Behavior / self-management",
                "Social / peer interaction",
                "Communication concerns",
                "Attendance / missed instruction",
                "Multiple areas of concern",
                "Other / Custom"
            ]
            concern_area_choices=[
                "Reading fluency","Reading comprehension","Phonics / decoding","Vocabulary",
                "Written expression","Spelling / conventions","Math computation","Math problem solving",
                "Following directions","Attention / focus","Task initiation","Task completion",
                "Organization","Behavior regulation","Peer interaction","Participation",
                "Independence","Communication","Other / Custom"
            ]
            requested_support_choices=[
                "Problem-solving team review",
                "Additional academic intervention",
                "Reading intervention",
                "Math intervention",
                "Behavior / self-management support",
                "Observation by support staff",
                "Progress-monitoring plan",
                "Additional classroom strategies",
                "Family-school support plan",
                "Consideration for further evaluation",
                "Other / Custom"
            ]

            st.markdown("#### Referral Information")
            c1,c2=st.columns(2)
            reason_choice=c1.selectbox("Primary reason for referral",referral_reason_choices,key="iat_reason")
            reason_custom=c2.text_input("Custom reason (if needed)",key="iat_reason_custom")
            referral_reason=reason_custom.strip() if reason_choice=="Other / Custom" and reason_custom.strip() else reason_choice

            selected_concerns=st.multiselect(
                "Areas of concern",
                concern_area_choices,
                key="iat_concerns"
            )
            custom_concern=st.text_input("Additional/custom concern",key="iat_custom_concern")
            concerns=[x for x in selected_concerns if x!="Other / Custom"]
            if custom_concern.strip():
                concerns.append(custom_concern.strip())

            a,b=st.columns(2)
            onset_choice=a.selectbox(
                "How long has the concern been observed?",
                ["Less than 1 month","1–2 months","3–4 months","Most of the school year","Since the beginning of the school year","Other / Custom"],
                key="iat_duration"
            )
            onset_custom=a.text_input("Other / custom duration",key="iat_duration_custom")
            onset=onset_custom.strip() if onset_choice=="Other / Custom" and onset_custom.strip() else onset_choice
            setting=b.multiselect(
                "Where is the concern most noticeable?",
                ["Whole group","Small group","Independent work","Assessments","Transitions","Unstructured time","Across settings","Other / Custom"],
                key="iat_settings"
            )
            setting_custom=b.text_input("Other / custom setting",key="iat_setting_custom")
            setting=[x for x in setting if x!="Other / Custom"] + ([setting_custom.strip()] if setting_custom.strip() else [])

            objective_description=st.text_area(
                "Objective description of the concern",
                placeholder="Describe what you are seeing using observable, specific language. Include examples when possible.",
                height=125,
                key="iat_objective"
            )

            impact=st.text_area(
                "How is this affecting the scholar's progress or access to instruction?",
                placeholder="Example: The scholar needs repeated prompting to begin independent work and often completes less than half of the assigned task.",
                height=105,
                key="iat_impact"
            )

            st.markdown("#### Supports Already Tried")
            selected_interventions=st.multiselect(
                "Classroom interventions / supports",
                [
                    "Small-group reteach","1:1 teacher conference","Preferential seating",
                    "Directions broken into steps","Visual checklist / reminder","Frequent check-ins",
                    "Extended time","Reduced task length","Graphic organizer","Model/example provided",
                    "Repeated practice","Read directions aloud","Peer support","Positive reinforcement",
                    "Behavior reminder / redirection","Home practice requested","Family contacted",
                    "Progress monitoring","Other / Custom"
                ],
                key="iat_interventions"
            )
            custom_intervention=st.text_input("Other / custom intervention",key="iat_intervention_custom")
            selected_interventions=[x for x in selected_interventions if x!="Other / Custom"] + ([custom_intervention.strip()] if custom_intervention.strip() else [])
            intervention_details=st.text_area(
                "Intervention details",
                placeholder="Include frequency, duration, and what the support looked like.",
                height=105,
                key="iat_intervention_details"
            )
            response_to_support=st.selectbox(
                "Overall response to interventions",
                [
                    "Improved with support",
                    "Some improvement, but concern remains",
                    "Improvement was temporary",
                    "Performance remains inconsistent",
                    "Minimal or no improvement",
                    "More data is needed",
                    "Other / Custom"
                ],
                key="iat_response"
            )
            response_custom=st.text_input("Other / custom response",key="iat_response_custom")
            response_to_support=response_custom.strip() if response_to_support=="Other / Custom" and response_custom.strip() else response_to_support

            st.markdown("#### Strengths & Family Communication")
            strengths_note=st.text_area(
                "Scholar strengths to highlight",
                value="; ".join([f"{subj}: {skill}" for subj,skill,pct in strengths[:4]]) if strengths else "",
                height=90,
                key="iat_strengths"
            )
            family_note=st.text_area(
                "Family communication / input",
                placeholder="Summarize relevant parent/guardian communication, concerns, or strategies discussed.",
                height=90,
                key="iat_family"
            )

            request_choice=st.multiselect(
                "What support are you requesting from IAT?",
                requested_support_choices,
                key="iat_request"
            )
            custom_request=st.text_area("Additional request / question for the team",height=85,key="iat_custom_request")

            if st.button("Generate IAT Referral",key="generate_iat_referral"):
                concern_text=", ".join(concerns) if concerns else referral_reason
                setting_text=", ".join(setting) if setting else "across classroom instruction"
                intervention_text=", ".join(selected_interventions) if selected_interventions else "classroom supports documented by the teacher"
                request_text=", ".join([x for x in request_choice if x!="Other / Custom"])
                if custom_request.strip():
                    request_text=(request_text + ("; " if request_text else "") + custom_request.strip())
                if not request_text:
                    request_text="team guidance regarding appropriate next steps and additional supports"

                # Pull a concise current grade snapshot
                grade_parts=[]
                for subj in ["ELA","Math","Science","Social Studies"]:
                    av,_=summary(sid,subj)
                    if av is not None:
                        grade_parts.append(f"{subj}: {av:.1f}% ({letter(av)})")
                grades_text=", ".join(grade_parts) if grade_parts else "limited current grade data"

                benchmark_parts=[]
                if br:
                    fp=br["fp_spring_level"] or br["fp_fall_level"]
                    nr=br["nwea_spring_reading"] or br["nwea_winter_reading"] or br["nwea_fall_reading"]
                    nmth=br["nwea_spring_math"] or br["nwea_winter_math"] or br["nwea_fall_math"]
                    if fp: benchmark_parts.append(f"F&P level {fp}")
                    if nr: benchmark_parts.append(f"NWEA Reading RIT {nr}")
                    if nmth: benchmark_parts.append(f"NWEA Math RIT {nmth}")
                benchmark_text=", ".join(benchmark_parts) if benchmark_parts else "no benchmark scores currently entered"

                evidence_lines=[]
                for _,r in support_history.head(4).iterrows():
                    evidence_lines.append(
                        f"{r.area}: {r.observation or 'documented concern'}; "
                        f"support: {r.intervention or '—'}; response: {r.response_to_intervention or '—'}"
                    )
                evidence_text=" | ".join(evidence_lines) if evidence_lines else "No prior intervention notes are currently saved."

                referral=(
                    f"IAT REFERRAL – {name}\n\n"
                    f"Reason for Referral:\n"
                    f"{referral_reason}. The primary areas of concern are {concern_text}. "
                    f"The concern has been observed for {onset.lower()} and is most noticeable during {setting_text}.\n\n"
                    f"Scholar Strengths:\n"
                    f"{strengths_note.strip() or f"{name} demonstrates strengths that should continue to be built upon during intervention planning."}\n\n"
                    f"Current Academic Data:\n"
                    f"Current classroom grades: {grades_text}. Current benchmark information: {benchmark_text}.\n\n"
                    f"Description of Concern:\n"
                    f"{objective_description.strip() or 'Teacher observation details should be added here.'}\n\n"
                    f"Educational Impact:\n"
                    f"{impact.strip() or f"The concern is affecting {poss_pr} consistent progress and/or independent access to classroom instruction."}\n\n"
                    f"Interventions / Supports Attempted:\n"
                    f"{intervention_text}. {intervention_details.strip()} "
                    f"Overall response: {response_to_support}.\n\n"
                    f"Existing Classroom Evidence:\n"
                    f"{evidence_text}\n\n"
                    f"Family Communication / Input:\n"
                    f"{family_note.strip() or 'Family communication details should be added if applicable.'}\n\n"
                    f"Request to IAT:\n"
                    f"I am requesting {request_text}. I would like the team to review the available data, "
                    f"consider additional supports, and help determine appropriate next steps for {name}."
                )
                st.session_state["iat_referral_text"]=referral
                editor_key=f"iat_referral_editor_{sid}"
                st.session_state[editor_key]=referral
                st.success("IAT referral generated below. You can edit it before saving.")

            editor_key=f"iat_referral_editor_{sid}"
            if editor_key not in st.session_state:
                st.session_state[editor_key]=st.session_state.get("iat_referral_text","")
            referral_text=st.text_area(
                "Editable IAT Referral",
                height=560,
                key=editor_key
            )

            action1,action2=st.columns(2)
            if action1.button("Save IAT Referral to Scholar Record",key="save_iat_referral"):
                if not referral_text.strip():
                    st.warning("Generate or enter the referral first.")
                else:
                    c=conn()
                    c.execute(
                        """INSERT INTO support_notes(
                            scholar_id,created_at,note_type,area,observation,frequency,
                            intervention,response_to_intervention,impact,concern_category,
                            frequency_choice,intervention_choice,response_choice,impact_choice)
                            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (
                            sid,
                            datetime.now().isoformat(timespec="minutes"),
                            "IAT Referral",
                            referral_reason,
                            referral_text.strip(),
                            onset,
                            ", ".join(selected_interventions),
                            response_to_support,
                            impact.strip(),
                            referral_reason,
                            onset,
                            ", ".join(selected_interventions),
                            response_to_support,
                            ", ".join(request_choice)
                        )
                    )
                    c.commit(); c.close()
                    st.success("IAT referral saved to the scholar's support history.")

            action2.download_button(
                "Download IAT Referral",
                data=referral_text or "No IAT referral generated yet.",
                file_name=f"{name.replace(' ','_')}_IAT_Referral.txt",
                mime="text/plain",
                key="download_iat_referral"
            )

            if not support_history.empty:
                with st.expander("Prior Intervention / Support Notes"):
                    st.dataframe(
                        support_history[
                            ["created_at","note_type","area","observation","frequency",
                             "intervention","response_to_intervention","impact"]
                        ],
                        hide_index=True,
                        use_container_width=True
                    )

            if not contact_history.empty:
                with st.expander("Recent Family Contact History"):
                    st.dataframe(contact_history.head(8),hide_index=True,use_container_width=True)

            if not work_history.empty:
                with st.expander("Recent Work-Sample Evidence"):
                    st.dataframe(work_history.head(8),hide_index=True,use_container_width=True)

        elif tool=="Parent Message":
            sid,scholar,name,pro,subj_pr,poss_pr=assistant_scholar_context(roster,"assistant_parent")
            gmap=guardian_display_map(sid)
            pref=update_preference(sid)
            default_gid=0
            if pref and pref["preferred_guardian_id"] and int(pref["preferred_guardian_id"]) in gmap:
                default_gid=int(pref["preferred_guardian_id"])

            st.markdown("### Parent Update Preference")
            requested=st.checkbox("Parent requested regular updates",
                                  value=bool(pref["requested_updates"]) if pref else False,
                                  key="assistant_requested_updates")
            frequency_options=["","Weekly","Every 2 Weeks","Monthly","End of Quarter","As Needed","Other / Custom"]
            current_freq=pref["update_frequency"] if pref and pref["update_frequency"] in frequency_options else ""
            frequency_choice=st.selectbox("Update frequency",frequency_options,index=frequency_options.index(current_freq),key="assistant_update_frequency")
            frequency_custom=st.text_input("Other / custom update frequency",key="assistant_update_frequency_other")
            frequency=frequency_custom.strip() if frequency_choice=="Other / Custom" and frequency_custom.strip() else frequency_choice
            gids=list(gmap)
            gid=st.selectbox("Parent / Guardian",gids,format_func=lambda x:gmap[x],
                             index=gids.index(default_gid) if default_gid in gids else 0,
                             key="assistant_msg_guardian")
            pref_notes=st.text_area("Update preference notes",
                                    value=pref["notes"] if pref else "",
                                    key="assistant_update_pref_notes")
            if st.button("Save Parent Update Preference",key="save_update_preference"):
                save_update_preference(sid,requested,frequency,gid,pref_notes)
                st.success("Parent update preference saved.")
                st.rerun()

            c1,c2=st.columns(2)
            subj_choice=c1.selectbox("Subject",["General","ELA","Math","Science","Social Studies","Behavior","Attendance","Other / Custom"],key="assistant_msg_subject")
            subject_custom=c1.text_input("Other / custom subject",key="assistant_msg_subject_other")
            subj=subject_custom.strip() if subj_choice=="Other / Custom" and subject_custom.strip() else subj_choice

            reason_choice=c2.selectbox("Message Type",[
                "Progress update","Growth update","At-risk / concern update","Positive update",
                "Homework reminder","Missing assignment","Behavior concern","Academic concern",
                "Assessment / test","Attendance / lateness","Conference request",
                "Injury / classroom incident","General update","Other / Custom"
            ],key="assistant_msg_reason")
            reason_custom=c2.text_input("Other / custom message type",key="assistant_msg_reason_other")
            reason=reason_custom.strip() if reason_choice=="Other / Custom" and reason_custom.strip() else reason_choice

            t1,t2=st.columns(2)
            tone_choice=t1.selectbox("Tone",[
                "Warm & supportive","Positive & celebratory","Direct & professional",
                "Concerned but supportive","Brief reminder","Other / Custom"
            ],key="assistant_msg_tone")
            tone_custom=t1.text_input("Other / custom tone",key="assistant_msg_tone_other")
            tone=tone_custom.strip() if tone_choice=="Other / Custom" and tone_custom.strip() else tone_choice

            next_choice=t2.selectbox("Requested next step",[
                "No action needed — informational only","Please discuss with scholar",
                "Please complete/return work","Please practice at home","Please reply to this message",
                "Please schedule a conference","I will follow up again","Other / Custom"
            ],key="assistant_msg_next_choice")
            next_custom=t2.text_input("Other / custom next step",key="assistant_msg_next_other")
            next_step=next_custom.strip() if next_choice=="Other / Custom" and next_custom.strip() else next_choice

            details=st.text_area("Specific information to add",key="assistant_msg_details")

            if st.button("Generate Data-Informed Parent Update",key="assistant_generate_msg"):
                parent_name=gmap.get(gid,"").split(" (")[0] if gid else ""
                greeting=f"Hello {parent_name}," if parent_name else "Hello,"
                if reason_choice in ["Progress update","Growth update","At-risk / concern update"]:
                    body=generate_parent_progress_update(sid,subj)
                    if details.strip():
                        body += " " + details.strip()
                elif reason_choice=="Positive update":
                    body=f"I wanted to share a positive update about {name}. {details}".strip()
                elif reason_choice=="Homework reminder":
                    body=f"This is a reminder regarding {name}'s {subj} homework. {details}".strip()
                elif reason_choice=="Behavior concern":
                    body=f"I am reaching out regarding {name}'s behavior. {details}".strip()
                elif reason_choice=="Injury / classroom incident":
                    body=f"I am contacting you to make you aware of an incident involving {name}. {details}".strip()
                else:
                    body=f"I am reaching out regarding {reason.lower()} for {name}. {details}".strip()

                tone_intro=""
                if tone=="Positive & celebratory":
                    tone_intro=" I am happy to share this update."
                elif tone=="Concerned but supportive":
                    tone_intro=" I want us to work together to support continued progress."
                elif tone=="Brief reminder":
                    tone_intro=" This is a quick reminder."
                elif tone and tone not in ["Warm & supportive","Direct & professional"]:
                    tone_intro=f" {tone}."

                step_text=""
                if next_step and next_step!="No action needed — informational only":
                    step_text=f" Next step: {next_step}."
                elif next_step=="No action needed — informational only":
                    step_text=" No action is needed at this time; I just wanted to keep you informed."

                msg=f"{greeting}{tone_intro} {body}{step_text} Thank you for your support."
                st.session_state["assistant_parent_msg"]=re.sub(r"\\s+"," ",msg).strip()

            msg=st.text_area("Edit / copy message",value=st.session_state.get("assistant_parent_msg",""),height=260)
            if st.button("Save Message to Communication Log",key="assistant_log_msg"):
                log_communication(sid,gid,"Text / School Message",subj,reason,details,msg)
                st.success("Message saved to the selected parent's communication log.")

        elif tool=="Phone Call Script":
            sid,scholar,name,pro,subj_pr,poss_pr=assistant_scholar_context(roster,"assistant_call")
            gmap=guardian_display_map(sid)
            gid=st.selectbox("Parent / Guardian",list(gmap),format_func=lambda x:gmap[x],key="assistant_call_guardian")

            c1,c2=st.columns(2)
            subj_choice=c1.selectbox("Subject",["General","ELA","Math","Science","Social Studies","Behavior","Attendance","Other / Custom"],key="assistant_call_subject")
            subj_custom=c1.text_input("Other / custom subject",key="assistant_call_subject_other")
            subj=subj_custom.strip() if subj_choice=="Other / Custom" and subj_custom.strip() else subj_choice

            reason_choice=c2.selectbox("Reason",[
                "Positive update","Progress update","Academic concern","Behavior concern","Missing work",
                "Attendance / lateness","Assessment / test","Injury / classroom incident",
                "Conference request","Follow-up from previous contact","General update","Other / Custom"
            ],key="assistant_call_reason")
            reason_custom=c2.text_input("Other / custom reason",key="assistant_call_reason_other")
            reason=reason_custom.strip() if reason_choice=="Other / Custom" and reason_custom.strip() else reason_choice

            t1,t2=st.columns(2)
            tone_choice=t1.selectbox("Tone",[
                "Warm & conversational","Positive & celebratory","Direct & professional",
                "Concerned but supportive","Other / Custom"
            ],key="assistant_call_tone")
            tone_custom=t1.text_input("Other / custom tone",key="assistant_call_tone_other")
            tone=tone_custom.strip() if tone_choice=="Other / Custom" and tone_custom.strip() else tone_choice

            next_choice=t2.selectbox("Desired outcome / next step",[
                "Inform family only","Ask family to reinforce at home","Create a school-home plan",
                "Schedule a conference","Ask family for input","Follow up again later","Other / Custom"
            ],key="assistant_call_next_choice")
            next_custom=t2.text_input("Other / custom outcome",key="assistant_call_next_other")
            next_step=next_custom.strip() if next_choice=="Other / Custom" and next_custom.strip() else next_choice

            details=st.text_area("Specific information / key details",key="assistant_call_details")

            if st.button("Generate Phone Script",key="assistant_generate_call"):
                parent_label=gmap.get(gid,"the parent/guardian").split(" (")[0]
                partnership={
                    "Inform family only":"I wanted to make sure you had this information directly from me.",
                    "Ask family to reinforce at home":"I would appreciate your help reinforcing this skill or expectation at home.",
                    "Create a school-home plan":"I would like us to agree on a simple school-home plan so we can support the same goal consistently.",
                    "Schedule a conference":"I would like to schedule a conference so we can discuss this more fully.",
                    "Ask family for input":"I would also like to hear what you are noticing at home and any strategies that work well.",
                    "Follow up again later":"I will continue monitoring this and follow up with you again."
                }.get(next_step,f"I would like to work with you on this next step: {next_step}.")

                tone_line=f"Tone: {tone}\n\n" if tone else ""
                st.session_state["assistant_phone_script"]=f"""Opening:
Hello, may I speak with {parent_label}? This is the teacher calling about {name}. Is now a good time to speak for a few minutes?

{tone_line}Reason:
I wanted to touch base regarding {reason.lower()} in {subj}.

Specific information:
{details or "I wanted to provide you with a clear classroom update."}

Partnership / Next Step:
{partnership}

Closing:
Thank you for taking the time to speak with me. Please let me know if you have any questions or anything you would like me to know."""

            script=st.text_area("Edit / use during call",value=st.session_state.get("assistant_phone_script",""),height=360)
            outcome=st.text_area("Call notes / parent response",key="assistant_call_outcome")
            if st.button("Save Call to Communication Log",key="assistant_log_call"):
                log_communication(sid,gid,"Phone Call",subj,reason,outcome or details,script)
                st.success("Phone call saved to the selected parent's communication log.")


elif page=="Bulletin Board":
    st.markdown('<div class="page-title">Bulletin Board</div><div class="page-subtitle">Turn scholar work into display-ready feedback, standards, vocabulary, and an optional grade.</div>',unsafe_allow_html=True)
    st.caption("Upload the scholar work and rubric, then build two glows and one grow in kid-friendly language with a little rigor.")

    roster=scholars_df(selected_class or None)

    if roster.empty:
        st.info("No scholars yet. The Bulletin Board workspace is ready now; scholar names will populate automatically after you add your roster.")
        b1,b2=st.columns(2)
        b1.selectbox("Scholar",["No scholars yet"],disabled=True,key="bb_empty_scholar")
        b2.selectbox("Subject",["ELA","Math","Science","Social Studies","Writing"],key="bb_empty_subject")
        st.text_input("Display / assignment title",placeholder="Example: Character Traits Response",key="bb_empty_title")
        st.file_uploader("Upload scholar work",type=["png","jpg","jpeg","pdf"],disabled=True,key="bb_empty_work")
        st.file_uploader("Upload rubric",type=["pdf","png","jpg","jpeg","docx"],disabled=True,key="bb_empty_rubric_file")
        st.text_area("Rubric / success criteria",placeholder="Paste the rubric or success criteria here.",key="bb_empty_rubric")
        st.text_area("Task / prompt",placeholder="What were scholars asked to do?",key="bb_empty_task")
        st.text_input("Vocabulary words used / expected",placeholder="Example: trait, evidence, infer, detail",key="bb_empty_vocab")
        st.text_input("Standards",placeholder="Example: 3R3, 3W5",key="bb_empty_standards")
        g1,g2=st.columns(2)
        g1.text_area("Glow #1",placeholder="What did the scholar do well?",key="bb_empty_glow1")
        g2.text_area("Glow #2",placeholder="What else did the scholar do well?",key="bb_empty_glow2")
        st.text_area("Grow",placeholder="What is the clearest next step?",key="bb_empty_grow")
        st.number_input("Grade",min_value=0.0,max_value=100.0,value=0.0,key="bb_empty_grade")
        st.checkbox("Include this grade in the gradebook / report-card average",value=False,key="bb_empty_include")
        st.button("Generate Bulletin Board Feedback",disabled=True,key="bb_empty_generate")
    else:
        top1,top2=st.columns(2)
        sid=top1.selectbox("Scholar",list(roster.id.astype(int)),
                           format_func=lambda x:nm(roster[roster.id==x].iloc[0]),key="bb_scholar")
        subject=top2.selectbox("Subject",["ELA","Math","Science","Social Studies","Writing"],key="bb_subject")
        scholar_name=nm(roster[roster.id==sid].iloc[0])

        title=st.text_input("Display / assignment title",placeholder="Example: Character Traits Response",key="bb_title")

        f1,f2=st.columns(2)
        work=f1.file_uploader("Upload scholar work",type=["png","jpg","jpeg","pdf"],key="bb_work")
        rubric_file=f2.file_uploader("Upload rubric",type=["pdf","png","jpg","jpeg","docx"],key="bb_rubric_file")

        rubric=st.text_area(
            "Rubric / success criteria",
            placeholder="Paste the rubric language or the success criteria you want the feedback and grade based on.",
            height=130,key="bb_rubric"
        )
        task=st.text_area("Task / prompt",placeholder="What were scholars asked to do?",height=90,key="bb_task")

        v1,v2=st.columns(2)
        vocab=v1.text_input("Vocabulary words used / expected",placeholder="Example: trait, evidence, infer, detail",key="bb_vocab")
        standards=v2.text_input("Standards",placeholder="Example: 3R3, 3W5",key="bb_standards")

        st.markdown("### Two Glows + One Grow")
        st.caption("Use evidence from the work/rubric. ChapLab turns these notes into a polished, scholar-friendly comment.")
        g1,g2=st.columns(2)
        glow1=g1.text_area("Glow #1",placeholder="Example: Answered all parts of the prompt correctly.",height=95,key="bb_glow1")
        glow2=g2.text_area("Glow #2",placeholder="Example: Used strong details from the text.",height=95,key="bb_glow2")
        grow=st.text_area("Grow",placeholder="Example: Use capital letters and ending punctuation in every sentence.",height=95,key="bb_grow")

        grade1,grade2,grade3=st.columns([1,1,2])
        grade_value=grade1.number_input("Grade",min_value=0.0,max_value=100.0,value=0.0,step=1.0,key="bb_grade")
        points_possible=grade2.number_input("Points Possible",min_value=1.0,max_value=1000.0,value=100.0,step=1.0,key="bb_points")
        include_grade=grade3.checkbox(
            "Include this grade in the gradebook / report-card average",
            value=False,key="bb_include_grade"
        )

        if st.button("✨ Generate Bulletin Board Feedback",key="bb_generate"):
            parts=[f"Great job, {scholar_name}!"]
            if glow1.strip():
                parts.append(glow1.strip().rstrip(".")+"." )
            if glow2.strip():
                parts.append(glow2.strip().rstrip(".")+"." )
            if grow.strip():
                parts.append("Next time, "+grow.strip()[0].lower()+grow.strip()[1:] if len(grow.strip())>1 else "Next time, "+grow.strip())
            feedback=" ".join(parts)
            st.session_state["bb_feedback"]=feedback

        feedback=st.text_area(
            "Editable feedback for the scholar",
            value=st.session_state.get("bb_feedback",""),
            height=155,key="bb_feedback_editor"
        )

        st.markdown("### Display Information")
        d1,d2=st.columns(2)
        with d1:
            st.text_area("Task to display",value=task,height=100,key="bb_task_display")
            st.text_area("Vocabulary to display",value=vocab,height=80,key="bb_vocab_display")
        with d2:
            st.text_area("Standards to display",value=standards,height=80,key="bb_standards_display")
            st.caption("These fields are editable so you can clean up wording before printing or copying.")

        if st.button("💾 Save Bulletin Board Entry",type="primary",key="bb_save"):
            if not title.strip():
                st.warning("Add a display / assignment title first.")
            else:
                file_name=""
                file_path=""
                if work is not None:
                    safe=datetime.now().strftime("%Y%m%d_%H%M%S_")+re.sub(r'[^A-Za-z0-9._-]','_',work.name)
                    folderp=UPLOAD_DIR/f"bulletin_board/{sid}"
                    folderp.mkdir(parents=True,exist_ok=True)
                    dest=folderp/safe
                    dest.write_bytes(work.getbuffer())
                    file_name=work.name
                    file_path=str(dest)

                assignment_id=None
                if include_grade:
                    c=conn()
                    existing=c.execute(
                        """SELECT id FROM assignments
                           WHERE class_id=? AND title=? AND subject=? ORDER BY id DESC LIMIT 1""",
                        (int(selected_class) if selected_class else None,title.strip(),subject)
                    ).fetchone()
                    if existing:
                        assignment_id=int(existing["id"])
                    else:
                        cur=c.execute(
                            """INSERT INTO assignments(title,subject,category,standard_code,points_possible,assignment_date,class_id,marking_period,include_in_average)
                               VALUES (?,?,?,?,?,?,?,?,1)""",
                            (title.strip(),subject,"Classwork",standards.split(",")[0].strip() if standards else "",
                             float(points_possible),str(date.today()),int(selected_class) if selected_class else None,
                             quarter_for_date(str(date.today()),current_academic_year()) or "Quarter 1")
                        )
                        assignment_id=int(cur.lastrowid)
                    earned=(float(grade_value)/100.0)*float(points_possible)
                    c.execute("INSERT OR REPLACE INTO grades(scholar_id,assignment_id,points_earned) VALUES (?,?,?)",
                              (int(sid),assignment_id,earned))
                    c.commit(); c.close()

                c=conn()
                c.execute("""INSERT INTO bulletin_board_work(
                    scholar_id,created_at,title,subject,task_text,vocabulary,standards,rubric_text,
                    work_file_name,work_file_path,glow_1,glow_2,grow,final_feedback,
                    grade_value,points_possible,include_in_gradebook,assignment_id)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (int(sid),datetime.now().isoformat(timespec="minutes"),title.strip(),subject,
                     task,vocab,standards,rubric,file_name,file_path,glow1,glow2,grow,feedback,
                     float(grade_value),float(points_possible),1 if include_grade else 0,assignment_id))
                c.commit(); c.close()
                st.success("Bulletin Board entry saved." + (" Grade added to the gradebook." if include_grade else " Grade kept out of the report-card average."))
                st.rerun()

        st.markdown("---")
        st.markdown("### Saved Bulletin Board Work")
        c=conn()
        saved=pd.read_sql_query(
            """SELECT b.id,b.created_at,b.title,b.subject,b.grade_value,b.include_in_gradebook,
                      b.final_feedback,s.first_name||' '||s.last_name scholar
               FROM bulletin_board_work b
               LEFT JOIN scholars s ON s.id=b.scholar_id
               WHERE b.scholar_id=?
               ORDER BY b.id DESC""",c,params=[int(sid)]
        )
        c.close()
        if saved.empty:
            st.caption("No Bulletin Board entries saved for this scholar yet.")
        else:
            show=saved.copy()
            show["Counts in Gradebook"]=show["include_in_gradebook"].map(lambda x:"Yes" if x else "No")
            st.dataframe(show[["created_at","title","subject","grade_value","Counts in Gradebook","final_feedback"]],
                         hide_index=True,use_container_width=True)

elif page=="Communication Log":
    st.markdown('<div class="page-title">Communication Log</div><div class="page-subtitle">Track family contacts and follow-up notes.</div>',unsafe_allow_html=True)
    roster=scholars_df(selected_class or None)

    st.markdown("### Add communication record")
    if roster.empty:
        st.info("No scholars yet. The communication form is ready below and the Scholar/Parent fields will populate automatically after you add a roster.")

        pc1,pc2=st.columns(2)
        pc1.selectbox("Scholar",["Scholar names will appear here"],disabled=True,key="comm_preview_scholar")
        pc2.selectbox("Parent / Guardian",["Parent/guardian contacts will appear here"],disabled=True,key="comm_preview_guardian")

        comm_type=st.selectbox("Communication Type",[
            "Phone Call","Voicemail","Text / School Message","Email",
            "In-Person Conversation","Conference","Letter / Notice Sent","Other"
        ],key="comm_preview_type")
        date_text=st.text_input("Date",value=date.today().strftime("%m/%d/%Y"),placeholder="MM/DD/YYYY",key="comm_preview_date")
        subject=st.selectbox("Subject",["General","ELA","Math","Science","Social Studies","Behavior","Attendance","Health / Injury"],key="comm_preview_subject")
        reason=st.selectbox("Reason",[
            "Positive update","Homework reminder","Missing assignment","Academic concern",
            "Behavior concern","Attendance / lateness","Injury / classroom incident",
            "Conference request","Progress update","Assessment / test","Supplies / materials",
            "Follow-up from previous contact","Parent question / concern","Other"
        ],key="comm_preview_reason")
        st.text_area("Notes / More Information",placeholder="Add what happened, what was discussed, parent response, next steps, etc.",key="comm_preview_notes")
        st.button("Save Communication Record",disabled=True,key="comm_preview_save")
    else:
        scholar_id=st.selectbox(
            "Scholar",
            list(roster.id.astype(int)),
            format_func=lambda x:nm(roster[roster.id==x].iloc[0]),
            key="comm_log_scholar"
        )
        gopts=guardian_options_for_scholar(scholar_id)
        guardian_id=st.selectbox("Parent / Guardian",list(gopts),format_func=lambda x:gopts[x],key="comm_log_guardian")
        comm_type=st.selectbox("Communication Type",[
            "Phone Call","Voicemail","Text / School Message","Email",
            "In-Person Conversation","Conference","Letter / Notice Sent","Other"
        ])
        date_text=st.text_input("Date",value=date.today().strftime("%m/%d/%Y"),placeholder="MM/DD/YYYY")
        subject=st.selectbox("Subject",["General","ELA","Math","Science","Social Studies","Behavior","Attendance","Health / Injury"])
        reason=st.selectbox("Reason",[
            "Positive update","Homework reminder","Missing assignment","Academic concern",
            "Behavior concern","Attendance / lateness","Injury / classroom incident",
            "Conference request","Progress update","Assessment / test","Supplies / materials",
            "Follow-up from previous contact","Parent question / concern","Other"
        ])
        notes_text=st.text_area("Notes / More Information",placeholder="Add what happened, what was discussed, parent response, next steps, etc.")
        if st.button("Save Communication Record"):
            c=conn()
            c.execute("""INSERT INTO communications(
                scholar_id,guardian_id,created_at,communication_type,subject,reason,details,generated_text)
                VALUES (?,?,?,?,?,?,?,?)""",
                (int(scholar_id),int(guardian_id) if guardian_id else None,date_text.strip(),
                 comm_type,subject,reason,notes_text.strip(),""))
            c.commit(); c.close()
            st.success("Communication record saved.")
            st.rerun()

    st.markdown("### Saved communication records")
    c=conn()
    q="""SELECT communications.id,communications.created_at,
         scholars.first_name||' '||scholars.last_name scholar,
         COALESCE(TRIM(guardians.first_name||' '||guardians.last_name),'') guardian,
         guardians.relationship,
         communications.communication_type,communications.subject,
         communications.reason,communications.details,communications.generated_text,
         communications.scholar_id, communications.guardian_id
         FROM communications
         LEFT JOIN scholars ON scholars.id=communications.scholar_id
         LEFT JOIN guardians ON guardians.id=communications.guardian_id
         WHERE 1=1"""
    params=[]
    if selected_class:
        q+=" AND scholars.class_id=?"
        params.append(int(selected_class))
    q+=" ORDER BY communications.id DESC"
    d=pd.read_sql_query(q,c,params=params); c.close()

    if d.empty:
        st.caption("No communication records yet. Saved records will appear here automatically.")
    else:
        show=d.copy()
        show["Parent / Guardian"]=show.apply(
            lambda r: (r["guardian"] + (f" ({r['relationship']})" if r["relationship"] else "")).strip(),
            axis=1
        )
        st.dataframe(show[["id","created_at","scholar","Parent / Guardian","communication_type","subject","reason","details"]],
                     hide_index=True,use_container_width=True)

        edit_id=st.selectbox(
            "Edit or delete record",
            list(d.id.astype(int)),
            format_func=lambda x:f"#{x} — {d[d.id==x].iloc[0].scholar} — {d[d.id==x].iloc[0].reason}",
            key="comm_edit_select"
        )
        r=d[d.id==edit_id].iloc[0]
        scholar_choices=scholars_df(selected_class or None)
        sch_ids=list(scholar_choices.id.astype(int))
        current_sch=int(r.scholar_id) if pd.notna(r.scholar_id) and int(r.scholar_id) in sch_ids else sch_ids[0]

        with st.expander("✏️ Edit selected communication"):
            with st.form("edit_comm"):
                e_sch=st.selectbox("Scholar",sch_ids,format_func=lambda x:nm(scholar_choices[scholar_choices.id==x].iloc[0]),
                                   index=sch_ids.index(current_sch))
                egopts=guardian_options_for_scholar(e_sch)
                gids=list(egopts)
                current_gid=int(r.guardian_id) if pd.notna(r.guardian_id) and int(r.guardian_id) in gids else 0
                e_guard=st.selectbox("Parent / Guardian",gids,format_func=lambda x:egopts[x],index=gids.index(current_gid))

                types=["Phone Call","Voicemail","Text / School Message","Email","In-Person Conversation","Conference","Letter / Notice Sent","Other"]
                e_type=st.selectbox("Communication Type",types,index=types.index(r.communication_type) if r.communication_type in types else 0)
                e_date=st.text_input("Date",value=str(r.created_at or ""))
                subjects=["General","ELA","Math","Science","Social Studies","Behavior","Attendance","Health / Injury"]
                e_subject=st.selectbox("Subject",subjects,index=subjects.index(r.subject) if r.subject in subjects else 0)
                reasons=["Positive update","Homework reminder","Missing assignment","Academic concern","Behavior concern","Attendance / lateness",
                         "Injury / classroom incident","Conference request","Progress update","Assessment / test","Supplies / materials",
                         "Follow-up from previous contact","Parent question / concern","Other"]
                e_reason=st.selectbox("Reason",reasons,index=reasons.index(r.reason) if r.reason in reasons else 0)
                e_notes=st.text_area("Notes / More Information",value=str(r.details or ""))

                if st.form_submit_button("Save Changes"):
                    c=conn()
                    c.execute("""UPDATE communications SET scholar_id=?,guardian_id=?,created_at=?,
                                 communication_type=?,subject=?,reason=?,details=? WHERE id=?""",
                              (int(e_sch),int(e_guard) if e_guard else None,e_date.strip(),
                               e_type,e_subject,e_reason,e_notes.strip(),int(edit_id)))
                    c.commit(); c.close()
                    st.success("Communication record updated.")
                    st.rerun()

            delete_ok=st.checkbox("I want to delete this communication record",key="comm_del_check")
            if st.button("Delete Selected Communication",disabled=not delete_ok):
                delete_record("communications",edit_id)
                st.success("Communication record deleted.")
                st.rerun()


elif page=="Web & Backup":
    st.markdown(
        '<div class="page-title">Web & Backup</div>'
        '<div class="page-subtitle">Cloud connection, database backups, restore/migration, and session controls.</div>',
        unsafe_allow_html=True
    )

    # ---------------- Cloud Status ----------------
    st.markdown("### ☁️ Cloud Storage")
    if cloud_configured():
        st.success("Cloud storage is connected.")
        cfg=cloud_config()
        st.caption(
            f"Bucket: {cfg['bucket']} • Database object: {cfg['db_object']}"
        )
        if st.session_state.get("_cloud_sync_error"):
            st.warning("Last sync message: "+str(st.session_state["_cloud_sync_error"]))
    else:
        st.info(
            "ChapLab is currently running without a configured cloud-storage connection. "
            "Local database tools below are still available."
        )

    # ---------------- Database Backup ----------------
    st.markdown("---")
    st.markdown("### 💾 Database Backup")
    st.caption(
        "Download a complete copy of the current ChapLab database. "
        "This includes classes, scholars, grades, assessments, communications, newsletter records, "
        "Bulletin Board entries, settings, and other saved app data."
    )

    db_path=Path(DB)
    if db_path.exists():
        st.download_button(
            "⬇️ Download Current ChapLab Database",
            data=db_path.read_bytes(),
            file_name=f"teacher_tracker_backup_{date.today().isoformat()}.db",
            mime="application/x-sqlite3",
            use_container_width=True,
            key="download_db_backup_v4017"
        )
    else:
        st.warning("The ChapLab database file is not currently available for download.")

    if cloud_configured():
        st.markdown("### Manual Cloud Sync")
        st.caption(
            "ChapLab does not automatically block startup waiting for the cloud. "
            "Use this button when you intentionally want to push the current database to cloud storage."
        )
        if st.button("☁️ Sync Database Now",use_container_width=True,key="sync_db_now_v4017"):
            if cloud_upload_file(DB):
                st.session_state.pop("_cloud_sync_error",None)
                st.success("Database synced to private cloud storage.")
            else:
                st.error("Cloud sync did not complete. Check the Supabase settings and bucket.")

    # ---------------- Restore / Migrate ----------------
    st.markdown("---")
    st.markdown("### ♻️ Restore / Migrate Existing ChapLab Data")
    st.caption(
        "Upload a previous ChapLab SQLite database if you need to restore a backup or move existing "
        "ChapLab data into this web version."
    )

    restore=st.file_uploader(
        "Upload an existing teacher_tracker.db or backup .db file",
        type=["db","sqlite","sqlite3"],
        key="restore_chaplab_db_v4017"
    )
    confirm_restore=st.checkbox(
        "I understand this will replace the database currently being used by this ChapLab installation.",
        key="confirm_restore_db_v4017"
    )

    if st.button(
        "Restore Uploaded Database",
        disabled=(restore is None or not confirm_restore),
        use_container_width=True,
        key="restore_db_button_v4017"
    ):
        raw=bytes(restore.getbuffer())
        tmp_path=Path(tempfile.gettempdir())/"chaplab_restore_check.db"
        tmp_path.write_bytes(raw)
        try:
            check=sqlite3.connect(str(tmp_path))
            integrity=check.execute("PRAGMA integrity_check").fetchone()[0]
            tables={r[0] for r in check.execute(
                "SELECT name FROM sqlite_master WHERE type='table'"
            ).fetchall()}
            check.close()

            if str(integrity).lower()!="ok":
                raise ValueError("SQLite integrity check did not pass.")
            if "scholars" not in tables or "settings" not in tables:
                raise ValueError("This does not appear to be a ChapLab database.")

            # Create an automatic safety copy before replacing the active DB.
            if Path(DB).exists():
                safety=Path(tempfile.gettempdir())/f"chaplab_before_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
                safety.write_bytes(Path(DB).read_bytes())

            Path(DB).write_bytes(raw)

            # Run normal migrations so older ChapLab databases gain newer tables/columns.
            st.cache_resource.clear()
            st.success("Database restored. ChapLab will reload and apply any required database updates.")
            st.rerun()

        except Exception as e:
            st.error(f"Could not restore this database: {e}")

    # ---------------- Creator / Demo Safety ----------------
    if is_creator_account():
        st.markdown("---")
        st.markdown("### 🧪 Creator / Demo Data")
        demo=demo_setting()
        st.write(f"Demo Class: **{'ON' if demo['enabled'] else 'OFF'}**")
        st.caption(
            "Use Profile Settings → Creator Rollout Controls to change Demo Mode or rollout settings. "
            "Turning Demo Mode off hides demo records and does not delete real teacher data."
        )

    # ---------------- Session ----------------
    if auth_config():
        st.markdown("---")
        st.markdown("### 🔐 Session")
        st.caption(f"Signed in as: {current_author_name()}")
        if st.button("Sign Out",key="chaplab_signout_v4017"):
            st.session_state["chaplab_authenticated"]=False
            st.session_state.pop("chaplab_username",None)
            st.rerun()

    # ---------------- Safety Note ----------------
    st.markdown("---")
    st.markdown("### Important")
    st.warning(
        "Keep downloaded database backups in a secure location. ChapLab may contain student information. "
        "Use only storage and sharing services approved by your school or organization. "
        "Avoid editing the same database from multiple devices at exactly the same time."
    )


st.markdown("""
<style>
/* ==========================================================
   ChapLab v4.0.5 — Sidebar Visibility Fix ONLY
   ========================================================== */

/* Re-enable the native Streamlit sidebar that v4.0 CSS was hiding */
section[data-testid="stSidebar"]{
    display:block !important;
    visibility:visible !important;
    width:270px !important;
    min-width:270px !important;
    max-width:270px !important;
    background:linear-gradient(180deg,#123f8c,#0d2f6a) !important;
    color:white !important;
    border-right:none !important;
    box-shadow:5px 0 18px rgba(0,0,0,.12) !important;
    z-index:10000 !important;
}
section[data-testid="stSidebar"] > div{
    display:block !important;
    visibility:visible !important;
}
[data-testid="collapsedControl"]{
    display:none !important;
}

/* Sidebar text */
section[data-testid="stSidebar"] h1,
section[data-testid="stSidebar"] h2,
section[data-testid="stSidebar"] h3,
section[data-testid="stSidebar"] p,
section[data-testid="stSidebar"] small,
section[data-testid="stSidebar"] label,
section[data-testid="stSidebar"] [data-testid="stCaptionContainer"]{
    color:white !important;
}

/* Native navigation buttons */
section[data-testid="stSidebar"] [data-testid="stButton"] button{
    width:100% !important;
    border-radius:10px !important;
    min-height:44px !important;
    font-weight:700 !important;
    text-align:left !important;
}
section[data-testid="stSidebar"] [data-testid="stButton"] button[kind="secondary"]{
    background:rgba(255,255,255,.08) !important;
    color:white !important;
    border:1px solid rgba(255,255,255,.08) !important;
}
section[data-testid="stSidebar"] [data-testid="stButton"] button[kind="secondary"] *{
    color:white !important;
}
section[data-testid="stSidebar"] [data-testid="stButton"] button[kind="primary"]{
    background:linear-gradient(90deg,#7669eb,#5f65e8) !important;
    color:white !important;
    border:none !important;
}
section[data-testid="stSidebar"] [data-testid="stButton"] button[kind="primary"] *{
    color:white !important;
}

/* Main content should account for the real sidebar now.
   Remove the old fake-sidebar 300px left padding. */
.block-container{
    max-width:1450px !important;
    margin:0 auto !important;
    padding:120px 34px 52px 34px !important;
}

/* Existing fixed class bar starts after the real sidebar */
.class-bar{
    left:270px !important;
}

/* Keep the old custom sidebar hidden if any remnants exist */
.teacher-sidebar{
    display:none !important;
}

@media(max-width:900px){
    section[data-testid="stSidebar"]{
        width:235px !important;
        min-width:235px !important;
        max-width:235px !important;
    }
    .class-bar{
        left:235px !important;
    }
}
</style>
""", unsafe_allow_html=True)


st.markdown("""
<style>
section[data-testid="stSidebar"],[data-testid="collapsedControl"]{display:none!important;}
.block-container,[data-testid="stMain"] .block-container{max-width:1500px!important;width:100%!important;margin:0 auto!important;padding:34px 28px 60px!important;box-sizing:border-box!important;}
.teacher-sidebar{display:none!important}.class-bar{position:static!important;left:auto!important;right:auto!important;top:auto!important}
.teacher-hub-banner{min-height:92px;display:flex;flex-direction:column;justify-content:center;padding:8px 10px}
.teacher-hub-banner-title{font-family:"Comic Sans MS","Segoe Print",cursive;font-size:clamp(28px,3.2vw,46px);font-weight:900;color:#26344b;line-height:1.08;margin-bottom:12px}
.teacher-hub-banner-details{color:#5e6674;font-size:15px;line-height:1.65;overflow-wrap:anywhere}
.accent-page-label{display:inline-block;margin:6px 0 18px;background:#c5a8e6;color:#253044;padding:7px 18px 9px;border-radius:8px;font-family:"Comic Sans MS","Segoe Print",cursive;font-weight:800}
</style>
""",unsafe_allow_html=True)

